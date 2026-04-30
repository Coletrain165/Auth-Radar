"""
Pace Auth PDF Extractor - Standalone Desktop Tool with OCR
Extracts Patient Name, Auth #, Date Approved, Date Auth Expire, and Patient ID
from authorization PDFs (including scanned documents) and exports to Excel.

No Azure required - uses local OCR (Tesseract).
"""

import os
import sys
import re
import pathlib
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
from threading import Thread

# Dynamic year detection for OCR date correction
# This avoids hardcoding years like "2025" in OCR patterns
def get_current_year():
    """Get current year as integer."""
    return datetime.now().year

def get_likely_year_suffix():
    """
    Get the likely 2-digit year suffix for OCR correction.
    Returns current year's last 2 digits (e.g., 26 for 2026).
    """
    return str(get_current_year())[-2:]

def get_likely_full_year():
    """
    Get the likely 4-digit year for OCR correction.
    Returns current year (e.g., 2026).
    """
    return str(get_current_year())

# Current and previous year for OCR pattern matching
# Auth documents may reference current year or previous year
CURRENT_YEAR = get_current_year()
CURRENT_YEAR_STR = str(CURRENT_YEAR)
CURRENT_YEAR_SUFFIX = CURRENT_YEAR_STR[-2:]  # e.g., "26"
PREV_YEAR = CURRENT_YEAR - 1
PREV_YEAR_STR = str(PREV_YEAR)
PREV_YEAR_SUFFIX = PREV_YEAR_STR[-2:]  # e.g., "25"
import json
import csv
import io
import base64

# Gmail API
try:
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    GMAIL_API_AVAILABLE = True
except ImportError:
    GMAIL_API_AVAILABLE = False

# HTTP requests for Caspio API
try:
    import requests
    REQUESTS_AVAILABLE = True
except ImportError:
    REQUESTS_AVAILABLE = False

# PDF and Excel libraries
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import pandas as pd
    from openpyxl import Workbook
except ImportError:
    pd = None

# OCR libraries
try:
    from pdf2image import convert_from_path
    import pytesseract
    from PIL import Image, ImageEnhance, ImageFilter
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

def preprocess_image_for_ocr(img):
    """
    Preprocess image to improve OCR accuracy.
    Steps: Convert to grayscale, increase contrast, sharpen, and apply threshold.
    """
    # Convert to grayscale if not already
    if img.mode != 'L':
        img = img.convert('L')
    
    # Increase contrast
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    
    # Sharpen
    img = img.filter(ImageFilter.SHARPEN)
    
    # Apply adaptive threshold for cleaner text
    # Convert to binary (black and white) using a threshold
    threshold = 150
    img = img.point(lambda p: 255 if p > threshold else 0, mode='1')
    
    # Convert back to L mode for Tesseract
    img = img.convert('L')
    
    return img

# ML libraries
# Set USE_ML = True to re-enable the ML model if/when it's retrained with more data
USE_ML = False
try:
    from ml_extractor import MLExtractor
    ML_AVAILABLE = True and USE_ML
except ImportError:
    ML_AVAILABLE = False
    print("ML module not available. Using regex extraction only.")

# Azure Document Intelligence
try:
    from azure.ai.formrecognizer import DocumentAnalysisClient
    from azure.core.credentials import AzureKeyCredential
    AZURE_AVAILABLE = True
except ImportError:
    AZURE_AVAILABLE = False
    print("Azure Document Intelligence not available. Using local OCR.")

# --- Centralized configuration (credentials, paths, field list) ---
# Loaded from config.py which reads .env for secrets
from config import (
    AZURE_ENDPOINT, AZURE_KEY, AZURE_MODEL_ID,
    CASPIO_ACCOUNT_ID, CASPIO_CLIENT_ID, CASPIO_CLIENT_SECRET, CASPIO_TABLE_NAME,
    FIELDS, APP_DIR, POPPLER_PATH, PATIENT_NAMES_FILE, PDF_PASSWORD,
)


class PatientNameMatcher:
    """Matches extracted patient names to database names using fuzzy matching."""
    
    def __init__(self, names_file=None):
        self.names_file = names_file or PATIENT_NAMES_FILE
        self.patients = []  # List of (last, first, full_name) tuples
        self.load_names()
    
    def load_names(self):
        """Load patient names from JSON file."""
        self.patients = []
        if not self.names_file.exists():
            return
        
        try:
            with open(self.names_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            for name in data.get("patients", []):
                if "," in name:
                    parts = name.split(",", 1)
                    last = parts[0].strip()
                    first = parts[1].strip()
                else:
                    parts = name.split()
                    last = parts[-1] if parts else ""
                    first = parts[0] if len(parts) > 1 else ""
                
                self.patients.append((last, first, name))
        except Exception as e:
            print(f"Error loading patient names: {e}")
    
    def reload_names(self):
        """Reload patient names from file (useful after syncing from Caspio)."""
        self.load_names()
    
    def normalize(self, s):
        """Normalize a string for comparison."""
        if not s:
            return ""
        # Convert to lowercase, remove punctuation except hyphens
        s = s.lower().strip()
        s = re.sub(r"[',.\s]+", " ", s)  # Replace punctuation with space
        s = re.sub(r'\s+', ' ', s)  # Normalize whitespace
        return s.strip()
    
    def get_first_name_only(self, first_name):
        """Extract just the first name, removing middle names/initials."""
        if not first_name:
            return ""
        parts = first_name.split()
        return parts[0] if parts else ""
    
    def clean_name(self, extracted_name):
        """
        Clean an extracted name by removing middle names/initials.
        Returns (last_name, first_name_only, clean_full_name).
        """
        if not extracted_name:
            return "", "", ""
        
        # Parse the name
        if "," in extracted_name:
            parts = extracted_name.split(",", 1)
            last_name = parts[0].strip()
            first_part = parts[1].strip() if len(parts) > 1 else ""
        else:
            # Assume "First Middle Last" format
            parts = extracted_name.split()
            if len(parts) >= 2:
                # Last word is last name, first word is first name
                last_name = parts[-1]
                first_part = parts[0]
            elif len(parts) == 1:
                last_name = parts[0]
                first_part = ""
            else:
                return "", "", ""
        
        # Clean last name - remove extra spaces, capitalize properly
        last_name = self.clean_last_name(last_name)
        
        # Get first name only (no middle names/initials)
        first_name = self.get_first_name_only(first_part)
        
        # Capitalize first name properly
        first_name = self.capitalize_name(first_name)
        
        # Build clean full name
        if last_name and first_name:
            clean_full = f"{last_name}, {first_name}"
        elif last_name:
            clean_full = last_name
        else:
            clean_full = first_name
        
        return last_name, first_name, clean_full
    
    def clean_last_name(self, name):
        """Clean and properly capitalize a last name."""
        if not name:
            return ""
        name = name.strip()
        # Handle compound names with spaces
        if ' ' in name:
            # Keep compound names together but capitalize each part
            parts = name.split()
            return ' '.join(self.capitalize_name(p) for p in parts)
        return self.capitalize_name(name)
    
    def capitalize_name(self, name):
        """Properly capitalize a name, handling special cases."""
        if not name:
            return ""
        name = name.strip()
        
        # Handle names that are already properly formatted (mixed case)
        # e.g., "McDonald", "DeLaCruz"
        if any(c.isupper() for c in name[1:]):
            # Already has internal capitals, likely correct
            return name[0].upper() + name[1:]
        
        # Simple capitalization
        return name.capitalize()
    
    def similarity_score(self, extracted_last, extracted_first, db_last, db_first):
        """
        Calculate similarity score between extracted name and database name.
        Returns a score from 0.0 to 1.0.
        """
        # Normalize all names
        ext_last = self.normalize(extracted_last)
        ext_first = self.normalize(extracted_first)
        db_last_norm = self.normalize(db_last)
        db_first_norm = self.normalize(db_first)
        
        # Get first name only (no middle names) for comparison
        ext_first_only = self.get_first_name_only(ext_first)
        db_first_only = self.get_first_name_only(db_first_norm)
        
        score = 0.0
        
        # Last name matching (worth 60% of score)
        if ext_last == db_last_norm:
            score += 0.6
        elif ext_last in db_last_norm or db_last_norm in ext_last:
            # Handle compound names like "DeLaCruz" matching "De La Cruz"
            score += 0.5
        elif self.normalize(ext_last.replace(" ", "")) == self.normalize(db_last_norm.replace(" ", "")):
            # Match if spaces are removed (e.g., "De La Cruz" == "DeLaCruz")
            score += 0.55
        else:
            # Check for partial match (typos, OCR errors)
            if len(ext_last) >= 3 and len(db_last_norm) >= 3:
                if ext_last[:3] == db_last_norm[:3]:
                    score += 0.3
        
        # First name matching (worth 40% of score)
        if ext_first_only == db_first_only:
            score += 0.4
        elif ext_first_only and db_first_only:
            if ext_first_only in db_first_only or db_first_only in ext_first_only:
                score += 0.35
            elif len(ext_first_only) >= 3 and len(db_first_only) >= 3:
                if ext_first_only[:3] == db_first_only[:3]:
                    score += 0.25
        
        return score
    
    def find_match(self, extracted_name, threshold=0.7):
        """
        Find the best matching patient name from the database.
        
        Args:
            extracted_name: The name extracted from the PDF
            threshold: Minimum similarity score to consider a match (0.0-1.0)
        
        Returns:
            (matched_last, matched_first, full_db_name, score) or (None, None, None, 0)
        """
        if not self.patients or not extracted_name:
            return None, None, None, 0
        
        # Parse extracted name
        if "," in extracted_name:
            parts = extracted_name.split(",", 1)
            ext_last = parts[0].strip()
            ext_first = parts[1].strip() if len(parts) > 1 else ""
        else:
            parts = extracted_name.split()
            if len(parts) >= 2:
                ext_first = parts[0]
                ext_last = " ".join(parts[1:])  # Handle multi-word last names
            elif len(parts) == 1:
                ext_last = parts[0]
                ext_first = ""
            else:
                return None, None, None, 0
        
        best_match = None
        best_score = 0
        
        for db_last, db_first, full_name in self.patients:
            score = self.similarity_score(ext_last, ext_first, db_last, db_first)
            
            if score > best_score:
                best_score = score
                best_match = (db_last, db_first, full_name)
        
        if best_match and best_score >= threshold:
            return best_match[0], best_match[1], best_match[2], best_score
        
        return None, None, None, best_score
    
    def match_and_format(self, extracted_name, threshold=0.7):
        """
        Match extracted name and return formatted database name.
        
        Returns:
            (last_name, first_name, matched_full_name, confidence) 
            If no match found, returns parsed extracted name with confidence=0
        """
        matched_last, matched_first, full_name, score = self.find_match(extracted_name, threshold)
        
        if matched_last:
            return matched_last, matched_first, full_name, score
        
        # No match found - return parsed extracted name
        if "," in extracted_name:
            parts = extracted_name.split(",", 1)
            last = parts[0].strip()
            first = parts[1].strip().split()[0] if len(parts) > 1 else ""
        else:
            parts = extracted_name.split()
            last = parts[-1] if parts else ""
            first = parts[0] if len(parts) > 1 else ""
        
        return last, first, extracted_name, 0


class CaspioAPI:
    """Handles communication with Caspio REST API."""
    
    def __init__(self, account_id=None, client_id=None, client_secret=None):
        self.account_id = account_id or CASPIO_ACCOUNT_ID
        self.client_id = client_id or CASPIO_CLIENT_ID
        self.client_secret = client_secret or CASPIO_CLIENT_SECRET
        self.base_url = f"https://{self.account_id}.caspio.com/rest/v2"
        self.token_url = f"https://{self.account_id}.caspio.com/oauth/token"
        self.access_token = None
        self.token_expiry = None
    
    def get_access_token(self):
        """Get OAuth2 access token from Caspio."""
        if not REQUESTS_AVAILABLE:
            raise ImportError("requests library is required. Run: pip install requests")
        
        # Check if we have a valid cached token
        if self.access_token and self.token_expiry:
            if datetime.now() < self.token_expiry:
                return self.access_token
        
        # Request new token
        data = {
            "grant_type": "client_credentials",
            "client_id": self.client_id,
            "client_secret": self.client_secret
        }
        
        response = requests.post(self.token_url, data=data)
        
        if response.status_code != 200:
            raise Exception(f"Failed to get Caspio token: {response.status_code} - {response.text}")
        
        token_data = response.json()
        self.access_token = token_data["access_token"]
        # Token typically expires in 24 hours, we'll refresh after 23 hours
        expires_in = token_data.get("expires_in", 86400)
        self.token_expiry = datetime.now() + timedelta(seconds=expires_in - 3600)
        
        return self.access_token
    
    def get_table_schema(self, table_name):
        """Get the field names and types for a Caspio table."""
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        url = f"{self.base_url}/tables/{table_name}/fields"
        response = requests.get(url, headers=headers)
        
        if response.status_code != 200:
            raise Exception(f"Failed to get table schema: {response.status_code} - {response.text}")
        
        data = response.json()
        # Return list of field info including read-only status
        fields = []
        for field in data.get("Result", []):
            field_type = field.get("Type", "")
            # Fields that are read-only in Caspio:
            # - AutoNumber (auto-increment IDs)
            # - Formula fields
            # - Timestamp fields (auto-updated)
            # - Fields starting with "PK_" typically
            is_readonly = (
                field_type in ("AutoNumber", "AUTONUMBER", "Formula", "FORMULA", "Timestamp", "TIMESTAMP") or
                field.get("Name", "").startswith("PK_") or
                field.get("IsAutoGenerated", False) or
                field.get("ReadOnly", False)
            )
            fields.append({
                "name": field.get("Name"),
                "type": field_type,
                "unique": field.get("UniqueValue", False),
                "required": not field.get("AllowNull", True),
                "readonly": is_readonly
            })
        return fields
    
    def get_writable_fields(self, table_name):
        """Get list of field names that can be written to."""
        fields = self.get_table_schema(table_name)
        return [f["name"] for f in fields if not f.get("readonly", False)]
    
    def insert_records(self, table_name, records):
        """Insert multiple records into a Caspio table."""
        import math
        
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        url = f"{self.base_url}/tables/{table_name}/records"
        
        results = {"success": 0, "failed": 0, "errors": []}
        
        # Get list of writable fields to filter out read-only fields
        try:
            writable_fields = self.get_writable_fields(table_name)
        except Exception as e:
            # If we can't get schema, proceed without filtering (will fail if read-only fields exist)
            writable_fields = None
        
        # Helper to clean NaN/NaT/non-JSON-compliant values
        def clean_value(val):
            """Convert any non-JSON-compliant value to empty string."""
            # None check
            if val is None:
                return ""
            
            # Float NaN check (NaN != NaN is True)
            if isinstance(val, float):
                if val != val or math.isnan(val) or math.isinf(val):
                    return ""
                return val
            
            # pandas NaT (Not a Time) and NaN check
            if pd:
                try:
                    if pd.isna(val):
                        return ""
                    # Handle Timestamp objects - convert to string
                    if hasattr(val, 'strftime'):
                        return val.strftime("%m/%d/%Y")
                except (TypeError, ValueError):
                    pass
            
            # numpy types check
            try:
                import numpy as np
                if isinstance(val, (np.floating, np.integer)):
                    if np.isnan(val) or np.isinf(val):
                        return ""
                    return float(val) if isinstance(val, np.floating) else int(val)
            except (ImportError, TypeError, ValueError):
                pass
            
            # String "nan" or "NaN" check
            if isinstance(val, str) and val.lower() in ('nan', 'nat', 'none', 'null'):
                return ""
            
            return val
        
        def clean_record(rec):
            """Clean all values in a record."""
            return {key: clean_value(val) for key, val in rec.items()}
        
        def filter_writable(rec, writable):
            """Filter record to only include writable fields."""
            if writable is None:
                return rec
            return {key: val for key, val in rec.items() if key in writable}
        
        for i, record in enumerate(records):
            # Clean the record to remove NaN/NaT values
            record = clean_record(record)
            # Filter out read-only fields
            record = filter_writable(record, writable_fields)
            try:
                response = requests.post(url, headers=headers, json=record)
                
                if response.status_code in (200, 201):
                    results["success"] += 1
                else:
                    results["failed"] += 1
                    results["errors"].append(f"Record {i+1}: {response.status_code} - {response.text}")
            except Exception as e:
                results["failed"] += 1
                results["errors"].append(f"Record {i+1}: {str(e)}")
        
        return results
    
    def fetch_patient_names(self, table_name="a_Patient"):
        """
        Fetch patient names from Caspio patient table.
        Returns list of dicts with last_name, first_name, full_name.
        """
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # Fetch records with pagination
        all_patients = []
        page_size = 1000
        page = 1
        
        while True:
            url = f"{self.base_url}/tables/{table_name}/records?q.pageSize={page_size}&q.pageNumber={page}"
            url += "&q.select=Box_2__Patient_Last_Name,Box_2__Patient_First_Name,Concantenated_Patient_Name"
            url += "&q.where=Unique_Payer_Identifier%3D'Innermark%20%3A%20WayStar%20(98481)'"
            
            response = requests.get(url, headers=headers)
            
            if response.status_code != 200:
                raise Exception(f"Failed to fetch patients: {response.status_code} - {response.text}")
            
            data = response.json()
            records = data.get("Result", [])
            
            for record in records:
                last_name = record.get("Box_2__Patient_Last_Name", "") or ""
                first_name = record.get("Box_2__Patient_First_Name", "") or ""
                full_name = record.get("Concantenated_Patient_Name", "") or ""
                
                # Skip empty records
                if not last_name and not first_name:
                    continue
                
                all_patients.append({
                    "last_name": last_name.strip(),
                    "first_name": first_name.strip(),
                    "full_name": full_name.strip() if full_name else f"{last_name}, {first_name}".strip(", ")
                })
            
            # Check if more pages
            if len(records) < page_size:
                break
            page += 1
        
        return all_patients
    
    def sync_patient_names_to_file(self, output_file=None):
        """
        Fetch patient names from Caspio and save to local JSON file.
        Returns the number of patients synced.
        """
        output_file = output_file or PATIENT_NAMES_FILE
        
        patients = self.fetch_patient_names()
        
        # Format for JSON file (Last, First format)
        patient_list = []
        for p in patients:
            if p["last_name"] and p["first_name"]:
                patient_list.append(f"{p['last_name']}, {p['first_name']}")
            elif p["last_name"]:
                patient_list.append(p["last_name"])
        
        # Remove duplicates and sort
        patient_list = sorted(set(patient_list))
        
        data = {
            "format": "Last, First",
            "source": "Caspio a_Patient table",
            "synced_at": datetime.now().isoformat(),
            "patients": patient_list
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2)
        
        return len(patient_list)
    
    def search_records(self, table_name, search_term, search_fields=None, max_results=100):
        """
        Search for records in a Caspio table.
        
        Args:
            table_name: Name of the Caspio table (e.g., 'a_Authorizations', 'a_Patient')
            search_term: The search term to look for
            search_fields: List of field names to search in (if None, searches common name fields)
            max_results: Maximum number of results to return
        
        Returns:
            List of matching records as dictionaries
        """
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # URL encode the search term
        import urllib.parse
        encoded_term = urllib.parse.quote(search_term)
        
        # Build WHERE clause for searching
        if search_fields:
            # Search in specified fields
            conditions = [f"{field} LIKE '%{search_term}%'" for field in search_fields]
            where_clause = " OR ".join(conditions)
        else:
            # Default search fields based on table
            if table_name == "a_Patient":
                where_clause = f"Box_2__Patient_Last_Name LIKE '%{search_term}%' OR Box_2__Patient_First_Name LIKE '%{search_term}%' OR Concantenated_Patient_Name LIKE '%{search_term}%'"
            elif table_name == "a_Authorizations":
                where_clause = f"Last_Name LIKE '%{search_term}%' OR First_Name LIKE '%{search_term}%' OR Auth_Number LIKE '%{search_term}%' OR Patient_ID LIKE '%{search_term}%'"
            else:
                # Generic search - will likely fail but try anyway
                where_clause = f"Name LIKE '%{search_term}%'"
        
        # URL encode the WHERE clause
        encoded_where = urllib.parse.quote(where_clause)
        
        url = f"{self.base_url}/tables/{table_name}/records?q.pageSize={max_results}&q.where={encoded_where}"
        
        try:
            response = requests.get(url, headers=headers)
            
            if response.status_code != 200:
                raise Exception(f"Search failed: {response.status_code} - {response.text}")
            
            data = response.json()
            return data.get("Result", [])
        except Exception as e:
            print(f"Search error: {e}")
            return []
    
    def get_all_records(self, table_name, select_fields=None, max_results=1000):
        """
        Fetch all records from a table (with optional field selection).
        
        Args:
            table_name: Name of the Caspio table
            select_fields: List of field names to return (None for all)
            max_results: Maximum total results
        
        Returns:
            List of records as dictionaries
        """
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        all_records = []
        page_size = min(1000, max_results)
        page = 1
        
        while len(all_records) < max_results:
            url = f"{self.base_url}/tables/{table_name}/records?q.pageSize={page_size}&q.pageNumber={page}"
            
            if select_fields:
                url += f"&q.select={','.join(select_fields)}"
            
            response = requests.get(url, headers=headers)
            
            if response.status_code != 200:
                raise Exception(f"Fetch failed: {response.status_code} - {response.text}")
            
            data = response.json()
            records = data.get("Result", [])
            
            all_records.extend(records)
            
            if len(records) < page_size:
                break
            page += 1
        
        return all_records[:max_results]

    def update_record(self, table_name, pk_field, pk_value, updates):
        """
        Update a record in a Caspio table.
        
        Args:
            table_name: Name of the Caspio table
            pk_field: Name of the primary key field (e.g., 'PK_ID')
            pk_value: Value of the primary key to identify the record
            updates: Dictionary of field names and their new values
        
        Returns:
            Dict with success status and any error message
        """
        import urllib.parse
        import math
        
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # Get writable fields to filter out read-only ones
        try:
            writable_fields = self.get_writable_fields(table_name)
            # Filter updates to only writable fields
            updates = {k: v for k, v in updates.items() if k in writable_fields}
        except Exception:
            pass  # Proceed without filtering
        
        # Clean NaN/None values
        def clean_value(val):
            if val is None:
                return ""
            if isinstance(val, float) and (val != val or math.isnan(val) or math.isinf(val)):
                return ""
            if isinstance(val, str) and val.lower() in ('nan', 'nat', 'none', 'null'):
                return ""
            return val
        
        updates = {k: clean_value(v) for k, v in updates.items()}
        
        # Build WHERE clause
        where_clause = f"{pk_field}='{pk_value}'"
        encoded_where = urllib.parse.quote(where_clause)
        
        url = f"{self.base_url}/tables/{table_name}/records?q.where={encoded_where}"
        
        try:
            response = requests.put(url, headers=headers, json=updates)
            
            if response.status_code in (200, 201, 204):
                return {"success": True, "message": "Record updated successfully"}
            else:
                return {"success": False, "message": f"Update failed: {response.status_code} - {response.text}"}
        except Exception as e:
            return {"success": False, "message": f"Update error: {str(e)}"}
    
    def delete_record(self, table_name, pk_field, pk_value):
        """
        Delete a record from a Caspio table.
        
        Args:
            table_name: Name of the Caspio table
            pk_field: Name of the primary key field
            pk_value: Value of the primary key to identify the record
        
        Returns:
            Dict with success status and any error message
        """
        import urllib.parse
        
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # Build WHERE clause
        where_clause = f"{pk_field}='{pk_value}'"
        encoded_where = urllib.parse.quote(where_clause)
        
        url = f"{self.base_url}/tables/{table_name}/records?q.where={encoded_where}"
        
        try:
            response = requests.delete(url, headers=headers)
            
            if response.status_code in (200, 204):
                return {"success": True, "message": "Record deleted successfully"}
            else:
                return {"success": False, "message": f"Delete failed: {response.status_code} - {response.text}"}
        except Exception as e:
            return {"success": False, "message": f"Delete error: {str(e)}"}
    
    def search_with_operator(self, table_name, field, operator, value, select_fields=None, max_results=100):
        """
        Search for records using a specific operator (like Caspio criteria builder).
        
        Args:
            table_name: Name of the Caspio table
            field: Field name to search on
            operator: Operator type - 'equals', 'contains', 'starts_with', 'ends_with', 
                      'not_equal', 'is_empty', 'is_not_empty', 'greater_than', 'less_than'
            value: The value to compare against
            select_fields: Optional list of fields to return
            max_results: Maximum number of results
        
        Returns:
            List of matching records
        """
        import urllib.parse
        
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        # Build WHERE clause based on operator
        if operator == 'equals':
            where_clause = f"{field}='{value}'"
        elif operator == 'not_equal':
            where_clause = f"{field}<>'{value}'"
        elif operator == 'contains':
            where_clause = f"{field} LIKE '%{value}%'"
        elif operator == 'starts_with':
            where_clause = f"{field} LIKE '{value}%'"
        elif operator == 'ends_with':
            where_clause = f"{field} LIKE '%{value}'"
        elif operator == 'is_empty':
            where_clause = f"({field} IS NULL OR {field}='')"
        elif operator == 'is_not_empty':
            where_clause = f"({field} IS NOT NULL AND {field}<>'')"
        elif operator == 'greater_than':
            where_clause = f"{field}>'{value}'"
        elif operator == 'less_than':
            where_clause = f"{field}<'{value}'"
        else:
            # Default to contains
            where_clause = f"{field} LIKE '%{value}%'"
        
        encoded_where = urllib.parse.quote(where_clause)
        
        url = f"{self.base_url}/tables/{table_name}/records?q.pageSize={max_results}&q.where={encoded_where}"
        
        if select_fields:
            url += f"&q.select={','.join(select_fields)}"
        
        try:
            response = requests.get(url, headers=headers)
            
            if response.status_code != 200:
                raise Exception(f"Search failed: {response.status_code} - {response.text}")
            
            data = response.json()
            return data.get("Result", [])
        except Exception as e:
            print(f"Search error: {e}")
            return []
    
    def insert_single_record(self, table_name, record):
        """
        Insert a single record into a Caspio table.
        
        Args:
            table_name: Name of the Caspio table
            record: Dictionary of field names and values
        
        Returns:
            Dict with success status, any error message, and the created record
        """
        import math
        
        token = self.get_access_token()
        
        headers = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
        
        url = f"{self.base_url}/tables/{table_name}/records"
        
        # Get writable fields
        try:
            writable_fields = self.get_writable_fields(table_name)
            record = {k: v for k, v in record.items() if k in writable_fields}
        except Exception:
            pass
        
        # Clean values
        def clean_value(val):
            if val is None:
                return ""
            if isinstance(val, float) and (val != val or math.isnan(val) or math.isinf(val)):
                return ""
            if isinstance(val, str) and val.lower() in ('nan', 'nat', 'none', 'null'):
                return ""
            return val
        
        record = {k: clean_value(v) for k, v in record.items()}
        
        try:
            response = requests.post(url, headers=headers, json=record)
            
            if response.status_code in (200, 201):
                result = response.json()
                return {"success": True, "message": "Record created successfully", "record": result.get("Result", {})}
            else:
                return {"success": False, "message": f"Insert failed: {response.status_code} - {response.text}"}
        except Exception as e:
            return {"success": False, "message": f"Insert error: {str(e)}"}


# Import timedelta for token expiry calculation
from datetime import timedelta

# Tesseract path (Windows default)
TESSERACT_PATHS = [
    r"C:\Program Files\Tesseract-OCR\tesseract.exe",
    r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
    r"C:\Users\{}\AppData\Local\Programs\Tesseract-OCR\tesseract.exe".format(os.environ.get("USERNAME", "")),
]


def find_tesseract():
    """Find Tesseract executable."""
    for path in TESSERACT_PATHS:
        if os.path.exists(path):
            return path
    return None


class PDFExtractor:
    """Extracts authorization data from PDFs using OCR and pattern matching."""
    
    # Patterns tuned for PACE Treatment Authorization Form OCR output
    # PSM 3 format: "Auth #: Date Approved: Date Auth. Expire: | | 25357360 fio/6/2025 10/16/2025"
    # PSM 11 format (preferred): values on separate lines, auth# may have leading '['
    PATTERNS = {
        # Patient Name - from "Participant's Name:" field
        "Patient Name": [
            r"Participant'?s?\s*Name[:\s]*\n?\s*([A-Z][A-Za-z]+(?:[,\s]+[A-Z][A-Za-z]+)+)",
            r"Member'?s?\s*Name[:\s]*\n?\s*([A-Z][A-Za-z]+(?:[,\s]+[A-Z][A-Za-z]+)+)",
            r"Patient'?s?\s*Name[:\s]*\n?\s*([A-Z][A-Za-z]+(?:[,\s]+[A-Z][A-Za-z]+)+)",
        ],
        # Auth Number - PSM 11: standalone line with optional '[' prefix; PSM 3: after '| |'
        "Auth #": [
            r"[\[\|](\d{7,10})\b",  # PSM 11: '[23924378' or '|23924378'
            r"(?m)^[\[\|]?(\d{7,10})\s*$",  # PSM 11: standalone line of digits
            r"\|\s*\|\s*(\d{7,10})",  # PSM 3: after double pipe "| | 25357360"
            r"Expire:?\s*\|?\s*\|?\s*(\d{7,10})",  # After Expire: | | 25357360
            r"Auth\s*#.*?\|\s*\|\s*(\d{7,10})",  # Full pattern with Auth #
            r"(\d{7,10})\s+\d{1,2}[/\-]",  # 7-10 digit number followed by date
            r"(\d{7,10})\s+[fF][iIlL1]",  # Auth number followed by garbled 'fi' date start
            r"Auth\s*#:?\s*(\d{7,10})",  # Simple "Auth #: 12345678" pattern
        ],
        # Date Approved - PSM 11: first clean date on its own line after auth#; PSM 3: inline
        "Date Approved": [
            r"[\[\|]?\d{7,10}\s*\n\s*(\d{1,2}/\d{1,2}/\d{4})",  # PSM 11: date on line after auth#
            r"\d{7,10}\s+(\d{1,2}/\d{1,2}/\d{4})",  # Clean date immediately after auth#
            r"\d{7,10}\s+(\S+/\d{1,2}/\d{4})",  # Any non-space chars with date format after auth#
            r"\d{7,10}\s+(\S+\s*/\d{1,2}/\d{4})",  # Allow space before first slash
            r"\d{8}\s+(.+?)\s+\d{1,2}/\d{1,2}/\d{4}",  # Capture everything between auth# and second date
        ],
        # Date Auth Expire - PSM 11: second date line after auth#; PSM 3: third token after auth#
        "Date Auth Expire": [
            r"[\[\|]?\d{7,10}\s*\n\s*\d{1,2}/\d{1,2}/\d{4}\s*\n\s*(\d{1,2}/\d{1,2}/\d{4})",  # PSM 11: second date after auth#
            r"\d{7,10}\s+\S+\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",  # PSM 3: second date after auth#
            r"(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})\s*\|?\s*UM",  # Date before "UM Decision"
            r"[/\-]\d{4}\s+(\d{1,2}[/\-]\d{1,2}[/\-]\d{4})",  # Date after another date
        ],
        # Patient ID - from "Participant ID:" field (handles newlines between label and value)
        "Patient ID": [
            r"Participant\s*ID:?\s*\n\s*(\d{6,12})",  # Label on one line, ID on next
            r"Participant\s*ID:?\s*\|?\s*(\d{6,12})",  # Inline format
            r"Participant\s*ID:?\s+(\d{6,12})",  # Space-separated
            r"Member\s*ID:?\s*\n?\s*(\d{6,12})",  # Member ID with optional newline
            r"Member\s*ID:?\s*\|?\s*(\d{6,12})",
            r"Policy\s*#:?\s*\n?\s*(\d{9,12})",  # Policy # with optional newline (common on referral pages)
            r"Policy\s*Number:?\s*(\d{9,12})",
            r"Masterld:?\s*(\d{9,12})",  # MasterID from face sheets
        ],
    }
    
    def __init__(self):
        self.results = []
        self.tesseract_path = find_tesseract()
        if self.tesseract_path and OCR_AVAILABLE:
            pytesseract.pytesseract.tesseract_cmd = self.tesseract_path
        
        # Initialize Azure Document Intelligence client
        self.azure_client = None
        self.use_azure = False
        if AZURE_AVAILABLE and AZURE_KEY and AZURE_ENDPOINT:
            try:
                self.azure_client = DocumentAnalysisClient(
                    AZURE_ENDPOINT, 
                    AzureKeyCredential(AZURE_KEY)
                )
                self.use_azure = True
                print(f"Azure Document Intelligence enabled (model: {AZURE_MODEL_ID}).")
            except Exception as e:
                print(f"Could not initialize Azure client: {e}. Using local OCR.")
                self.use_azure = False
        
        # Initialize ML extractor if available
        self.ml_extractor = None
        self.use_ml = False
        if ML_AVAILABLE:
            model_path = APP_DIR / "auth_form_ner_model"
            if model_path.exists():
                try:
                    self.ml_extractor = MLExtractor(model_path=str(model_path))
                    self.use_ml = True
                    print("ML model loaded. Using hybrid extraction (ML + Regex).")
                except Exception as e:
                    print(f"Could not load ML model: {e}. Using regex only.")
                    self.use_ml = False
            else:
                print("ML model not found. Using regex extraction only.")
                print(f"To use ML, train a model with: python ml_trainer.py --sample --output {model_path}")
    
    def extract_with_azure(self, pdf_path):
        """
        Extract fields using Azure Document Intelligence custom model.
        Returns dict with extracted fields or None if extraction fails.
        """
        if not self.use_azure or not self.azure_client:
            return None
        
        try:
            with open(pdf_path, "rb") as f:
                poller = self.azure_client.begin_analyze_document(AZURE_MODEL_ID, f)
                result = poller.result()
            
            extracted = {}
            for doc in result.documents:
                for name, field in doc.fields.items():
                    if field.value:
                        # Map Azure field names to our field names
                        field_name = name.replace("_", " ")
                        # Normalize field names
                        if "patient name" in field_name.lower() or field_name == "Patient Name":
                            extracted["Patient Name"] = field.value
                            extracted["Patient Name_confidence"] = field.confidence
                            extracted["Patient Name_method"] = "azure"
                        elif "auth" in field_name.lower() and "#" in field_name:
                            extracted["Auth #"] = str(field.value)
                            extracted["Auth #_confidence"] = field.confidence
                            extracted["Auth #_method"] = "azure"
                        elif "date approved" in field_name.lower():
                            # Handle date value - could be string or date object
                            if hasattr(field.value, 'strftime'):
                                extracted["Date Approved"] = field.value.strftime("%m/%d/%Y")
                            else:
                                extracted["Date Approved"] = str(field.value)
                            extracted["Date Approved_confidence"] = field.confidence
                            extracted["Date Approved_method"] = "azure"
                        elif "date auth expire" in field_name.lower() or "expire" in field_name.lower():
                            if hasattr(field.value, 'strftime'):
                                extracted["Date Auth Expire"] = field.value.strftime("%m/%d/%Y")
                            else:
                                extracted["Date Auth Expire"] = str(field.value)
                            extracted["Date Auth Expire_confidence"] = field.confidence
                            extracted["Date Auth Expire_method"] = "azure"
                        elif "patient id" in field_name.lower() or "participant id" in field_name.lower():
                            extracted["Patient ID"] = str(field.value)
                            extracted["Patient ID_confidence"] = field.confidence
                            extracted["Patient ID_method"] = "azure"
            
            return extracted if extracted else None
            
        except Exception as e:
            print(f"Azure extraction error: {e}")
            return None
    
    def find_auth_page(self, page_texts):
        """Find the page containing the TREATMENT AUTHORIZATION FORM."""
        # The auth form page MUST have "TREATMENT AUTHORIZATION FORM" header
        # This is the only page we should extract Auth #, dates, name, and ID from
        
        for i, text in enumerate(page_texts):
            if "TREATMENT AUTHORIZATION FORM" in text.upper():
                # Found the auth form page - return ONLY this page
                return page_texts[i], page_texts[i], i + 1
        
        # Fallback: look for key fields if header not found
        best_page = None
        best_score = 0
        auth_keywords = ["Auth #", "Date Approved", "Date Auth. Expire", "Participant ID", "MEMBER (Participant)"]
        
        for i, text in enumerate(page_texts):
            score = sum(1 for kw in auth_keywords if kw.upper() in text.upper())
            if score > best_score:
                best_score = score
                best_page = i
        
        if best_page is not None and best_score >= 3:
            return page_texts[best_page], page_texts[best_page], best_page + 1
        
        # No auth form found - return all text combined
        all_text = "\n".join(page_texts)
        return all_text, all_text, 0
        
    def extract_text_from_pdf(self, pdf_path):
        """Extract text from PDF, using OCR if needed. Returns text from auth form page."""
        page_texts = []
        method = "none"
        
        # Try pdfplumber first (for native PDFs)
        if pdfplumber:
            try:
                with pdfplumber.open(pdf_path, password=PDF_PASSWORD) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text() or ""
                        if page_text.strip():
                            page_texts.append(page_text)
                if page_texts:
                    method = "text"
            except Exception as e:
                print(f"pdfplumber error: {e}")
        
        # If no text found, try PyPDF2
        if not page_texts and PyPDF2:
            try:
                with open(pdf_path, 'rb') as f:
                    reader = PyPDF2.PdfReader(f)
                    # Decrypt if encrypted
                    if reader.is_encrypted:
                        reader.decrypt(PDF_PASSWORD)
                    for page in reader.pages:
                        page_text = page.extract_text() or ""
                        if page_text.strip():
                            page_texts.append(page_text)
                if page_texts:
                    method = "text"
            except Exception as e:
                print(f"PyPDF2 error: {e}")
        
        # Use OCR for scanned PDFs (no text extracted)
        if not page_texts and OCR_AVAILABLE and self.tesseract_path:
            try:
                poppler_path = str(POPPLER_PATH) if POPPLER_PATH.exists() else None
                # 400 DPI significantly improves accuracy for form field values
                images = convert_from_path(pdf_path, poppler_path=poppler_path, dpi=400, userpw=PDF_PASSWORD)
                
                # PSM 3 = fully automatic page segmentation (good for general text/page structure)
                # PSM 11 = sparse text, no OSD (best for isolated form field values like auth# and dates)
                # OEM 3 = default, uses LSTM if available
                config_psm3 = '--psm 3 --oem 3 -c tessedit_do_invert=0'
                config_psm11 = '--psm 11 --oem 3 -c tessedit_do_invert=0'
                
                for img in images:
                    # PSM 3 on raw image: captures general text and page structure
                    page_text = pytesseract.image_to_string(img, config=config_psm3)
                    
                    # PSM 11 on raw image: much better at reading isolated values (auth#, dates)
                    # on their own lines without layout interference
                    page_text_psm11 = pytesseract.image_to_string(img, config=config_psm11)
                    
                    # Combine both passes so patterns for both formats can match
                    combined = page_text
                    if page_text_psm11.strip():
                        combined = page_text + "\n--- PSM11 ---\n" + page_text_psm11
                    
                    if combined.strip():
                        page_texts.append(combined)
                
                if page_texts:
                    method = "ocr"
            except Exception as e:
                print(f"OCR error: {e}")
        
        if not page_texts:
            return "", "", "none", 0
        
        # Find the page with the authorization form
        # Returns (auth_page_text, all_text, page_num)
        auth_text, all_text, page_num = self.find_auth_page(page_texts)
        return auth_text, all_text, method, page_num
    
    def clean_ocr_date(self, value):
        """Fix common OCR errors in dates (e.g., 'fio/6/2025' -> '10/16/2025').
        
        Common OCR misreads from these PDFs:
        - "fi" at start -> "1" (e.g., fi2 -> 12, fi1 -> 11)
        - "fio" at start -> "10"
        - "fioo" -> "10/0"
        - "fin" -> "11" (n misread for 1)
        - "fia" -> "12" (a misread for 2)
        - "fioso" -> "10/80" or similar 
        - "es" at end of year -> "25" 
        - "e0" in year -> "20"
        - letters mixed with digits generally
        - "foo" -> "10/0" (common OCR error)
        - "[os" -> "08" (bracket misread as character)
        """
        if not value:
            return value
        
        # Remove any spaces within the date string first
        value = re.sub(r'\s+', '', value)
        
        # === HANDLE "oi7" PATTERN (very common OCR error for "01/") ===
        # "oi719/2026" -> o=0, i=1, 7=/ -> "01/19/2026"
        # "oi7i7/2026" -> "01/17/2026" (if both slashes read as 7, second value is i7 for 17)
        # Also handle with leading bracket: "[oi7..."
        
        # First try to detect "oi7i7" pattern (01/17) before the simpler oi7 (01/)
        # "oi7i7" = 01/17 where both slashes are read as 7 and second 1 as i
        value = re.sub(r'^[\[\(]?oi7i7', '01/17', value, flags=re.IGNORECASE)
        value = re.sub(r'^[\[\(]?oi7i9', '01/19', value, flags=re.IGNORECASE)  # In case it's really 19
        
        # Handle "oi719" as potentially "01/17" - the "9" might be a garbled "7"
        # When 17 is scanned badly, the two digits can merge to look like "19"
        # Check: if next char after oi719 is /, it's likely MMDD format where DD=17 not 19
        value = re.sub(r'^[\[\(]?oi717(?=/)', '01/17', value, flags=re.IGNORECASE)  # oi717/YYYY
        
        # Now handle single oi7 pattern
        value = re.sub(r'^[\[\(]?oi7', '01/', value, flags=re.IGNORECASE)
        
        # Handle "foi7" pattern: f often appears before oi -> foi7 = 01/
        value = re.sub(r'^foi7', '01/', value, flags=re.IGNORECASE)
        
        # Handle "foi" followed by digits -> "01/" + digits
        # E.g., "foi19/2026" -> "01/19/2026"
        value = re.sub(r'^foi(\d)', r'01/\1', value, flags=re.IGNORECASE)
        
        # === HANDLE DATES WITHOUT SLASHES (MMDDYYYY format) ===
        # Pattern: fo6724725 -> needs to become 06/24/2025
        # OCR sometimes removes slashes entirely
        
        # Check if value looks like MMDDYYYY without slashes
        # "fo" prefix followed by digits (fo = 0)
        no_slash_match = re.match(r'^fo(\d{7,8})$', value, re.IGNORECASE)
        if no_slash_match:
            digits = '0' + no_slash_match.group(1)  # fo -> 0
            # Try to parse as MMDDYYYY: first 2 = month, next 2 = day, last 4 = year
            if len(digits) >= 8:
                # Take last 8 digits as MMDDYYYY
                mmddyyyy = digits[-8:]
                month = mmddyyyy[0:2]
                day = mmddyyyy[2:4]
                year = mmddyyyy[4:8]
                # Validate
                try:
                    m, d, y = int(month), int(day), int(year)
                    if 1 <= m <= 12 and 1 <= d <= 31 and 2020 <= y <= 2030:
                        return f"{month}/{day}/{year}"
                except:
                    pass
        
        # Handle "fo67" pattern where "7" is misread "/" -> fo6/
        # E.g., "fo6724725" should be "06/24/2025" (fo=0, 6=6, 7=/, 24=24, 7=/, 2025 with missing 0)
        # Pattern: fo + digit + 7 + 2 digits + 7 + 3-4 digits
        fo_slash_match = re.match(r'^fo(\d)7(\d{2})7(\d{2,3})(\d{2})$', value, re.IGNORECASE)
        if fo_slash_match:
            # fo6724725 -> groups: (6, 24, 25, last2)? Let me think...
            # Actually: fo + 6 + 7 + 24 + 7 + 25 = fo672425 (8 chars after fo)
            # Our string is fo6724725 (9 chars) - extra char
            pass  # Will try simpler pattern below
        
        # Simpler: "fo" followed by digit, then "7" as slash separator
        # fo6724725 = "0" + "6" + "/" + "24" + "/" + "2" + "025" but 0 is missing from 2025
        # Actually try: treat 7 as / throughout
        value_with_slashes = re.sub(r'^fo(\d)7', r'0\1/', value, flags=re.IGNORECASE)
        if value_with_slashes != value:
            # Converted fo67 -> 06/
            # Now handle remaining 7 as / if it makes sense
            # "24725" -> "24/25" if 7 is /? No, year should be 4 digits
            # Let me try: "24725" where 7 = / and year = 2025 (missing leading 0 or it's "725" = "025"?)
            # Actually 24725 = 24 / 2025 with the "/" misread as nothing and year truncated
            # Try: last 4 digits as year, preceding 2 as day
            remaining = value_with_slashes[3:]  # After "06/"
            if len(remaining) >= 6:
                day = remaining[0:2]
                # Check if remaining has format like "24725" or "247025"
                year_part = remaining[2:]
                # Try inserting a slash and 20 for year
                if re.match(r'^\d{2}7\d{2,3}$', remaining):  # like 24725
                    # 24 + 7 + 25 -> 24 / 2025 (insert 20 before last 2)
                    day = remaining[0:2]
                    year_suffix = remaining[-2:]  # last 2 digits
                    value = f"0{value_with_slashes[1]}/{day}/20{year_suffix}"
                    # Validate
                    try:
                        m, d, y = int(value.split('/')[0]), int(day), int(f"20{year_suffix}")
                        if 1 <= m <= 12 and 1 <= d <= 31:
                            return value
                    except:
                        pass
        
        # Also check pure digit strings that might be MMDDYYYY
        pure_digit_match = re.match(r'^(\d{8})$', value)
        if pure_digit_match:
            mmddyyyy = pure_digit_match.group(1)
            month = mmddyyyy[0:2]
            day = mmddyyyy[2:4]
            year = mmddyyyy[4:8]
            try:
                m, d, y = int(month), int(day), int(year)
                if 1 <= m <= 12 and 1 <= d <= 31 and 2020 <= y <= 2030:
                    return f"{month}/{day}/{year}"
            except:
                pass
        
        # Handle severely corrupted dates like "fioos/20es" -> try to fix
        # Pattern: "fi" prefix followed by garbled text with slashes
        
        # NEW: Handle bracket "[" being misread - "[1/02" should be "01/02" (bracket = 0)
        # Pattern: [1/DD/YYYY -> 01/DD/YYYY (bracket represents 0)
        value = re.sub(r'^\[(\d)', r'0\1', value)  # [1 -> 01, [2 -> 02, etc.
        value = re.sub(r'\[', '1', value)  # Replace remaining brackets with 1
        
        # NEW: Handle parenthesis "(" being misread 
        value = re.sub(r'^\(', '', value)  # Remove leading paren
        value = re.sub(r'\(', '', value)  # Remove remaining parens
        
        # NEW: Handle pipe "|" being misread - often "|" is "1" at start of date
        value = re.sub(r'^\|', '', value)  # Remove leading pipe
        value = re.sub(r'\|', '1', value)  # Replace remaining pipes with 1
        
        # NEW: Handle "jo1oz" -> "01/02" (very common OCR for 01/02)
        value = re.sub(r'^jo1oz(?=/|\d)', '01/02', value, flags=re.IGNORECASE)
        
        # NEW: Handle "|oz" -> "02" (pipe-o-z = 02)
        value = re.sub(r'^oz(?=/|\d)', '02', value, flags=re.IGNORECASE)
        
        # NEW: Handle "or" -> "01" at start (o=0, r=1)
        value = re.sub(r'^or(?=/|\d)', '01', value, flags=re.IGNORECASE)
        
        # NEW: Handle "o7" -> "07" (o misread for 0)
        value = re.sub(r'^o(\d)', r'0\1', value, flags=re.IGNORECASE)
        
        # NEW: Handle "foo" -> "10/0" (very common OCR error for "10/0")
        value = re.sub(r'^foo(?=\d|/)', '10/0', value, flags=re.IGNORECASE)
        
        # NEW: Handle "fos" -> "10/8" (very common OCR error - s=8)
        value = re.sub(r'^fos(?=/|\d)', '10/8', value, flags=re.IGNORECASE)
        
        # NEW: Handle "[os" or "os" at start -> "08" 
        value = re.sub(r'^os(?=\d|/|e)', '08/', value, flags=re.IGNORECASE)
        
        # NEW: Handle "ose" pattern -> "08/2" (common misread for 08/2x)
        value = re.sub(r'^08e', '08/2', value, flags=re.IGNORECASE)
        
        # NEW: Handle "fia" followed by digit -> "12/1X" (common for 12/1X dates where /1 is lost)
        # E.g., "fia6/2025" -> "12/16/2025" (the OCR lost the /1 between 12 and 6)
        value = re.sub(r'^fia(\d)(?=/)', r'12/1\1', value, flags=re.IGNORECASE)
        
        # Handle "fia" -> "12" (a misread for 2, common for December dates)
        value = re.sub(r'^fia(?=/)', '12', value, flags=re.IGNORECASE)
        
        # NEW: Handle "fiz" -> "12" (z misread for 2) BUT only when followed by specific patterns
        # Be careful: "fz" might be "01" (f=0, z=1) in some cases like "fz 0/2025" -> "01/09/2025"
        value = re.sub(r'^fiz(?=/|\d|e|y)', '12', value, flags=re.IGNORECASE)
        
        # NEW: Handle "fz" -> "01" (f=0, z=1) when followed by a digit (like "fz0" -> "010" -> "01/0")
        # This pattern appears when OCR reads "01" as "fz"
        value = re.sub(r'^fz(\d)', r'01/0\1', value, flags=re.IGNORECASE)  # fz9 -> 01/09
        
        # NEW: Handle "fz" followed by "/" -> "01/" (f=0, z=1)
        value = re.sub(r'^fz(?=/)', '01', value, flags=re.IGNORECASE)  # fz/ -> 01/
        
        # NEW: Handle single digit day like "01/0/2025" -> might be missing day digit
        # If we have format like MM/D/YYYY with single digit day, it's likely correct as-is
        # But "01/0/2025" looks wrong - the 0 alone isn't a valid day
        
        # NEW: Handle extremely garbled patterns like "fizeayaces" -> "12/24/2025"
        # After fiz->12, we get "12eayaces" 
        # Break down: 12 + e (/) + a (2) + y (4) + aces (2025)
        # Try to detect this pattern: 12 followed by letters that should be /DD/YYYY
        value = re.sub(r'^12ea', '12/24', value, flags=re.IGNORECASE)  # ea -> /24 (e=/, a=2, implies 4)
        value = re.sub(r'^12e([a-z])', r'12/2\1', value, flags=re.IGNORECASE)  # 12e + letter -> 12/2 + letter
        
        # NEW: Handle "ay" in date -> "24" (a=2, y=4)
        value = re.sub(r'/ay(?=/|\d)', '/24', value, flags=re.IGNORECASE)
        value = re.sub(r'(\d)ay(?=/|\d)', r'\g<1>/24', value, flags=re.IGNORECASE)
        
        # NEW: Handle "y" followed by digits -> likely day part missing slash
        value = re.sub(r'^(\d{1,2})y(\d)', r'\1/\2', value, flags=re.IGNORECASE)
        
        # NEW: Handle "year" in date text -> strip it (e.g., "12year2025")
        value = re.sub(r'year(?=\d{4})', '/', value, flags=re.IGNORECASE)
        
        # Dynamic year handling - use current year instead of hardcoded 2025
        # This handles OCR errors like "yeajo0es" -> "/YYYY" where YYYY is current year
        current_year = CURRENT_YEAR_STR
        current_suffix = CURRENT_YEAR_SUFFIX  # e.g., "26" for 2026
        prev_year = PREV_YEAR_STR
        prev_suffix = PREV_YEAR_SUFFIX  # e.g., "25" for 2025
        
        # Handle garbled year patterns - try current year first, then previous year
        # Patterns like "yeajo0es", "yea...es" -> likely current or previous year
        # For "es" suffix: 25 (2025) or 26 (2026)
        # For "e6" or similar: likely 26 (2026)
        
        # Pattern "yea...es" - "es" could be 25 or 26, use current year
        value = re.sub(r'yeajo0es$', f'/{current_year}', value, flags=re.IGNORECASE)
        value = re.sub(r'yeaj00es$', f'/{current_year}', value, flags=re.IGNORECASE)
        value = re.sub(r'yea[a-z0-9]*e[56s]$', f'/{current_year}', value, flags=re.IGNORECASE)  # Catch-all for yea...e5/e6/es patterns
        
        # NEW: Handle "yeaj" or "yea" followed by digits OR more letters - OCR variant of "year"
        # In patterns like "12yeajo0es", yeaj is followed by 'o' not a digit
        # Convert the entire 'yeaj...' or 'yea...' sequence that precedes year numbers
        value = re.sub(r'yeaj[a-z0]*(?=\d)', '/', value, flags=re.IGNORECASE)  # yeajo0 -> / (before digits)
        value = re.sub(r'yeaj(?=\d)', '/', value, flags=re.IGNORECASE)
        value = re.sub(r'yea(?=\d)', '/', value, flags=re.IGNORECASE)
        
        # NEW: Handle "fizyear2025" type patterns 
        # After fiz->12: "12year2025", then year->/ gives "12/2025"
        # This is missing day, so we'd need a smarter approach
        
        # NEW: Handle "fin" -> "11" (n misread for 1, common for November dates)
        value = re.sub(r'^fin(?=/|\d)', '11', value, flags=re.IGNORECASE)
        
        # "fi" at start often means "1" - e.g., "fi2" -> "12", "fi1" -> "11"
        value = re.sub(r'^fi(\d)', r'1\1', value, flags=re.IGNORECASE)
        
        # Special case: "fio" -> "10"
        value = re.sub(r'^fio(?=/)', '10', value, flags=re.IGNORECASE)
        
        # "fioo" or "fi0o" patterns -> likely "10/0" or similar 
        value = re.sub(r'^fioo', '10/0', value, flags=re.IGNORECASE)
        value = re.sub(r'^fi0o', '10/0', value, flags=re.IGNORECASE)
        
        # NEW: Handle complex garbled patterns like "fioso77e0es" 
        # This is likely "10/07/2025" - the OCR heavily corrupted it
        # fioso -> 10/0, 77 -> /7 (duplicated), e0es -> 2025
        value = re.sub(r'^fioso', '10/0', value, flags=re.IGNORECASE)
        
        # NEW: Handle "fioye" pattern specifically -> likely "10/2" (e.g., "fioye7/20es" -> "10/27/2025")
        # The "y" is often a misread "/" and "e" is often a misread "2"
        value = re.sub(r'^fioye(\d)', r'10/2\1', value, flags=re.IGNORECASE)
        
        # NEW: Handle "fioy" -> "10/" (y misread for slash)
        value = re.sub(r'^fioy(?=\d)', '10/', value, flags=re.IGNORECASE)
        value = re.sub(r'^fi0y(?=\d)', '10/', value, flags=re.IGNORECASE)
        
        # Fallback: "fioy" -> "10/0" if followed by non-digit
        value = re.sub(r'^fioy', '10/0', value, flags=re.IGNORECASE)
        value = re.sub(r'^fi0y', '10/0', value, flags=re.IGNORECASE)
        
        # Special case: "fio/6/" is likely "10/16/" (OCR dropped the 1 from 16)
        value = re.sub(r'^fio/6/', '10/16/', value, flags=re.IGNORECASE)
        value = re.sub(r'^f[1l]o/6/', '10/16/', value, flags=re.IGNORECASE)
        
        # Common OCR misreads for "10": fio, f1o, flo, 1o, lo, lO, etc.
        value = re.sub(r'^f[1l]o(?=/)', '10', value, flags=re.IGNORECASE)
        value = re.sub(r'^[1l]o(?=/)', '10', value, flags=re.IGNORECASE)
        value = re.sub(r'^[1l][oO0](?=/)', '10', value)
        
        # Fix "s" or "es" being read instead of digits
        # "os" or "oos" -> "08" (common misread)
        value = re.sub(r's/', '8/', value, flags=re.IGNORECASE)
        value = re.sub(r'/os', '/08', value, flags=re.IGNORECASE)
        value = re.sub(r'/oos', '/08', value, flags=re.IGNORECASE)
        
        # NEW: Fix "e9" in middle -> "29" (e misread for 2)  
        value = re.sub(r'/e9/', '/29/', value, flags=re.IGNORECASE)
        value = re.sub(r'^e9/', '29/', value, flags=re.IGNORECASE)
        
        # NEW: Fix "y" misread as part of number - often means there's a missing slash
        # e.g., "0y3" might be "0/3" or "03"
        value = re.sub(r'(\d)y(\d)', r'\1/\2', value, flags=re.IGNORECASE)
        
        # NEW: Fix "e" between digits -> probably "2" or separator issue
        value = re.sub(r'(\d)e(\d)(?=/)', r'\1/\2', value, flags=re.IGNORECASE)
        
        # NEW: Fix duplicated digits that OCR sometimes produces
        # e.g., "77" when it should be "/7" or "7/"
        value = re.sub(r'/(\d)\1(?=/)', r'/\1', value)  # /77/ -> /7/
        
        # NEW: Fix "e0" in year -> "20" (e misread for 2)
        value = re.sub(r'/e0(\d\d)$', r'/20\1', value, flags=re.IGNORECASE)
        value = re.sub(r'/(\d)e(\d\d)$', r'/\g<1>0\2', value, flags=re.IGNORECASE)  # 7e25 -> 7025 (then needs more fix)
        
        # Dynamic year patterns using current year (handles 2025, 2026, etc.)
        # "es" at end -> likely "25" or "26" depending on current year
        # Match /20es, /20e5, /20e6 and replace with current year
        # IMPORTANT: Don't match valid years like /2025, /2026 - only OCR errors
        value = re.sub(r'/20e[s]$', f'/{current_year}', value, flags=re.IGNORECASE)
        value = re.sub(r'/2oe[s]$', f'/{current_year}', value, flags=re.IGNORECASE)
        value = re.sub(r'/202s$', f'/{current_year}', value, flags=re.IGNORECASE)  # Only letter s, not digits
        value = re.sub(r'e0e[s]$', current_year, value, flags=re.IGNORECASE)  # Handle just "e0es" -> current year
        
        # Handle "eces" at end -> current year (e=2, c=0, e=2, s=5)
        value = re.sub(r'ece[s56]$', current_year, value, flags=re.IGNORECASE)
        
        # Handle "aces" at end -> current year (common OCR error for 2025/2026)
        # IMPORTANT: yaces must come BEFORE aces to avoid leaving 'y' behind
        value = re.sub(r'yace[s56]$', f'/{current_year}', value, flags=re.IGNORECASE)  # yaces -> /YYYY (y is the slash)
        value = re.sub(r'/ace[s56]$', f'/{current_year}', value, flags=re.IGNORECASE)
        value = re.sub(r'ace[s56]$', current_year, value, flags=re.IGNORECASE)
        
        # Handle "jo0es", "j00es", etc. at end -> current year (j=2, o/0=0, es=25/26)
        value = re.sub(r'jo0e[s56]$', current_year, value, flags=re.IGNORECASE)
        value = re.sub(r'j00e[s56]$', current_year, value, flags=re.IGNORECASE)
        value = re.sub(r'j0e[s56]$', current_year, value, flags=re.IGNORECASE)
        
        # Handle "o0es" at end -> current year 
        value = re.sub(r'/o0e[s56]$', f'/{current_year}', value, flags=re.IGNORECASE)
        value = re.sub(r'o0e[s56]$', current_year, value, flags=re.IGNORECASE)
        
        # Handle "/re/" pattern -> "/01/" (r=0, e=1)
        value = re.sub(r'/re/', '/01/', value, flags=re.IGNORECASE)
        
        # NEW: Handle 2026-specific OCR patterns
        # "e6" at end often means "26" for 2026
        value = re.sub(r'/20e6$', f'/{current_year}', value, flags=re.IGNORECASE)
        value = re.sub(r'/202e6$', f'/{current_year}', value, flags=re.IGNORECASE)
        # "2o26" -> "2026" (o misread for 0)
        value = re.sub(r'/2o26$', f'/{current_year}', value, flags=re.IGNORECASE)
        # "20z6" -> "2026" (z misread for 2)
        value = re.sub(r'/20z6$', f'/{current_year}', value, flags=re.IGNORECASE)
        
        # NEW: Handle completely garbled years like "year2025" -> "2025"
        value = re.sub(r'year(\d{4})$', r'/\1', value, flags=re.IGNORECASE)
        
        # NEW: Fix 3-digit years (OCR corruption like "995" -> "2025" or "926" -> "2026")
        # Pattern: MM/DD/9XX -> MM/DD/20XX, MM/DD/0XX -> MM/DD/20XX
        value = re.sub(r'/9(\d{2})$', r'/20\1', value)  # /925 -> /2025, /926 -> /2026
        value = re.sub(r'/0(\d{2})$', r'/20\1', value)  # /025 -> /2025, /026 -> /2026
        
        # Fix within date (after first /)
        value = re.sub(r'(?<=/)[fli1][iloO0](?=/)', '10', value)
        
        # Fix O/o read as 0 in dates
        value = re.sub(r'[oO](?=\d)', '0', value)
        value = re.sub(r'(?<=\d)[oO]', '0', value)
        
        # NEW: Final cleanup - ensure we have proper date separators
        # If we have more than 2 slashes, something went wrong, try to fix
        slash_count = value.count('/')
        if slash_count > 2:
            # Keep only first two slashes
            parts = value.split('/')
            if len(parts) >= 3:
                # Try to reconstruct: take first part, second part, and join rest for year
                value = f"{parts[0]}/{parts[1]}/{parts[2]}"
        
        return value
    
    def extract_ocr_date(self, raw_text):
        """
        Try to extract a date from OCR text that might be garbled.
        Returns cleaned date or None if can't parse.
        """
        # First clean the input
        cleaned = self.clean_ocr_date(raw_text)
        
        # Try to match standard date pattern
        match = re.match(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', cleaned)
        if match:
            return f"{match.group(1)}/{match.group(2)}/{match.group(3)}"
        
        return None
    
    def extract_smart(self, text):
        """
        Smart extraction for PACE Treatment Authorization Form.
        
        Form layout (all on same page with "TREATMENT AUTHORIZATION FORM" header):
        - Auth #: [8-digit number]
        - Date Approved: [MM/DD/YYYY]  
        - Date Auth. Expire: [MM/DD/YYYY]
        - Participant's Name: [LAST, FIRST MIDDLE]
        - Participant ID: [9-digit number]
        """
        results = {}
        
        # ===== PRE-PROCESSING: fix common OCR artifacts before pattern matching =====
        # 1. Collapse spaces within auth numbers: "[205175 56" -> "[20517556"
        #    When we see digits-space-digits with no slash (not a date), combine them
        text = re.sub(r'([\[\(]?\d{4,7})\s+(\d{1,4})(?!\s*[/\d])', 
                      lambda m: m.group(1) + m.group(2), text)
        # 2. Collapse spaces within dates after '[': "[o6/ 02/2025" -> "[o6/02/2025"
        text = re.sub(r'([\[\(]?[o0\d]{1,2}/)\s+(\d{1,2}/\d{4})', 
                      lambda m: m.group(1) + m.group(2), text)
        # 3. Remove spaces between '[' and digits: "[ 23924378" -> "[23924378"
        text = re.sub(r'[\[\(]\s+(\d)', r'[\1', text)
        
        # ===== HEADER LINE EXTRACTION =====
        # Form has: "Auth #: Date Approved: Date Auth. Expire:\n[AUTH DATE1 DATE2"
        # Try to extract all 3 values from a single header line first
        # NOTE: OCR sometimes adds spaces within dates (e.g., "fia 6/2025" should be "12/16/2025")
        # NOTE: Auth# may have digits split by a space (e.g., "205175 56" = "20517556") - handled by pre-processing above
        
        # PSM11 BLOCK MATCH: Labels appear all at once, then values all at once
        # "Auth 4:\n\nDate Approved:\n\nDate Auth. Expire:\n\n[20517556\n\n[o6/02/2025\n\n[12/07/2025"
        psm11_block = re.search(
            r'Auth\s*[#4]:?\s*Date\s*Approved:?\s*Date\s*Auth\.?\s*Expire:?\s*'
            r'([\[\(|]?[\d\s]{6,12})\s+'     # Auth# value (digits, may have space in them)
            r'([\[\(|]?[o0\d]{1,2}/\S+\d{4})\s+'  # Date Approved (may have OCR prefix)
            r'([\[\(|]?\d{1,2}/\d{1,2}/\d{4})',   # Date Expire (usually cleaner)
            text, re.IGNORECASE
        )
        if psm11_block:
            raw_auth = re.sub(r'^[\[\(|]', '', psm11_block.group(1).strip())
            raw_auth = re.sub(r'\s', '', raw_auth)  # Remove any spaces within number
            raw_auth = re.sub(r'[^\d]', '', raw_auth)
            if len(raw_auth) >= 7:
                results["Auth #"] = raw_auth
            
            raw_approved = re.sub(r'^[\[\(|]', '', psm11_block.group(2).strip())
            cleaned_approved = self.clean_ocr_date(raw_approved)
            if cleaned_approved and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned_approved):
                results["Date Approved"] = cleaned_approved
            
            raw_expire = re.sub(r'^[\[\(|]', '', psm11_block.group(3).strip())
            cleaned_expire = self.clean_ocr_date(raw_expire)
            if cleaned_expire and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned_expire):
                results["Date Auth Expire"] = cleaned_expire
        
        # PSM11 BLOCK VARIANT B: Labels grouped vertically, then values as date_approved, date_expire, auth#
        # Seen in some PDFs where PSM11 reads columns as: all labels first, then all values
        # "Auth 4:\n\nDate Approved:\n\nDate Auth. Expire:\n\n11/10/2025\n\n[04/30/2026\n\n[25881916"
        if not all(k in results for k in ["Auth #", "Date Approved", "Date Auth Expire"]):
            psm11_block_v2 = re.search(
                r'Auth\s*[#4?]:?\s+'
                r'Date\s+Approved:?\s+'
                r'Date\s+Auth\.?\s*Expire:?\s+'
                r'([\[\(|]?[o0\d]{1,2}/\S{1,5}/\d{4})\s+'   # date_approved (first value after labels)
                r'([\[\(|]?\d{1,2}/\S{1,5}/\d{4})\s+'        # date_expire (second value)
                r'([\[\(|]?[\d]{7,10})',                       # auth_num (third value)
                text, re.IGNORECASE
            )
            if psm11_block_v2:
                raw_auth = re.sub(r'^[\[\(|]', '', psm11_block_v2.group(3).strip())
                raw_auth = re.sub(r'[^\d]', '', raw_auth)
                if len(raw_auth) >= 7 and "Auth #" not in results:
                    results["Auth #"] = raw_auth

                raw_approved = re.sub(r'^[\[\(|IlL]', '', psm11_block_v2.group(1).strip())
                cleaned_approved = self.clean_ocr_date(raw_approved)
                if cleaned_approved and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned_approved) and "Date Approved" not in results:
                    results["Date Approved"] = cleaned_approved

                raw_expire = re.sub(r'^[\[\(|IlL]', '', psm11_block_v2.group(2).strip())
                cleaned_expire = self.clean_ocr_date(raw_expire)
                if cleaned_expire and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned_expire) and "Date Auth Expire" not in results:
                    results["Date Auth Expire"] = cleaned_expire

        # PSM3 HEADER MATCH: All labels on one line, values on next line
        # "Auth #: Date Approved: Date Auth. Expire:\n[AUTH DATE1 DATE2"
        header_match = re.search(
            r'Auth\s*#:.*?Date\s*Approved:.*?Date\s*Auth\.?\s*Expire:\s*\n+\s*'
            r'[\[\(]?([^\s]+)\s+'  # Auth # (may have OCR letters mixed with digits)
            r'[\[\(|]?([^\s]+)\s+'  # Date Approved part 1 (may be garbled)
            r'[\[\(|]?([^\s]+)',  # Date Approved part 2 OR Date Auth Expire
            text, re.IGNORECASE
        )
        
        if header_match and not all(k in results for k in ["Auth #", "Date Approved", "Date Auth Expire"]):
            raw_auth = header_match.group(1)
            raw_approved = header_match.group(2)
            raw_third = header_match.group(3)  # Could be Date Approved part 2 or Date Auth Expire
            
            # Check if raw_approved looks incomplete (missing slash or year)
            # E.g., "fia" without "/YYYY" means the date is split across tokens
            looks_like_split_date = not re.search(r'/\d{4}$', raw_approved)
            
            if looks_like_split_date and re.search(r'/\d{4}$', raw_third):
                # Date Approved is split: combine raw_approved + raw_third
                # And then look for Date Auth Expire in the next token
                raw_approved = raw_approved + raw_third  # e.g., "fia" + "6/2025" = "fia6/2025"
                # Find Date Auth Expire - it's the next token after raw_third in the text
                expire_match = re.search(
                    re.escape(raw_third) + r'\s+[\[\(|]?(\d{1,2}/\d{1,2}/\d{4})',
                    text
                )
                raw_expire = expire_match.group(1) if expire_match else None
            else:
                # Normal case: raw_third is Date Auth Expire
                raw_expire = raw_third
            
            # Clean Auth # - convert OCR letters to digits
            if "Auth #" not in results:
                clean_auth = raw_auth
                clean_auth = re.sub(r'^[\[\(|f]', '', clean_auth)
                clean_auth = re.sub(r'[fpF]', '0', clean_auth)
                clean_auth = re.sub(r'[eE]', '6', clean_auth)
                clean_auth = re.sub(r'[oO]', '0', clean_auth)
                clean_auth = re.sub(r'[lIi]', '1', clean_auth)
                clean_auth = re.sub(r'[cC]', '2', clean_auth)
                clean_auth = re.sub(r'[rR]', '7', clean_auth)
                clean_auth = re.sub(r'[aA]', '4', clean_auth)
                clean_auth = re.sub(r'[tT]', '7', clean_auth)
                clean_auth = re.sub(r'[sS]', '8', clean_auth)
                clean_auth = re.sub(r'[zZ]', '2', clean_auth)
                clean_auth = re.sub(r'[^\d]', '', clean_auth)
                if len(clean_auth) >= 7:
                    results["Auth #"] = clean_auth
            
            # Clean Date Approved
            if "Date Approved" not in results:
                cleaned_approved = self.clean_ocr_date(raw_approved)
                if cleaned_approved and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned_approved):
                    results["Date Approved"] = cleaned_approved
            
            # Clean Date Auth Expire
            if "Date Auth Expire" not in results and raw_expire:
                cleaned_expire = self.clean_ocr_date(raw_expire)
                if cleaned_expire and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned_expire):
                    results["Date Auth Expire"] = cleaned_expire
        
        # ===== 1. AUTH # (fallback if not found above) =====
        # Look for Auth # followed by 7-8 digit number
        # OCR may add [ or ( before the number
        # Auth # can be on a separate line from the label
        if "Auth #" not in results:
            auth_patterns = [
                r'Auth\s*#:?\s*\n?\s*[\[\(]?(\d{7,10})',  # Auth #: [23924378 or Auth #:\n23924378
                r'Auth\s*#:?\s*[|\s]*[\[\(]?(\d{7,10})',  # With pipes
                r'[\[\(](\d{8})\s+(?:\d{4}/\d{4}|\d{1,2}/\d{1,2}/\d{4})',  # [26370364 1203/2025 or [26370364 12/03/2025
                r'[\[\(I](\d{8})\s+',  # [26370364 - auth at start of line with [ or I
            ]
            for pattern in auth_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    results["Auth #"] = match.group(1).strip()
                    break
        
        # ===== 2. DATE APPROVED =====
        # Look for Date Approved: followed by date
        # Also check header row: "26370364 1203/2025 02/28/2026"
        if "Date Approved" not in results:
            date_approved_patterns = [
                r'Date\s*Approved:?\s*\n?\s*(\d{1,2}/\d{1,2}/\d{4})',  # Clean date
                r'Date\s*Approved:?\s*\n?\s*[\[\(|]?([o0\d]{1,2}/\d{1,2}/\d{4})',  # [o6/02/2025 -> 06/02/2025
                r'Date\s*Approved:?\s*\n?\s*[\[\(|]?([fioa\d]{1,5}/\d{1,2}/\d{4})',  # With OCR errors + bracket
                r'Date\s*Approved:?\s*\n?\s*(\d{4}/\d{4})',  # MMDD/YYYY format (OCR missing slash)
            ]
            for pattern in date_approved_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    raw_date = match.group(1).strip()
                    # Fix missing slash: MMDD/YYYY -> MM/DD/YYYY
                    if re.match(r'^\d{4}/\d{4}$', raw_date):
                        raw_date = raw_date[:2] + '/' + raw_date[2:]  # Insert slash after 2 digits
                    cleaned = self.clean_ocr_date(raw_date)
                    if cleaned and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned):
                        results["Date Approved"] = cleaned
                        break
        
        # Fallback: Look for dates after Auth# in header row
        # Format: "[26370364 1203/2025 (02/28/2026" where first date is Approved, second is Expire
        # OCR may add bracket before auth number
        # Also handle OCR errors like "[oi22/2026" where "oi" = "01"
        if "Date Approved" not in results and "Auth #" in results:
            auth_num = results["Auth #"]
            # More flexible pattern: capture any non-whitespace with a slash, then a clean date
            auth_header = re.search(
                rf'[\[\(]?{auth_num}\s+[\[\(]?([^\s]+/\d{{4}})\s*[\[\(]?(\d{{1,2}}/\d{{1,2}}/\d{{4}})',
                text
            )
            if auth_header:
                raw_approved = auth_header.group(1)
                raw_expire = auth_header.group(2)
                # Clean OCR prefix like "[oi" -> "01", "oi" -> "01"
                raw_approved = re.sub(r'^[\[\(]?o[i1l]', '01', raw_approved, flags=re.IGNORECASE)
                raw_approved = re.sub(r'^[\[\(]', '', raw_approved)
                # Handle MMDD/YYYY format
                if re.match(r'^\d{4}/\d{4}$', raw_approved):
                    raw_approved = raw_approved[:2] + '/' + raw_approved[2:]
                cleaned_approved = self.clean_ocr_date(raw_approved)
                if cleaned_approved and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned_approved):
                    results["Date Approved"] = cleaned_approved
        
        # ===== 3. DATE AUTH EXPIRE =====
        # Look for Date Auth. Expire: or Date Auth Expire: followed by date
        if "Date Auth Expire" not in results:
            date_expire_patterns = [
                r'Date\s*Auth\.?\s*Expire:?\s*\n?\s*(\d{1,2}/\d{1,2}/\d{4})',
                r'Date\s*Auth\.?\s*Expire:?\s*\n?\s*([fioa\d]{1,5}/\d{1,2}/\d{4})',
                r'Date\s*Auth\.?\s*Expire:?\s*\n?\s*[\[\(|]?(\d{1,2}/\d{1,2}/\d{4})',  # With leading bracket/pipe
            ]
            for pattern in date_expire_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    cleaned = self.clean_ocr_date(match.group(1).strip())
                    if cleaned and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned):
                        results["Date Auth Expire"] = cleaned
                        break
        
        # Fallback: Look for valid date at end of header line (after Auth # and garbled Date Approved)
        # Pattern: "22928116 fo6724725 12/23/2025" - the last token is Date Auth Expire
        if "Date Auth Expire" not in results and "Auth #" in results:
            auth_num = results["Auth #"]
            # Look for Auth # followed by anything, then a clean date at end
            expire_fallback = re.search(
                rf'{auth_num}\s+\S+\s+(\d{{1,2}}/\d{{1,2}}/\d{{4}})',
                text
            )
            if expire_fallback:
                cleaned = self.clean_ocr_date(expire_fallback.group(1).strip())
                if cleaned and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned):
                    results["Date Auth Expire"] = cleaned
        
        # ===== 4. PARTICIPANT NAME =====
        # OCR always adds a leading '[', 'I', or 'l' artifact before the uppercase name.
        # Layouts observed:
        #   PSM3 inline:   "Participant's Name: DOB:\n[CROSS, WILLIAM EARL [07/06/1955"
        #   PSM3 split:    "Participant's Name: DOB:\n\nFIGUEROA, REINA\n\n[08/13/1950"
        #   PSM11 split:   "Participant's Name:\n\nDOB:\n\nICROSS, WILLIAM EARL\n"
        #   Footer clean:  "WILLIAM CROSS(DOB:07/06/1955) Pagel of 2"  (FIRST LAST order)
        OCR_PREFIX = r"[\[\(|IlL1]?"  # leading artifact to strip
        NAME_CHARS = r"[A-Z][A-Z,\s\-'\.]+"  # valid name characters (all caps, comma, space, hyphen)
        name_patterns = [
            # PSM3 inline: name and DOB on same value line
            rf"Participant'?s?\s*Name:?\s*DOB:?\s*\n+\s*{OCR_PREFIX}({NAME_CHARS}?)\s+{OCR_PREFIX}\d{{1,2}}/\d{{1,2}}/\d{{4}}",
            # PSM3 split: name on own line after combined label, no DOB on same line
            rf"Participant'?s?\s*Name:?\s*DOB:?\s*\n+\s*{OCR_PREFIX}({NAME_CHARS}?)(?:\n|$)",
            # PSM11 split: name appears AFTER DOB label on its own line
            rf"Participant'?s?\s*Name:?\s*\n+\s*DOB:?\s*\n+\s*{OCR_PREFIX}({NAME_CHARS}?)(?:\n|$)",
            # Older formats / same-line with Name only label
            rf"Participant'?s?\s*Name:?\s*\n?\s*{OCR_PREFIX}({NAME_CHARS}?)(?:\s+DOB|\n\s*Medical)",
        ]
        for pattern in name_patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                name = match.group(1).strip()
                # Strip leading OCR artifact: '[', '(', '|', 'I'/'l'/'1' before an uppercase letter
                name = re.sub(r'^[I\[\(|l1](?=[A-Z])', '', name)
                # Strip trailing OCR artifacts and whitespace
                name = re.sub(r'[\[\]\|\(\)]+', '', name)
                name = re.sub(r'\s+', ' ', name).strip()
                if len(name) >= 3 and ',' in name and not name.replace(',', '').replace('-', '').replace(' ', '').isdigit():
                    results["Patient Name"] = name
                    break
        
        # Footer fallback: page footer always has clean "FIRSTNAME LASTNAME(DOB:MM/DD/YYYY)"
        # Convert FIRST LAST -> LAST, FIRST for consistency
        if "Patient Name" not in results:
            footer_match = re.search(
                r'\b([A-Z]{2,}(?:\s+[A-Z]{2,})+)\(DOB:\d{1,2}/\d{1,2}/\d{4}\)',
                text
            )
            if footer_match:
                full_name = footer_match.group(1).strip()
                parts = full_name.split()
                if len(parts) >= 2:
                    # Last word is the last name, rest is first/middle
                    last = parts[-1]
                    first = ' '.join(parts[:-1])
                    results["Patient Name"] = f"{last}, {first}"
        
        # ===== 5. PARTICIPANT ID =====
        # Look for Participant ID: followed by 9-digit number (may be on next line)
        # OCR often garbles "10" as "fo", "fro" and "100" as "fio", "fron", "foo", "fot", "froo" etc.
        # Also may add [ or f before the number, and mix letters with digits (s=8, e=6, o=0, t=7)
        # May have SPACE in the middle of the ID!
        # Examples from real OCR:
        #   "Participant ID:\n\nfron 308633" -> ID is "100308633"
        #   "Participant ID:\nfot 390355" -> ID is "100390355"
        #   "Participant ID:\n\nfo1419975" -> ID is "101419975"
        #   "Participant ID:\n\nfro0038800" -> ID is "1000038800"
        #   "Participant ID:\n\nfroosose31" -> ID is "100808631" (froo=100, s=8, o=0, s=8, e=6)
        #   "Participant ID:\n\nfroo151102" -> ID is "100151102" (froo=100)
        #   "Participant ID:\n\nfrooet 5980" -> ID is "100675980" (froo=100, e=6, t=7, space, 5980)
        #   "Participant ID:\nfiotazaci 6" -> ID is "107424216" (fio=10, t=7, a=4, z=2, a=4, c=2, i=1, space, 6)
        id_patterns = [
            r'Participant\s*ID:?\s*\n?\s*[\[\(f]?(\d{6,})',  # Clean digit ID (min 6), with optional [ or f
            r'Participant\s*ID:?\s*\n?\s*(fio[a-z0-9 ]{6,})',  # fio + alphanumeric (fio=10), allows internal space
            r'Participant\s*ID:?\s*\n?\s*(fron[a-z0-9 ]{5,})',  # fron + mixed (fron=100)
            r'Participant\s*ID:?\s*\n?\s*(froo[a-z0-9 ]{5,})',  # froo + mixed (froo=100), allows internal space
            r'Participant\s*ID:?\s*\n?\s*(foo[a-z0-9 ]{5,})',  # foo + mixed (foo=100)
            r'Participant\s*ID:?\s*\n?\s*(fot[a-z0-9 ]{5,})',  # fot + mixed (fot=100)
            r'Participant\s*ID:?\s*\n?\s*(fo[a-z0-9 ]{6,})',  # fo + mixed (fo=10)
            r'Participant\s*ID:?\s*\n?\s*(fro[a-z0-9 ]{6,})',  # fro + mixed (fro=10)
            r'Participant\s*ID:?\s*\n?\s*[f\[]?([1ifl][0o]0[a-z0-9 ]{5,})',  # 100 with OCR errors
        ]
        for pattern in id_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                patient_id = match.group(1).strip()
                # Remove spaces within the ID
                patient_id = re.sub(r'\s+', '', patient_id)
                # Clean OCR errors - convert letter prefixes to numbers
                patient_id = re.sub(r'^fron\s*', '100', patient_id, flags=re.IGNORECASE)
                patient_id = re.sub(r'^froo', '100', patient_id, flags=re.IGNORECASE)  # froo before fro
                patient_id = re.sub(r'^foo\s*', '100', patient_id, flags=re.IGNORECASE)
                patient_id = re.sub(r'^fot\s*', '100', patient_id, flags=re.IGNORECASE)
                patient_id = re.sub(r'^fro', '10', patient_id, flags=re.IGNORECASE)
                patient_id = re.sub(r'^fio', '10', patient_id, flags=re.IGNORECASE)
                patient_id = re.sub(r'^fo', '10', patient_id, flags=re.IGNORECASE)
                patient_id = re.sub(r'^[1ifl][0o]0', '100', patient_id, flags=re.IGNORECASE)
                # Convert OCR letters to digits throughout
                patient_id = re.sub(r'[oO]', '0', patient_id)  # o -> 0
                patient_id = re.sub(r'[sS]', '8', patient_id)  # s -> 8
                patient_id = re.sub(r'[eE]', '6', patient_id)  # e -> 6
                patient_id = re.sub(r'[tT]', '7', patient_id)  # t -> 7
                patient_id = re.sub(r'[iIlL]', '1', patient_id)  # i/l -> 1
                patient_id = re.sub(r'[aA]', '4', patient_id)  # a -> 4
                patient_id = re.sub(r'[zZ]', '2', patient_id)  # z -> 2
                patient_id = re.sub(r'[cC]', '2', patient_id)  # c -> 2 (common)
                patient_id = re.sub(r'[^\d]', '', patient_id)  # Remove remaining non-digits
                if len(patient_id) >= 6:  # Minimum 6 digits for a valid ID
                    results["Patient ID"] = patient_id
                    break
        
        # ===== 6. FALLBACK: DATE APPROVED FROM REVIEWER DATE =====
        # If we couldn't extract Date Approved, look in "Reviewer's Name: Date:" section
        # That date typically matches Date Approved
        # Format: "Reviewer's Name: Date:\n\nSandra Coronel, RS Mgr. 06/23/2025"
        if "Date Approved" not in results:
            reviewer_match = re.search(
                r"Reviewer'?s?\s*Name:?\s*Date:?\s*\n*\s*(?:[A-Za-z\s,.\[\]]+?)\s*[fhi]*(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})",
                text,
                re.IGNORECASE
            )
            if reviewer_match:
                month, day, year = reviewer_match.groups()
                # Fix 2-digit year
                if len(year) == 2:
                    year = "20" + year
                raw_date = f"{month}/{day}/{year}"
                cleaned = self.clean_ocr_date(raw_date)
                if cleaned and re.match(r'\d{1,2}/\d{1,2}/\d{4}', cleaned):
                    results["Date Approved"] = cleaned
        
        return results
    
    def validate_and_fix_dates(self, result):
        """
        Validate extracted dates and fix common issues:
        1. Swap dates if Expire is before Approved
        2. Validate year is reasonable (2020-2030)
        3. Flag suspicious same-day dates
        4. Fix corrupted years like "995" -> "2025" or "926" -> "2026"
        """
        from datetime import datetime, timedelta
        
        date_approved = result.get("Date Approved")
        date_expire = result.get("Date Auth Expire")
        
        # Helper to parse date safely
        def parse_date(date_str):
            if not date_str:
                return None
            try:
                return datetime.strptime(date_str, "%m/%d/%Y")
            except:
                return None
        
        # Helper to fix corrupted years
        def fix_year(date_str):
            if not date_str:
                return date_str
            # Pattern: check if year part is corrupted
            match = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{3,4})$', date_str)
            if match:
                month, day, year = match.groups()
                year_int = int(year)
                # Fix 3-digit years (like 995, 925, 926)
                if len(year) == 3:
                    if year.startswith('9'):
                        # 925 -> 2025, 926 -> 2026
                        year = f"20{year[1:]}"
                    else:
                        # Other 3-digit -> assume 2020s
                        year = f"2{year}"
                # Validate year range
                year_int = int(year)
                if year_int < 2020:
                    # Likely should be 2020s - fix last 2 digits
                    if 20 <= year_int % 100 <= 30:
                        year = f"20{year_int % 100}"
                elif year_int > 2030:
                    # Too far future, might be corrupted
                    # Check if it's a simple corruption
                    if year.endswith('25'):
                        year = '2025'
                    elif year.endswith('26'):
                        year = '2026'
                return f"{month}/{day}/{year}"
            return date_str
        
        # Fix corrupted years first
        if date_approved:
            result["Date Approved"] = fix_year(date_approved)
            date_approved = result["Date Approved"]
        if date_expire:
            result["Date Auth Expire"] = fix_year(date_expire)
            date_expire = result["Date Auth Expire"]
        
        # Parse dates
        approved_dt = parse_date(date_approved)
        expire_dt = parse_date(date_expire)
        
        # Validate and fix order if needed
        if approved_dt and expire_dt:
            # If dates are swapped (Expire before Approved), consider swapping them
            # BUT only swap if the difference is significant (> 7 days)
            # Small differences (1-7 days) could be OCR errors like 17->19
            if expire_dt < approved_dt:
                days_diff = (approved_dt - expire_dt).days
                if days_diff > 7:
                    # Significant difference - likely genuinely swapped
                    result["Date Approved"], result["Date Auth Expire"] = result["Date Auth Expire"], result["Date Approved"]
                    result["dates_swapped"] = True
                else:
                    # Small difference - could be OCR error, flag it but don't swap
                    result["dates_close_warning"] = f"Dates within {days_diff} days, might be OCR error"
            
            # If dates are the same, flag it but keep them (might be legitimate 1-day auth)
            if expire_dt == approved_dt:
                result["dates_same_warning"] = True
        
        return result
    
    def extract_field(self, text, field_name):
        """Extract a field value using pattern matching."""
        patterns = self.PATTERNS.get(field_name, [])
        
        for pattern in patterns:
            # Use DOTALL to match across newlines
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE | re.DOTALL)
            if match:
                value = match.group(1).strip()
                # Clean up the value - remove extra whitespace/newlines
                value = re.sub(r'[\n\r]+', ' ', value)
                value = re.sub(r'\s+', ' ', value)
                # Fix OCR errors in dates
                if "Date" in field_name:
                    value = self.clean_ocr_date(value)
                return value
        
        return None
    
    def extract_from_filename(self, filename):
        """Extract patient name and date from filename as fallback."""
        # Pattern: "First Last ... MMDDYYYY.pdf"
        name_match = re.match(r'^([A-Z][a-z]+(?:\s+[A-Z][a-z-]+)+)', filename)
        date_match = re.search(r'(\d{6,8})(?:[-.]|\.pdf)?$', filename, re.IGNORECASE)
        
        result = {"Patient Name": None, "date_hint": None}
        
        if name_match:
            result["Patient Name"] = name_match.group(1)
        
        if date_match:
            date_str = date_match.group(1)
            try:
                if len(date_str) == 8:
                    result["date_hint"] = datetime.strptime(date_str, "%m%d%Y")
                elif len(date_str) == 6:
                    result["date_hint"] = datetime.strptime(date_str, "%m%d%y")
            except ValueError:
                pass
        
        return result
    
    def process_pdf(self, pdf_path):
        """Process a single PDF and extract all fields."""
        pdf_path = pathlib.Path(pdf_path)
        result = {
            "file": pdf_path.name,
            "extracted_at": datetime.now().isoformat(),
        }
        
        # Initialize all fields as None
        for field in FIELDS:
            result[field] = None
        
        # Get fallback data from filename
        filename_data = self.extract_from_filename(pdf_path.stem)
        
        # TRY AZURE DOCUMENT INTELLIGENCE FIRST (best quality)
        if self.use_azure:
            azure_results = self.extract_with_azure(pdf_path)
            if azure_results:
                result["extraction_method"] = "azure"
                for field in FIELDS:
                    if field in azure_results and azure_results[field]:
                        result[field] = azure_results[field]
                        if f"{field}_confidence" in azure_results:
                            result[f"{field}_confidence"] = azure_results[f"{field}_confidence"]
                        if f"{field}_method" in azure_results:
                            result[f"{field}_method"] = azure_results[f"{field}_method"]
                
                # If Azure got all fields, we're done
                if all(result.get(field) for field in FIELDS):
                    # Use filename fallback for patient name if needed
                    if not result["Patient Name"] and filename_data["Patient Name"]:
                        result["Patient Name"] = filename_data["Patient Name"]
                        result["Patient Name_method"] = "filename"
                    return result
        
        # FALLBACK TO LOCAL OCR + ML/REGEX
        # Extract text from PDF (finds the auth form page automatically)
        # auth_text = just the auth form page, all_text = all pages combined
        auth_text, all_text, method, page_num = self.extract_text_from_pdf(pdf_path)
        result["extraction_method"] = method if not self.use_azure else f"azure + {method}"
        result["auth_page"] = page_num if page_num > 0 else "all"
        
        # Use auth_text for main extraction (Auth #, dates, name)
        # Use all_text for Patient ID extraction (may be on different pages)
        text = auth_text  # Primary text source
        
        if not text.strip():
            result["error"] = "Could not extract text (encrypted or corrupted)"
            if filename_data["Patient Name"]:
                result["Patient Name"] = filename_data["Patient Name"]
            return result
        
        # Store raw text for debugging (first 800 chars to see more context)
        result["raw_text_preview"] = text[:800].replace('\n', ' | ')
        
        # Run smart regex extraction FIRST (most reliable for auth form fields)
        smart_results = self.extract_smart(text)
        for field in FIELDS:
            if not result[field] and field in smart_results and smart_results[field]:
                result[field] = smart_results[field]
                result[f"{field}_method"] = "regex_smart"
        
        # Try ML extraction for any fields regex couldn't find
        ml_results = None
        if self.use_ml and self.ml_extractor:
            if not all(result.get(field) for field in FIELDS):
                ml_results = self.ml_extractor.extract_with_confidence(text)
                result["extraction_method"] = f"{method} + ML"
                
                # Use ML results only for fields still missing after regex
                for field in FIELDS:
                    if not result[field] and field in ml_results and ml_results[field]["value"]:
                        result[field] = ml_results[field]["value"]
                        result[f"{field}_confidence"] = ml_results[field]["confidence"]
                        result[f"{field}_method"] = ml_results[field]["method"]
        
        # Fallback: Use smart extraction again for any still-missing fields (already done above)
        
        # Fallback: try old pattern matching for any missing fields
        for field in FIELDS:
            if not result[field]:
                value = self.extract_field(text, field)
                if value:
                    result[field] = value
                    if not result.get(f"{field}_method"):
                        result[f"{field}_method"] = "regex_pattern"
        
        # AGGRESSIVE FALLBACK for Auth # - search for any 7-10 digit number near "Auth" or on auth form page
        if not result["Auth #"]:
            # Try to find auth number by looking for 7-10 digit numbers that aren't phone/ID patterns
            auth_candidates = []
            
            # Pattern 1: Numbers right after "Auth #:" or similar labels
            auth_label_match = re.search(r'Auth[:\s#]+(\d{7,10})', text, re.IGNORECASE)
            if auth_label_match:
                auth_candidates.append(auth_label_match.group(1))
            
            # Pattern 2: Standalone 8-digit numbers (most common auth format)
            standalone_8digit = re.findall(r'\b(\d{8})\b', text)
            for num in standalone_8digit:
                # Filter out obvious non-auth numbers (phone patterns, years repeated, etc.)
                if not (num.startswith('1') and len(num) == 10):  # Not phone number
                    if num not in auth_candidates:
                        auth_candidates.append(num)
            
            # Pattern 3: 7-digit numbers (less common but possible)
            standalone_7digit = re.findall(r'\b(\d{7})\b', text)
            for num in standalone_7digit:
                # Filter out zip codes and dates (9999999 format unlikely to be auth)
                if not num.startswith('9') and not all(c == num[0] for c in num):
                    if num not in auth_candidates:
                        auth_candidates.append(num)
            
            # Use the first candidate that's on the auth form page
            for candidate in auth_candidates:
                # Verify it appears near auth-related content
                candidate_pos = text.find(candidate)
                if candidate_pos >= 0:
                    surrounding_text = text[max(0, candidate_pos-100):candidate_pos+100].lower()
                    if any(word in surrounding_text for word in ['auth', 'approved', 'expire', 'treatment authorization']):
                        result["Auth #"] = candidate
                        result["Auth #_method"] = "fallback_search"
                        break
        
        # AGGRESSIVE FALLBACK for Patient ID - search for 9-digit participant IDs
        # Search ALL pages (all_text) since Patient ID may be on referral pages
        if not result["Patient ID"]:
            # Pattern 1: Look for Participant ID: followed by number (flexible spacing, including newlines)
            id_patterns = [
                r'Participant\s*ID:?\s*\n\s*(\d{9,12})',  # Label on one line, ID on next line
                r'Participant\s*ID[:\s]*(\d{9,12})',  # "Participant ID: 123456789" or "ParticipantID:123456789"
                r'Participant\s+ID[:\s]+(\d{9})',  # Strict space version
                r'(?:Participant|Member)\s*ID[:\s]*(\d{9})',
                r'Policy\s*#:?\s*(\d{9})',  # Sometimes listed as Policy #
                r'Policy\s*Number[:\s]*(\d{9})',  # "Policy Number:"
                r'Insurance[:\s]*.*?(\b10[01]\d{6}\b)',  # Near insurance label
                r'\n\s*(\d{9})\s*\n',  # Standalone 9-digit number on its own line
            ]
            for pattern in id_patterns:
                id_match = re.search(pattern, all_text, re.IGNORECASE)
                if id_match:
                    result["Patient ID"] = id_match.group(1)
                    result["Patient ID_method"] = "fallback_search"
                    break
            
            # Pattern 2: Look for 9-digit numbers starting with 100 or 101 (common PACE ID formats)
            if not result["Patient ID"]:
                pace_id_pattern = r'\b(10[01]\d{6})\b'  # 100xxxxxx or 101xxxxxx
                pace_matches = re.findall(pace_id_pattern, all_text)
                # Filter out numbers that are clearly dates or phone numbers
                for candidate in pace_matches:
                    # Verify it's near "Participant" or "ID" context OR just take first match
                    candidate_pos = all_text.find(candidate)
                    if candidate_pos >= 0:
                        surrounding = all_text[max(0, candidate_pos-100):candidate_pos+50].lower()
                        if any(word in surrounding for word in ['participant', 'member', 'id', 'patient', 'policy', 'insurance', 'pace']):
                            result["Patient ID"] = candidate
                            result["Patient ID_method"] = "fallback_search"
                            break
            
            # Pattern 3: If still no ID, look for any 9-digit number that appears in insurance/patient context
            if not result["Patient ID"]:
                # Find all 9-digit numbers
                all_9digit = re.findall(r'\b(\d{9})\b', all_text)
                # Filter out obvious non-IDs (phone numbers, dates, etc.)
                for candidate in all_9digit:
                    # Skip phone-like patterns (area codes starting with 6, 7, 8, 9)
                    if candidate[0] in '6789' and candidate[1:4].isdigit():
                        continue
                    # Skip if it looks like a date (has / or - nearby)
                    candidate_pos = all_text.find(candidate)
                    if candidate_pos >= 0:
                        nearby = all_text[max(0, candidate_pos-5):candidate_pos+15]
                        if '/' in nearby or '-' in nearby:
                            continue
                        # Good candidate - use it
                        result["Patient ID"] = candidate
                        result["Patient ID_method"] = "fallback_broad_search"
                        break
        
        # AGGRESSIVE FALLBACK for dates - if missing either date, search more broadly
        if not result["Date Approved"] or not result["Date Auth Expire"]:
            # Find all dates in MM/DD/YYYY format
            all_dates = re.findall(r'(\d{1,2})/(\d{1,2})/(\d{4})', text)
            if all_dates:
                # Convert to date objects and sort
                parsed_dates = []
                for m, d, y in all_dates:
                    try:
                        date_str = f"{m}/{d}/{y}"
                        # Skip dates that look like DOB (likely 1940s-1980s)
                        year = int(y)
                        if 2020 <= year <= 2030:  # Only recent dates
                            parsed_dates.append(date_str)
                    except:
                        continue
                
                # Remove duplicates while preserving order
                seen = set()
                unique_dates = []
                for d in parsed_dates:
                    if d not in seen:
                        seen.add(d)
                        unique_dates.append(d)
                
                # Assign first two dates as Approved and Expire
                if len(unique_dates) >= 1 and not result["Date Approved"]:
                    result["Date Approved"] = unique_dates[0]
                    result["Date Approved_method"] = "fallback_search"
                if len(unique_dates) >= 2 and not result["Date Auth Expire"]:
                    result["Date Auth Expire"] = unique_dates[1]
                    result["Date Auth Expire_method"] = "fallback_search"
        
        # ===== DATE VALIDATION AND CORRECTION =====
        # Fix common issues: same dates, expire before approved, invalid years
        result = self.validate_and_fix_dates(result)
        
        # ===== CLEAN UP EXTRACTED VALUES =====
        # Fix Patient ID that has OCR artifacts
        if result.get("Patient ID"):
            patient_id = result["Patient ID"]
            # Remove spaces within the ID
            patient_id = re.sub(r'\s+', '', patient_id)
            # Fix common OCR prefix errors (order matters - longer prefixes first)
            patient_id = re.sub(r'^fron\s*', '100', patient_id, flags=re.IGNORECASE)
            patient_id = re.sub(r'^froo', '100', patient_id, flags=re.IGNORECASE)
            patient_id = re.sub(r'^foo\s*', '100', patient_id, flags=re.IGNORECASE)
            patient_id = re.sub(r'^fot\s*', '100', patient_id, flags=re.IGNORECASE)
            patient_id = re.sub(r'^fro', '10', patient_id, flags=re.IGNORECASE)
            patient_id = re.sub(r'^fio', '10', patient_id, flags=re.IGNORECASE)
            patient_id = re.sub(r'^fo', '10', patient_id, flags=re.IGNORECASE)
            patient_id = re.sub(r'^fi', '1', patient_id, flags=re.IGNORECASE)
            patient_id = re.sub(r'^[1ifl][0o]0', '100', patient_id, flags=re.IGNORECASE)
            # Convert OCR letters to digits throughout
            patient_id = re.sub(r'[oO]', '0', patient_id)  # o -> 0
            patient_id = re.sub(r'[sS]', '8', patient_id)  # s -> 8
            patient_id = re.sub(r'[eE]', '6', patient_id)  # e -> 6
            patient_id = re.sub(r'[tT]', '7', patient_id)  # t -> 7
            patient_id = re.sub(r'[iIlL]', '1', patient_id)  # i/l -> 1
            patient_id = re.sub(r'[aA]', '4', patient_id)  # a -> 4
            patient_id = re.sub(r'[zZ]', '2', patient_id)  # z -> 2
            patient_id = re.sub(r'[cC]', '2', patient_id)  # c -> 2
            # Remove any remaining non-digit characters
            patient_id = re.sub(r'[^\d]', '', patient_id)
            result["Patient ID"] = patient_id if patient_id else None
        
        # Validate dates - reject if too long or contain non-date characters
        for date_field in ["Date Approved", "Date Auth Expire"]:
            if result.get(date_field):
                date_val = result[date_field]
                # Reject if too long (valid dates are like "01/15/2026" = max 10 chars)
                if len(date_val) > 12:
                    result[date_field] = None
                # Reject if contains letters (except OCR artifacts we can fix)
                elif re.search(r'[a-zA-Z]', date_val):
                    result[date_field] = None
        
        # Use filename fallback for patient name if not found
        if not result["Patient Name"] and filename_data["Patient Name"]:
            result["Patient Name"] = filename_data["Patient Name"]
            result["Patient Name_method"] = "filename"
        
        return result
    
    def process_folder(self, folder_path, progress_callback=None):
        """Process all PDFs in a folder."""
        folder = pathlib.Path(folder_path)
        pdf_files = sorted(folder.glob("*.pdf"))
        
        self.results = []
        total = len(pdf_files)
        
        for i, pdf_file in enumerate(pdf_files):
            try:
                result = self.process_pdf(pdf_file)
                self.results.append(result)
            except Exception as e:
                self.results.append({
                    "file": pdf_file.name,
                    "error": str(e),
                    "extracted_at": datetime.now().isoformat(),
                })
            
            if progress_callback:
                progress_callback(i + 1, total, pdf_file.name)
        
        return self.results

    def process_all_files(self, folder_path, progress_callback=None):
        """Process all supported files in a folder (PDF, CSV, XLSX, PNG, JPG).

        PDFs use the existing process_pdf pipeline.
        Other file types are routed through extraction.router which picks
        the best extraction method (structured parse, OCR, etc.).
        Results from the router are converted into the same dict format
        that process_pdf returns so format_results() works unchanged.
        """
        from extraction.router import classify_file, route_file
        from extraction.structured_extractor import extract_csv_rows
        from extraction.excel_extractor import extract_xlsx_rows
        from config import SUPPORTED_EXTENSIONS

        folder = pathlib.Path(folder_path)

        # Collect all supported files (recursive glob for subfolder support)
        all_files = []
        for ext in SUPPORTED_EXTENSIONS:
            all_files.extend(folder.rglob(f"*{ext}"))
        all_files = sorted(set(all_files))

        self.results = []
        total = len(all_files)

        for i, file_path in enumerate(all_files):
            file_type = classify_file(str(file_path))

            try:
                if file_type == "pdf":
                    # Use existing battle-tested PDF pipeline
                    result = self.process_pdf(file_path)
                    self.results.append(result)
                elif file_type in ("csv", "xlsx"):
                    # Structured files can contain multiple rows
                    if file_type == "csv":
                        rows = extract_csv_rows(str(file_path))
                    else:
                        rows = extract_xlsx_rows(str(file_path))
                    for er in rows:
                        self.results.append(self._extraction_result_to_dict(er))
                else:
                    # Images and anything else -> router
                    er = route_file(str(file_path))
                    self.results.append(self._extraction_result_to_dict(er))
            except Exception as e:
                self.results.append({
                    "file": file_path.name,
                    "error": str(e),
                    "extracted_at": datetime.now().isoformat(),
                })

            if progress_callback:
                progress_callback(i + 1, total, file_path.name)

        return self.results

    @staticmethod
    def _extraction_result_to_dict(er):
        """Convert an ExtractionResult into the dict format process_pdf returns."""
        result = {
            "file": pathlib.Path(er.source_file).name if er.source_file else "",
            "extracted_at": datetime.now().isoformat(),
            "extraction_method": er.extraction_method,
        }
        for field in FIELDS:
            result[field] = er.fields.get(field)
        if er.error:
            result["error"] = er.error
        if er.raw_text:
            result["raw_text_preview"] = er.raw_text[:800]
        if er.confidence:
            for field, conf in er.confidence.items():
                result[f"{field}_confidence"] = conf
        if er.warnings:
            for w in er.warnings:
                if "_method" not in w:
                    result.setdefault("warnings", []).append(w)
        return result
    
    def get_auth_type_from_filename(self, filename):
        """
        Determine authorization type from filename keywords.
        Returns "Skilled", "Unskilled", "Escort", or "" if unknown.
        """
        filename_lower = filename.lower()
        
        # Check for "Skilled" first (more specific - must not contain "unskilled")
        if "skilled" in filename_lower and "unskilled" not in filename_lower:
            return "Skilled"
        # Check for "Escort assistance"
        elif "escort" in filename_lower:
            return "Escort"
        # Check for "Unskilled"
        elif "unskilled" in filename_lower:
            return "Unskilled"
        else:
            return ""

    def get_cpt_codes(self, filename):
        """
        Determine CPT codes based on filename keywords.
        Returns a list of up to 5 CPT codes based on filename.
        - Unskilled or escort assistance -> [S9122]
        - Skilled -> [S9123, S9124]
        """
        filename_lower = filename.lower()
        
        # Check for "Skilled" first (more specific - must not contain "unskilled")
        if "skilled" in filename_lower and "unskilled" not in filename_lower:
            return ["S9123", "S9124", "", "", ""]
        # Check for "Escort assistance" or "Unskilled"
        elif "escort" in filename_lower or "unskilled" in filename_lower:
            return ["S9122", "", "", "", ""]
        else:
            # Default: no CPT codes
            return ["", "", "", "", ""]
    
    def format_results(self):
        """
        Format results into the desired output format.
        Creates ONE row per PDF file with 5 CPT code columns.
        Uses name matching against Caspio patient database for accurate names.
        """
        formatted_rows = []
        
        # Initialize name matcher to match against Caspio patient names
        name_matcher = PatientNameMatcher()
        
        # Track match statistics
        match_stats = {"matched": 0, "not_matched": 0, "no_name": 0}
        
        for result in self.results:
            # Skip errors
            if result.get("error"):
                continue
            
            filename = result.get("file", "")
            patient_name = result.get("Patient Name", "")
            
            # Clean patient name of common OCR artifacts
            if patient_name:
                patient_name = re.sub(r'^DOB:?\s*', '', patient_name, flags=re.IGNORECASE)
                patient_name = re.sub(r'[\[\]\|]', '', patient_name)
                # Remove leading/trailing non-alpha characters (except commas and hyphens which are name parts)
                patient_name = re.sub(r'^[^A-Za-z]+', '', patient_name)
                patient_name = re.sub(r'[^A-Za-z,\-\s]+$', '', patient_name)
                patient_name = patient_name.strip()
            
            extracted_name = patient_name  # Keep original for reference
            
            # Try to match against Caspio patient database
            last_name = ""
            first_name = ""
            clean_full_name = ""
            match_found = False
            
            if patient_name:
                # First, try to find a match in the Caspio patient database
                matched_last, matched_first, db_full_name, score = name_matcher.find_match(patient_name, threshold=0.7)
                
                if matched_last and score >= 0.7:
                    # Match found - use the Caspio database format
                    last_name = matched_last
                    first_name = matched_first
                    clean_full_name = db_full_name
                    match_found = True
                    match_stats["matched"] += 1
                else:
                    # No match found - clean the extracted name (remove middle names/initials)
                    # This handles new patients not yet in Caspio
                    last_name, first_name, clean_full_name = name_matcher.clean_name(patient_name)
                    match_stats["not_matched"] += 1
            else:
                match_stats["no_name"] += 1
            
            # Get CPT codes based on filename (5 codes, no mileage)
            cpt_codes = self.get_cpt_codes(filename)
            
            # Get auth type to determine Service_Type_Identifier
            auth_type = self.get_auth_type_from_filename(filename)
            service_type_identifier = "Escort" if auth_type == "Escort" else ""
            
            # Build warning summary from per-field fallback methods
            warnings = []
            for f in FIELDS:
                meth = result.get(f"{f}_method", "")
                if meth and "fallback" in meth:
                    warnings.append(f"{f}: {meth}")
            if result.get("warnings"):
                warnings.extend(result["warnings"])

            # Create ONE row per file
            formatted_rows.append({
                "Last Name": last_name,
                "First Name": first_name,
                "Patient Name": clean_full_name,  # Clean name (matched or cleaned)
                "Extracted Name": extracted_name,  # Original extracted name for reference
                "Patient ID": result.get("Patient ID", ""),
                "Service_Type_Identifier": service_type_identifier,
                "CPT Code": cpt_codes[0],
                "CPT Code 2": cpt_codes[1],
                "CPT Code 3": cpt_codes[2],
                "CPT Code 4": cpt_codes[3],
                "CPT Code 5": cpt_codes[4],
                "Auth Number": result.get("Auth #", ""),
                "Date Approved": result.get("Date Approved", ""),
                "Date Auth Expired": result.get("Date Auth Expire", ""),
                "Clearing House Payer ID": "98481",
                "Location ID": "CAENC",
                "Unique Payer Identifier": "Innermark : WayStar (98481)",
                "Extraction Method": result.get("extraction_method", ""),
                "Source File": filename,
                "Warnings": "; ".join(warnings) if warnings else "",
            })
        
        # Store match stats for later access
        self.last_match_stats = match_stats
        
        return formatted_rows

    def export_to_excel(self, output_path):
        """Export results to Excel with Raw Data and Formatted sheets."""
        if not pd:
            raise ImportError("pandas is required for Excel export")
        
        if not self.results:
            raise ValueError("No results to export")
        
        # === RAW DATA SHEET ===
        df_raw = pd.DataFrame(self.results)
        
        # Parse date columns
        for col in ["Date Approved", "Date Auth Expire"]:
            if col in df_raw.columns:
                df_raw[col] = pd.to_datetime(df_raw[col], errors="coerce")
        
        # Preferred column order for raw data
        ordered = [
            "file",
            "Patient Name", "Auth #", "Date Approved", "Date Auth Expire", "Patient ID",
            "extraction_method", "raw_text_preview", "extracted_at", "error"
        ]
        cols = [c for c in ordered if c in df_raw.columns] + [c for c in df_raw.columns if c not in ordered]
        df_raw = df_raw[cols]
        
        # === ERRORS SHEET ===
        # Filter for records with errors
        error_records = [r for r in self.results if r.get("error")]
        df_errors = pd.DataFrame(error_records) if error_records else pd.DataFrame(columns=["file", "error"])
        if not df_errors.empty:
            # Keep only relevant columns for errors sheet
            error_columns = ["file", "error", "extraction_method"]
            error_cols = [c for c in error_columns if c in df_errors.columns]
            df_errors = df_errors[error_cols]
        
        # === FORMATTED DATA SHEET ===
        formatted_data = self.format_results()
        df_formatted = pd.DataFrame(formatted_data)
        
        # Parse date columns in formatted data
        for col in ["Date Approved", "Date Auth Expired"]:
            if col in df_formatted.columns:
                df_formatted[col] = pd.to_datetime(df_formatted[col], errors="coerce")
        
        # Ensure column order (5 CPT code columns starting at position F)
        formatted_columns = [
            "Last Name", "First Name", "Patient Name", "Extracted Name",
            "Patient ID", "Service_Type_Identifier",
            "CPT Code", "CPT Code 2", "CPT Code 3", "CPT Code 4", "CPT Code 5",
            "Auth Number", "Date Approved", "Date Auth Expired",
            "Clearing House Payer ID", "Location ID", "Unique Payer Identifier"
        ]
        df_formatted = df_formatted[formatted_columns]
        
        # Write to Excel with both sheets
        with pd.ExcelWriter(output_path, engine="openpyxl", 
                           datetime_format="mm/dd/yyyy", 
                           date_format="mm/dd/yyyy") as writer:
            # Raw Data sheet
            df_raw.to_excel(writer, sheet_name="Raw Data", index=False)
            ws_raw = writer.sheets["Raw Data"]
            ws_raw.freeze_panes = "A2"
            
            # Auto-adjust column widths for raw data
            for col_cells in ws_raw.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        v = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(v))
                    except:
                        pass
                ws_raw.column_dimensions[col_letter].width = min(max_len + 2, 50)
            
            # Formatted sheet
            df_formatted.to_excel(writer, sheet_name="Formatted", index=False)
            ws_fmt = writer.sheets["Formatted"]
            ws_fmt.freeze_panes = "A2"
            
            # Auto-adjust column widths for formatted data
            for col_cells in ws_fmt.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        v = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(v))
                    except:
                        pass
                ws_fmt.column_dimensions[col_letter].width = min(max_len + 2, 50)
            
            # Errors sheet (files that had extraction errors)
            df_errors.to_excel(writer, sheet_name="Errors", index=False)
            ws_err = writer.sheets["Errors"]
            ws_err.freeze_panes = "A2"
            
            # Auto-adjust column widths for errors
            for col_cells in ws_err.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        v = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(v))
                    except:
                        pass
                ws_err.column_dimensions[col_letter].width = min(max_len + 2, 60)
        
        return output_path


class SplashScreen:
    """Loading splash screen shown before the main application launches."""

    STEPS = [
        (10,  "Loading interface..."),
        (25,  "Reading configuration..."),
        (45,  "Checking authorization models..."),
        (62,  "Connecting services..."),
        (78,  "Preparing workspace..."),
        (92,  "Almost ready..."),
        (100, "Launching Auth Radar..."),
    ]
    STEP_DELAY_MS = 380

    def __init__(self, win, on_complete):
        self.win = win
        self.on_complete = on_complete

        # Colors — matches Auth Radar theme
        self.bg     = '#0F172A'
        self.panel  = '#1E293B'
        self.accent = '#22C55E'
        self.text   = '#E5E7EB'
        self.muted  = '#94A3B8'
        self.border = '#334155'

        # Window chrome — frameless, centered
        w, h = 480, 340
        sw = win.winfo_screenwidth()
        sh = win.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.win.overrideredirect(True)
        self.win.geometry(f"{w}x{h}+{x}+{y}")
        self.win.configure(bg=self.border)   # 1-px border effect
        self.win.lift()
        self.win.attributes('-topmost', True)

        self._bar_widget = None
        self._track_widget = None
        self._status_var = tk.StringVar(value="Initializing...")

        self._build_ui()
        self._schedule_steps()

    # ------------------------------------------------------------------
    def _build_ui(self):
        # Inner frame (1-px border achieved by bg on outer win)
        inner = tk.Frame(self.win, bg=self.bg)
        inner.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)

        # ── Logo + branding ──────────────────────────────────────────
        top = tk.Frame(inner, bg=self.bg)
        top.pack(expand=True, fill=tk.BOTH, padx=50, pady=(44, 8))

        logo_path = APP_DIR / "Auth Radar Logo.png"
        self._logo_img = None
        if logo_path.exists():
            try:
                from PIL import Image, ImageTk
                raw = Image.open(str(logo_path)).convert("RGBA")
                bg_img = Image.new("RGBA", raw.size, self.bg)
                bg_img.alpha_composite(raw)
                img = bg_img.resize((200, 200), Image.LANCZOS).convert("RGB")
                self._logo_img = ImageTk.PhotoImage(img)
                tk.Label(top, image=self._logo_img, bg=self.bg, bd=0, highlightthickness=0).pack()
            except Exception:
                pass

        # ── Progress section ─────────────────────────────────────────
        progress_frame = tk.Frame(inner, bg=self.bg)
        progress_frame.pack(fill=tk.X, padx=50, pady=(4, 0))

        self._status_var.set("Initializing...")
        tk.Label(progress_frame, textvariable=self._status_var,
                 font=('Segoe UI', 8), bg=self.bg, fg=self.muted,
                 anchor=tk.W).pack(fill=tk.X, pady=(0, 5))

        # Track (background)
        track = tk.Frame(progress_frame, bg=self.panel, height=5)
        track.pack(fill=tk.X)
        track.pack_propagate(False)
        self._track_widget = track

        # Fill bar (foreground)
        bar = tk.Frame(track, bg=self.accent, height=5, width=0)
        bar.place(x=0, y=0, relheight=1, width=0)
        self._bar_widget = bar

        # ── Footer ───────────────────────────────────────────────────
        tk.Label(inner, text="Pace Healthcare  ·  Auth Radar",
                 font=('Segoe UI', 7), bg=self.bg, fg=self.border).pack(pady=(10, 14))

    # ------------------------------------------------------------------
    def _schedule_steps(self):
        delay = 120
        for progress, status in self.STEPS:
            self.win.after(delay, lambda p=progress, s=status: self._advance(p, s))
            delay += self.STEP_DELAY_MS
        self.win.after(delay + 200, self._finish)

    def _advance(self, value, status):
        self._status_var.set(status)
        self.win.update_idletasks()
        tw = self._track_widget.winfo_width()
        if tw > 1:
            self._bar_widget.place(x=0, y=0, relheight=1, width=int(tw * value / 100))

    def _finish(self):
        self.win.destroy()
        self.on_complete()


class LandingPage:
    """Landing page for selecting payer type."""
    
    # Define available payer types
    PAYER_TYPES = {
        "Pace": {
            "name": "Pace",
            "description": "PACE Treatment Authorization Forms",
            "icon": "📋"
        },
        # Future payers can be added here:
        # "Molina": {
        #     "name": "Molina",
        #     "description": "Molina Healthcare Authorization Forms",
        #     "icon": "📄"
        # },
    }
    
    def __init__(self, root):
        self.root = root
        self.root.title("Auth Radar")
        self.root.geometry("600x500")
        self.root.minsize(500, 400)
        self.selected_payer = None
        
        self.setup_ui()
        
    def setup_ui(self):
        """Set up the landing page UI."""
        # Auth Radar dark navy theme
        bg_color = '#0F172A'
        panel_color = '#1E293B'
        text_color = '#E5E7EB'
        accent_color = '#22C55E'
        text_muted = '#94A3B8'
        border_color = '#334155'
        
        self.root.configure(bg=bg_color)
        
        # Main container
        main_frame = tk.Frame(self.root, bg=bg_color)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=40, pady=40)
        
        # Auth Radar branding — logo only
        self._landing_logo = None
        logo_path = APP_DIR / "Auth Radar Logo.png"
        if logo_path.exists():
            try:
                from PIL import Image, ImageTk
                raw = Image.open(str(logo_path)).convert("RGBA")
                bg_img = Image.new("RGBA", raw.size, bg_color)
                bg_img.alpha_composite(raw)
                img = bg_img.resize((180, 180), Image.LANCZOS).convert("RGB")
                self._landing_logo = ImageTk.PhotoImage(img)
                tk.Label(main_frame, image=self._landing_logo, bg=bg_color, bd=0, highlightthickness=0).pack(pady=(20, 20))
            except Exception:
                pass
        
        # Instruction
        instruction_label = tk.Label(main_frame, text="Select a payer type to begin:",
                                     font=('Segoe UI', 10),
                                     bg=bg_color, fg=text_color)
        instruction_label.pack(pady=(0, 20))
        
        # Payer buttons container
        buttons_frame = tk.Frame(main_frame, bg=bg_color)
        buttons_frame.pack(fill=tk.X, pady=10)
        
        # Create a button for each payer type
        for payer_key, payer_info in self.PAYER_TYPES.items():
            self.create_payer_button(buttons_frame, payer_key, payer_info, bg_color, panel_color, text_color, accent_color)
        
        # Coming soon placeholder for future payers
        coming_soon_frame = tk.Frame(main_frame, bg=bg_color)
        coming_soon_frame.pack(fill=tk.X, pady=30)
        
        coming_soon_label = tk.Label(coming_soon_frame, 
                                     text="More payers coming soon...",
                                     font=('Segoe UI', 9, 'italic'),
                                     bg=bg_color, fg=text_muted)
        coming_soon_label.pack()
    
    def create_payer_button(self, parent, payer_key, payer_info, bg_color, panel_color, text_color, accent_color):
        """Create a styled button for a payer type."""
        # Button frame with visible border effect
        btn_container = tk.Frame(parent, bg='#334155', padx=1, pady=1)
        btn_container.pack(fill=tk.X, pady=8)
        
        # Inner button area
        btn_frame = tk.Frame(btn_container, bg=panel_color, cursor="hand2")
        btn_frame.pack(fill=tk.X)
        
        # Content frame
        content = tk.Frame(btn_frame, bg=panel_color)
        content.pack(fill=tk.X, padx=25, pady=20)
        
        # Icon and name
        name_frame = tk.Frame(content, bg=panel_color)
        name_frame.pack(fill=tk.X)
        
        icon_label = tk.Label(name_frame, text=payer_info['icon'],
                              font=('Segoe UI', 18),
                              bg=panel_color, fg=text_color)
        icon_label.pack(side=tk.LEFT)
        
        name_label = tk.Label(name_frame, text=payer_info['name'],
                              font=('Segoe UI', 14, 'bold'),
                              bg=panel_color, fg=text_color)
        name_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # Arrow indicator
        arrow_label = tk.Label(name_frame, text="→",
                               font=('Segoe UI', 14),
                               bg=panel_color, fg='#a0a0a0')
        arrow_label.pack(side=tk.RIGHT)
        
        # Hover colors - subtle lift effect
        hover_bg = '#334155'
        
        # Hover effects
        def on_enter(e):
            btn_frame.configure(bg=hover_bg)
            content.configure(bg=hover_bg)
            name_frame.configure(bg=hover_bg)
            for widget in [icon_label, name_label, arrow_label]:
                widget.configure(bg=hover_bg)
            arrow_label.configure(fg=accent_color)
            btn_container.configure(bg=accent_color)
        
        def on_leave(e):
            btn_frame.configure(bg=panel_color)
            content.configure(bg=panel_color)
            name_frame.configure(bg=panel_color)
            for widget in [icon_label, name_label, arrow_label]:
                widget.configure(bg=panel_color)
            arrow_label.configure(fg='#94A3B8')
            btn_container.configure(bg='#334155')
        
        def on_click(e):
            self.select_payer(payer_key)
        
        # Bind events to all clickable elements
        for widget in [btn_frame, content, name_frame, icon_label, name_label, arrow_label]:
            widget.bind('<Enter>', on_enter)
            widget.bind('<Leave>', on_leave)
            widget.bind('<Button-1>', on_click)
    
    def select_payer(self, payer_key):
        """Handle payer selection."""
        self.selected_payer = payer_key
        # Destroy landing page and launch main app
        for widget in self.root.winfo_children():
            widget.destroy()
        # Resize window for main app
        self.root.geometry("1100x800")
        self.root.minsize(900, 700)
        # Launch main app with selected payer
        AuthExtractorApp(self.root, payer_type=payer_key)


class AuthExtractorApp:
    """GUI Application for PDF extraction."""
    
    def __init__(self, root, payer_type="Pace"):
        self.root = root
        self.payer_type = payer_type
        self.root.title(f"Auth Radar — {payer_type}")
        self.root.geometry("1100x800")
        self.root.minsize(900, 700)
        
        self.extractor = PDFExtractor()
        self.input_folder = tk.StringVar()
        self.output_file = tk.StringVar()
        self.status_text = tk.StringVar(value="Ready")
        self.is_processing = False
        
        # File Finder variables
        self.finder_source = tk.StringVar()
        self.finder_dest = tk.StringVar()
        
        # File Finder results storage (for Excel export)
        self.finder_found_matches = []  # (name, auth_type, filename, last_dos)
        self.finder_not_found = []  # (name, auth_type, reason, last_dos)
        self.finder_duplicate_imports = []  # Duplicate names from import
        self.finder_original_count = 0  # Original count before dedup

        # Search tab variables
        self.search_auth_term_var = tk.StringVar()  # Authorization search term
        self.search_patient_term_var = tk.StringVar()  # Patient search term
        self.search_auths_results = []  # Authorization search results
        self.search_patients_results = []  # Patient search results
        
        # In-app results storage for inline editing
        self.current_results_df = None  # DataFrame of formatted results
        self.results_columns = []  # Column names for results table
        self.editing_cell = None  # Track cell being edited (item, column)

        # Dropbox integration
        self.dropbox_service = None  # DropboxService instance
        self.dropbox_files = []     # List of FileMetadata from last listing
        self.dropbox_folder_var = tk.StringVar(value="")  # Selected Dropbox folder path
        self.dropbox_keyword_var = tk.StringVar(value="")  # filename keyword filter
        self.dropbox_name_filter_var = tk.StringVar(value="")  # patient name in filename
        self.dropbox_name_filter_list = []  # list of patient names
        self.dropbox_status_var = tk.StringVar(value="(not connected)")
        self.dropbox_name_count_var = tk.StringVar(value="(none)")
        self.finder_source_type = tk.StringVar(value="local")  # "local" or "dropbox"
        
        self.setup_ui()
        self.check_ocr_status()
        
    def check_ocr_status(self):
        """Check and display OCR status."""
        if OCR_AVAILABLE and self.extractor.tesseract_path:
            self.log("✅ OCR Ready (Tesseract found)")
            self.log(f"   Path: {self.extractor.tesseract_path}")
            if POPPLER_PATH and POPPLER_PATH.exists():
                self.log(f"   Poppler: {POPPLER_PATH}")
            else:
                self.log(f"⚠️ Poppler not found at expected location")
        elif OCR_AVAILABLE:
            self.log("⚠️ OCR libraries loaded but Tesseract not found")
            self.log("   Install Tesseract: winget install UB-Mannheim.TesseractOCR")
        else:
            self.log("⚠️ OCR not available - install Tesseract for scanned PDFs")
            self.log("   Run: winget install UB-Mannheim.TesseractOCR")
        self.log("")
    
    def return_to_landing(self):
        """Return to the landing page to select a different payer."""
        # Destroy all widgets
        for widget in self.root.winfo_children():
            widget.destroy()
        # Launch landing page
        LandingPage(self.root)
        
    def setup_ui(self):
        """Set up the user interface."""
        # Auth Radar dark navy theme
        self.colors = {
            'bg': '#0F172A',            # Dark navy background
            'panel': '#1E293B',          # Slate panel/card
            'text': '#E5E7EB',           # Light gray text
            'text_light': '#94A3B8',     # Muted slate text
            'accent': '#22C55E',         # Radar green - primary
            'accent_light': '#4ADE80',   # Lighter green hover
            'accent_dark': '#16A34A',    # Darker green pressed
            'accent_blue': '#06B6D4',    # Teal/cyan - secondary
            'accent_blue_light': '#22D3EE',
            'accent_blue_dark': '#0891B2',
            'border': '#334155',         # Slate border
            'success': '#22C55E',        # Radar green - valid
            'warning': '#F59E0B',        # Amber - warnings
            'error': '#EF4444',          # Alert red - invalid
            'row_alt': '#162033',        # Alternating row
            'selected': '#1D4ED8',       # Blue selection
            'entry_bg': '#1E293B',       # Entry field background
        }
        
        # Apply clean styling
        self.apply_theme()
        
        # Header with Auth Radar branding
        header_frame = ttk.Frame(self.root)
        header_frame.pack(fill=tk.X, padx=20, pady=(15, 10))
        
        # Left side: logo + title
        brand_frame = tk.Frame(header_frame, bg=self.colors['bg'])
        brand_frame.pack(side=tk.LEFT)
        
        # Load logo if available
        self._header_logo = None
        logo_path = APP_DIR / "Auth Radar Logo.png"
        if logo_path.exists():
            try:
                from PIL import Image, ImageTk
                raw = Image.open(str(logo_path)).convert("RGBA")
                bg_img = Image.new("RGBA", raw.size, self.colors['bg'])
                bg_img.alpha_composite(raw)
                img = bg_img.resize((48, 48), Image.LANCZOS).convert("RGB")
                self._header_logo = ImageTk.PhotoImage(img)
                logo_lbl = tk.Label(brand_frame, image=self._header_logo, bg=self.colors['bg'], bd=0, highlightthickness=0)
                logo_lbl.pack(side=tk.LEFT, padx=(0, 8))
            except Exception:
                pass
        
        # Payer indicator
        payer_lbl = ttk.Label(header_frame, text=f"  •  {self.payer_type}",
                              style='Desc.TLabel')
        payer_lbl.pack(side=tk.LEFT, padx=(8, 0))
        
        # Right side: status indicator + back button
        right_frame = ttk.Frame(header_frame)
        right_frame.pack(side=tk.RIGHT)
        
        self.radar_status_label = tk.Label(right_frame, text="● ONLINE",
                                           font=('Segoe UI', 8, 'bold'),
                                           bg=self.colors['bg'], fg='#22C55E')
        self.radar_status_label.pack(side=tk.LEFT, padx=(0, 12))
        
        back_btn = ttk.Button(right_frame, text="← Change Payer", 
                              command=self.return_to_landing)
        back_btn.pack(side=tk.LEFT)
        
        # Create notebook for tabs
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 20))
        
        # ===== MAIN TAB 1: Auth Management =====
        self.auth_management_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.auth_management_tab, text="   Auth Management   ")
        
        # Create nested notebook inside Auth Management
        self.auth_notebook = ttk.Notebook(self.auth_management_tab, style='Sub.TNotebook')
        self.auth_notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Sub-Tab 1: File Finder
        self.finder_tab = ttk.Frame(self.auth_notebook)
        self.auth_notebook.add(self.finder_tab, text="  1. Find Auth PDFs  ")
        self.setup_finder_tab()
        
        # Sub-Tab 2: Extract PDFs
        self.extractor_tab = ttk.Frame(self.auth_notebook)
        self.auth_notebook.add(self.extractor_tab, text="  2. Download PDFs  ")
        self.setup_extractor_tab()
        
        # Sub-Tab 3: Review & Edit
        self.review_tab = ttk.Frame(self.auth_notebook)
        self.auth_notebook.add(self.review_tab, text="   Review & Edit   ")
        self.setup_review_tab()

        # Sub-Tab 4: Email
        self.email_tab = ttk.Frame(self.auth_notebook)
        self.auth_notebook.add(self.email_tab, text="   Email   ")
        self.setup_email_tab()

        # ===== MAIN TAB 2: Search =====
        self.search_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.search_tab, text="   Search   ")
        self.setup_search_tab()
    
    def apply_theme(self):
        """Apply modern dark theme styling to the application."""
        style = ttk.Style()
        style.theme_use('clam')
        
        c = self.colors
        
        # Root window
        self.root.configure(bg=c['bg'])
        
        # Main Notebook tabs - larger, more prominent
        style.configure('TNotebook', background=c['bg'], borderwidth=0, tabmargins=[0, 0, 0, 0])
        style.configure('TNotebook.Tab',
            background=c['panel'],
            foreground=c['text_light'],
            padding=[40, 14],
            font=('Segoe UI', 11, 'bold'))
        style.map('TNotebook.Tab',
            background=[('selected', c['accent'])],
            foreground=[('selected', '#ffffff')],
            padding=[('selected', [40, 14])])
        
        # Sub-notebook tabs - smaller, nested look
        style.configure('Sub.TNotebook', background=c['bg'], borderwidth=0, tabmargins=[0, 0, 0, 0])
        style.configure('Sub.TNotebook.Tab',
            background=c['panel'],
            foreground=c['text_light'],
            padding=[20, 8],
            font=('Segoe UI', 9))
        style.map('Sub.TNotebook.Tab',
            background=[('selected', c['accent_blue'])],
            foreground=[('selected', '#ffffff')],
            padding=[('selected', [20, 8])])
        
        # Frames - dark background
        style.configure('TFrame', background=c['bg'])
        
        # LabelFrames - subtle borders
        style.configure('TLabelframe', background=c['bg'], bordercolor=c['border'])
        style.configure('TLabelframe.Label',
            background=c['bg'],
            foreground=c['text_light'],
            font=('Segoe UI', 9, 'bold'))
        
        # Labels
        style.configure('TLabel',
            background=c['bg'],
            foreground=c['text'],
            font=('Segoe UI', 9))
        
        style.configure('Title.TLabel',
            background=c['bg'],
            foreground=c['text'],
            font=('Segoe UI', 18, 'bold'))
        
        style.configure('Desc.TLabel',
            background=c['bg'],
            foreground=c['text_light'],
            font=('Segoe UI', 9))
        
        # Regular buttons - subtle dark look
        style.configure('TButton',
            background=c['panel'],
            foreground=c['text'],
            bordercolor=c['border'],
            padding=[14, 8],
            font=('Segoe UI', 9))
        style.map('TButton',
            background=[('active', c['border']), ('pressed', c['bg'])],
            foreground=[('active', c['text']), ('pressed', c['text'])],
            bordercolor=[('active', c['accent'])])
        
        # Action buttons - radar green accent
        style.configure('Action.TButton',
            background=c['accent'],
            foreground='#ffffff',
            padding=[18, 10],
            font=('Segoe UI', 10, 'bold'))
        style.map('Action.TButton',
            background=[('active', c['accent_light']), ('pressed', c['accent_dark'])],
            foreground=[('active', '#ffffff'), ('pressed', '#ffffff')])
        
        # Blue action buttons - for informational actions
        style.configure('Blue.TButton',
            background=c['accent_blue'],
            foreground='#ffffff',
            padding=[18, 10],
            font=('Segoe UI', 10, 'bold'))
        style.map('Blue.TButton',
            background=[('active', c['accent_blue_light']), ('pressed', c['accent_blue_dark'])],
            foreground=[('active', '#ffffff'), ('pressed', '#ffffff')])
        
        # Entries - dark fields
        style.configure('TEntry',
            fieldbackground=c['entry_bg'],
            foreground=c['text'],
            insertcolor=c['text'],
            bordercolor=c['border'],
            padding=8)
        
        # Combobox - dark, subtle
        style.configure('TCombobox',
            fieldbackground=c['entry_bg'],
            background=c['panel'],
            foreground=c['text'],
            arrowcolor=c['text_light'],
            bordercolor=c['border'],
            padding=5)
        style.map('TCombobox',
            fieldbackground=[('readonly', c['entry_bg'])],
            selectbackground=[('readonly', c['accent'])],
            foreground=[('readonly', c['text'])])
        
        # Treeview - dark modern theme
        style.configure('Treeview',
            background=c['panel'],
            foreground=c['text'],
            fieldbackground=c['panel'],
            bordercolor=c['border'],
            font=('Segoe UI', 9),
            rowheight=28)
        style.configure('Treeview.Heading',
            background=c['bg'],
            foreground=c['text_light'],
            font=('Segoe UI', 9, 'bold'),
            bordercolor=c['border'],
            padding=[5, 8])
        style.map('Treeview',
            background=[('selected', c['accent'])],
            foreground=[('selected', '#ffffff')])
        
        # Progressbar - radar green
        style.configure('TProgressbar',
            background=c['accent'],
            troughcolor=c['panel'],
            bordercolor=c['border'])
        
        # Scanning progressbar - pulsing blue for processing
        style.configure('Scan.TProgressbar',
            background=c['accent_blue'],
            troughcolor=c['panel'],
            bordercolor=c['border'])
        
        # Scrollbar - dark theme
        style.configure('TScrollbar',
            background=c['panel'],
            troughcolor=c['bg'],
            bordercolor=c['border'],
            arrowcolor=c['text_light'])
        style.map('TScrollbar',
            background=[('active', c['accent'])])
        
        # Checkbutton - dark theme
        style.configure('TCheckbutton',
            background=c['bg'],
            foreground=c['text'])
        style.map('TCheckbutton',
            background=[('active', c['bg'])],
            foreground=[('active', c['accent'])])
        
        # Radiobutton - dark theme
        style.configure('TRadiobutton',
            background=c['bg'],
            foreground=c['text'])
        style.map('TRadiobutton',
            background=[('active', c['bg'])],
            foreground=[('active', c['accent'])])
        
        # PanedWindow
        style.configure('TPanedwindow', background=c['bg'])
        
        # Separator
        style.configure('TSeparator', background=c['border'])
    
    def style_popup(self, popup):
        """Apply dark theme styling to a popup window."""
        c = self.colors
        popup.configure(bg=c['bg'])
        
    def setup_extractor_tab(self):
        """Set up the extractor tab UI."""
        # Main frame with padding
        main_frame = ttk.Frame(self.extractor_tab, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Description
        desc_label = ttk.Label(main_frame, 
            text="Download and unlock authorization PDFs so you can review them or upload to ChatGPT for extraction.",
            style='Desc.TLabel')
        desc_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Source selection frame
        source_frame = ttk.LabelFrame(main_frame, text="Step 1: Select PDF Source", padding="10")
        source_frame.pack(fill=tk.X, pady=5)
        
        # Radio buttons for source type
        self.extract_source_type = tk.StringVar(value="manual")
        
        # Option 1: Use File Finder results
        finder_option_frame = ttk.Frame(source_frame)
        finder_option_frame.pack(fill=tk.X, pady=2)
        
        self.finder_radio = ttk.Radiobutton(finder_option_frame, text="Use File Finder results",
                                             variable=self.extract_source_type, value="finder",
                                             command=self.update_source_selection)
        self.finder_radio.pack(side=tk.LEFT)
        
        self.finder_files_count_var = tk.StringVar(value="(No files found yet - run File Finder first)")
        self.finder_files_label = ttk.Label(finder_option_frame, textvariable=self.finder_files_count_var,
                                            font=("Segoe UI", 9, "italic"), foreground="gray")
        self.finder_files_label.pack(side=tk.LEFT, padx=(10, 0))
        
        # Option 2: Manual folder selection
        manual_option_frame = ttk.Frame(source_frame)
        manual_option_frame.pack(fill=tk.X, pady=2)
        
        self.manual_radio = ttk.Radiobutton(manual_option_frame, text="Select folder manually",
                                             variable=self.extract_source_type, value="manual",
                                             command=self.update_source_selection)
        self.manual_radio.pack(side=tk.LEFT)
        
        # Manual folder selection (shown when manual is selected)
        self.manual_folder_frame = ttk.Frame(source_frame)
        self.manual_folder_frame.pack(fill=tk.X, pady=(5, 0))
        
        input_entry = ttk.Entry(self.manual_folder_frame, textvariable=self.input_folder, width=60)
        input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(20, 10))
        
        browse_btn = ttk.Button(self.manual_folder_frame, text="📁 Browse Folder...", command=self.browse_input)
        browse_btn.pack(side=tk.RIGHT)

        # Output file selection
        output_frame = ttk.LabelFrame(main_frame, text="Step 2: Choose Output Folder", padding="10")
        output_frame.pack(fill=tk.X, pady=5)
        
        output_entry = ttk.Entry(output_frame, textvariable=self.output_file, width=60)
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        save_btn = ttk.Button(output_frame, text="Save As...", command=self.browse_output)
        save_btn.pack(side=tk.RIGHT)

        # Status bar — pack at bottom first so it's always visible
        status_bar = ttk.Label(main_frame, textvariable=self.status_text,
                               relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, pady=(4, 0), side=tk.BOTTOM)

        # Buttons — pack at bottom before progress so they're always visible
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(6, 2), side=tk.BOTTOM)

        self.extract_btn = ttk.Button(button_frame, text="� Download & Unlock PDFs",
                                      command=self.start_extraction, style='Action.TButton')
        self.extract_btn.pack(side=tk.LEFT, padx=5)

        self.test_btn = ttk.Button(button_frame, text="🔍 Test Single PDF",
                                   command=self.test_single_pdf)
        self.test_btn.pack(side=tk.LEFT, padx=5)

        self.sync_btn = ttk.Button(button_frame, text="🔄 Sync Patient Names",
                                   command=self.sync_patient_names_from_caspio)
        self.sync_btn.pack(side=tk.LEFT, padx=5)

        self.goto_review_btn = ttk.Button(button_frame, text="📋 Go to Review Tab",
                                          command=lambda: self.auth_notebook.select(self.review_tab))
        self.goto_review_btn.pack(side=tk.LEFT, padx=5)
        
        # Progress section — expands to fill remaining space between steps and buttons
        progress_frame = ttk.LabelFrame(main_frame, text="Progress & Log", padding="10")
        progress_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        self.progress_bar = ttk.Progressbar(progress_frame, mode="determinate")
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        
        # Log text area - dark terminal style
        self.log_text = tk.Text(progress_frame, height=6, state=tk.DISABLED, 
                                font=("Consolas", 9),
                                bg='#0B1222',
                                fg='#22C55E',
                                insertbackground='#22C55E',
                                relief='solid',
                                borderwidth=1,
                                padx=8, pady=8)
        scrollbar = ttk.Scrollbar(progress_frame, command=self.log_text.yview)
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    def on_results_double_click(self, event):
        """Handle double-click to edit a cell in results table."""
        # Identify the row and column clicked
        region = self.results_tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        
        item = self.results_tree.identify_row(event.y)
        column = self.results_tree.identify_column(event.x)
        
        if not item or not column:
            return
        
        # Get column index (column is like "#1", "#2", etc.)
        col_idx = int(column[1:]) - 1
        if col_idx < 0 or col_idx >= len(self.results_columns):
            return
        
        col_name = self.results_columns[col_idx]
        
        # Get cell bounding box
        bbox = self.results_tree.bbox(item, column)
        if not bbox:
            return
        
        x, y, width, height = bbox
        
        # Get current value
        current_values = self.results_tree.item(item, "values")
        current_value = current_values[col_idx] if col_idx < len(current_values) else ""
        
        # Create entry widget for editing
        self.edit_entry = ttk.Entry(self.results_tree, width=width//8)
        self.edit_entry.insert(0, current_value)
        self.edit_entry.select_range(0, tk.END)
        
        # Place the entry over the cell
        self.edit_entry.place(x=x, y=y, width=width, height=height)
        self.edit_entry.focus_set()
        
        # Store editing context
        self.editing_cell = (item, col_idx, col_name)
        
        # Bind events
        self.edit_entry.bind("<Return>", self.save_cell_edit)
        self.edit_entry.bind("<Escape>", self.cancel_cell_edit)
        self.edit_entry.bind("<FocusOut>", self.save_cell_edit)
    
    def save_cell_edit(self, event=None):
        """Save the edited cell value."""
        if not hasattr(self, 'edit_entry') or not self.editing_cell:
            return
        
        item, col_idx, col_name = self.editing_cell
        new_value = self.edit_entry.get()
        
        # Get old value for audit trail
        current_values = list(self.results_tree.item(item, "values"))
        old_value = current_values[col_idx] if col_idx < len(current_values) else ""
        
        # Update treeview
        current_values[col_idx] = new_value
        self.results_tree.item(item, values=current_values)
        
        # Update the DataFrame
        if self.current_results_df is not None:
            # Find the row index from item (item is like "I001", "I002", etc.)
            try:
                row_idx = self.results_tree.index(item)
                if col_name in self.current_results_df.columns:
                    self.current_results_df.at[row_idx, col_name] = new_value
            except Exception:
                pass
        
        # Audit log the edit (non-blocking)
        if str(old_value) != str(new_value):
            try:
                # Try to get source file from the row for context
                source_file = ""
                src_idx = None
                if "Source File" in self.results_columns:
                    src_idx = self.results_columns.index("Source File")
                if src_idx is not None and src_idx < len(current_values):
                    source_file = current_values[src_idx]
                
                from audit.logger import AuditLogger
                AuditLogger().log_review(
                    source_file=source_file or "unknown",
                    reviewer="user",
                    edits={col_name: {"old": str(old_value), "new": str(new_value)}},
                )
            except Exception:
                pass  # audit failure should not block editing
        
        # Cleanup
        self.edit_entry.destroy()
        self.editing_cell = None
    
    def cancel_cell_edit(self, event=None):
        """Cancel cell editing."""
        if hasattr(self, 'edit_entry'):
            self.edit_entry.destroy()
        self.editing_cell = None
    
    def populate_results_table(self, df, errors_df=None, total_files=0):
        """Populate the results table with DataFrame data and highlight expired auths."""
        # Clear existing items
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        
        # Clear errors table too
        if hasattr(self, 'errors_tree'):
            for item in self.errors_tree.get_children():
                self.errors_tree.delete(item)
        
        successful_count = 0
        expired_count = 0
        error_count = 0
        
        if df is not None and not df.empty:
            # Store the DataFrame
            self.current_results_df = df.copy()
            
            # Map DataFrame columns to display columns (direct 1:1 mapping)
            col_mapping = {
                "Last Name": "Last Name",
                "First Name": "First Name",
                "Patient Name": "Patient Name",
                "Patient ID": "Patient ID",
                "Auth Number": "Auth Number",
                "Date Approved": "Date Approved",
                "Date Auth Expired": "Date Auth Expired",
                "Last DOS": "Last DOS",
                "CPT Code": "CPT Code",
                "CPT Code 2": "CPT Code 2",
                "CPT Code 3": "CPT Code 3",
                "CPT Code 4": "CPT Code 4",
                "CPT Code 5": "CPT Code 5",
                "Clearing House Payer ID": "Clearing House Payer ID",
                "Location ID": "Location ID",
                "Unique Payer Identifier": "Unique Payer Identifier",
                "Extraction Method": "Extraction Method",
                "Source File": "Source File",
                "Warnings": "Warnings",
            }
            
            # Parse date columns for comparison
            date_expired_col = None
            last_dos_col = None
            
            if "Date Auth Expired" in df.columns:
                date_expired_col = pd.to_datetime(df["Date Auth Expired"], errors="coerce")
            if "Last DOS" in df.columns:
                last_dos_col = pd.to_datetime(df["Last DOS"], errors="coerce")
            
            # Insert rows
            for idx, row in df.iterrows():
                values = []
                for col in self.results_columns:
                    df_col = col_mapping.get(col, col)
                    if df_col in df.columns:
                        val = row[df_col]
                        # Format dates
                        if pd and pd.notna(val) and hasattr(val, 'strftime'):
                            val = val.strftime("%m/%d/%Y")
                        elif pd and pd.isna(val):
                            val = ""
                        values.append(str(val) if val else "")
                    else:
                        values.append("")
                
                # Determine row tag: expired takes priority, then method-based, then warnings
                is_expired = False
                if date_expired_col is not None and last_dos_col is not None:
                    expire_val = date_expired_col.iloc[idx] if idx < len(date_expired_col) else None
                    dos_val = last_dos_col.iloc[idx] if idx < len(last_dos_col) else None
                    if pd.notna(expire_val) and pd.notna(dos_val) and dos_val > expire_val:
                        is_expired = True
                        expired_count += 1
                
                if is_expired:
                    tag = "expired"
                else:
                    # Color by extraction method
                    ext_method = ""
                    if "Extraction Method" in df.columns:
                        ext_method = str(row.get("Extraction Method", "")).lower()
                    has_warnings = ""
                    if "Warnings" in df.columns:
                        has_warnings = str(row.get("Warnings", ""))
                    
                    if ext_method in ("csv_parse", "excel_parse"):
                        tag = "structured"
                    elif "ocr" in ext_method:
                        tag = "ocr"
                    elif has_warnings:
                        tag = "warning"
                    else:
                        tag = "normal"
                
                self.results_tree.insert("", tk.END, values=values, tags=(tag,))
                successful_count += 1
        else:
            self.current_results_df = None
        
        # Populate errors table
        if errors_df is not None and not errors_df.empty and hasattr(self, 'errors_tree'):
            for _, row in errors_df.iterrows():
                filename = str(row.get("file", row.get("File Name", "Unknown")))
                error_msg = str(row.get("error", row.get("Error", "Unknown error")))
                # Determine auth type from filename
                auth_type = self.get_auth_type_from_filename(filename)
                self.errors_tree.insert("", tk.END, values=(filename, error_msg, auth_type), tags=("error",))
                error_count += 1
        
        # Update counts
        if total_files == 0:
            total_files = successful_count + error_count
        
        valid_count = successful_count - expired_count
        self.results_count_var.set(f"{successful_count} records ({valid_count} valid, {expired_count} expired)")
        
        if hasattr(self, 'valid_count_var'):
            self.valid_count_var.set(f"✓ {valid_count} valid auths ready for upload")
        
        if hasattr(self, 'errors_count_var'):
            self.errors_count_var.set(f"{error_count} errors")
        
        if hasattr(self, 'view_errors_btn'):
            btn_text = f"⚠️ View Errors ({error_count})" if error_count > 0 else "⚠️ View Errors"
            self.view_errors_btn.config(text=btn_text)
        
        if hasattr(self, 'review_summary_var'):
            self.review_summary_var.set(
                f"Total Files: {total_files}  |  Successful: {successful_count}  |  Valid: {valid_count}  |  "
                f"Expired: {expired_count}  |  Errors: {error_count}"
            )
    
    def setup_review_tab(self):
        """Set up the Review & Edit tab UI."""
        main_frame = ttk.Frame(self.review_tab, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Description
        desc_label = ttk.Label(main_frame, 
            text="Review extracted data, edit cells (double-click), and upload to Caspio",
            style='Desc.TLabel')
        desc_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Summary counts frame
        summary_frame = ttk.Frame(main_frame)
        summary_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.review_summary_var = tk.StringVar(value="No extraction data yet. Run extraction from Extract PDFs tab.")
        ttk.Label(summary_frame, textvariable=self.review_summary_var, 
                  font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        
        # === SUCCESSFUL EXTRACTIONS (full height) ===
        success_frame = ttk.LabelFrame(main_frame, text="Successful Extractions - Ready for Upload (double-click to edit)", padding="10")
        success_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))
        
        # Toolbar for Add/Edit/Delete at top of successful extractions
        success_toolbar = ttk.Frame(success_frame)
        success_toolbar.pack(fill=tk.X, pady=(0, 5))
        
        self.add_entry_btn = ttk.Button(success_toolbar, text="➕ Add", 
                                         command=self.show_add_entry_dialog)
        self.add_entry_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.edit_entry_btn = ttk.Button(success_toolbar, text="✏️ Edit", 
                                          command=self.edit_selected_entry)
        self.edit_entry_btn.pack(side=tk.LEFT, padx=5)
        
        self.delete_row_btn = ttk.Button(success_toolbar, text="🗑️ Delete", 
                                          command=self.delete_selected_row)
        self.delete_row_btn.pack(side=tk.LEFT, padx=5)
        
        self.mark_error_btn = ttk.Button(success_toolbar, text="⚠️ Mark as Error", 
                                          command=self.mark_selected_as_error)
        self.mark_error_btn.pack(side=tk.LEFT, padx=5)
        
        # Create treeview with scrollbars
        tree_container = ttk.Frame(success_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)
        
        # Define columns for editable results - ALL fields for Caspio mapping
        self.results_columns = [
            "Last Name", "First Name", "Patient Name", "Patient ID", "Service_Type_Identifier",
            "Auth Number", "Date Approved", "Date Auth Expired", "Last DOS",
            "CPT Code", "CPT Code 2", "CPT Code 3", "CPT Code 4", "CPT Code 5",
            "Clearing House Payer ID", "Location ID", "Unique Payer Identifier",
            "Extraction Method", "Source File", "Warnings"
        ]
        
        self.results_tree = ttk.Treeview(tree_container, columns=self.results_columns, show="headings", height=10)
        
        # Configure columns
        col_widths = {
            "Last Name": 110, "First Name": 90, "Patient Name": 140, "Patient ID": 70, "Service_Type_Identifier": 85,
            "Auth Number": 90, "Date Approved": 95, "Date Auth Expired": 105, "Last DOS": 85,
            "CPT Code": 70, "CPT Code 2": 75, "CPT Code 3": 75, "CPT Code 4": 75, "CPT Code 5": 75,
            "Clearing House Payer ID": 130, "Location ID": 80, "Unique Payer Identifier": 130,
            "Extraction Method": 95, "Source File": 150, "Warnings": 180,
        }
        for col in self.results_columns:
            self.results_tree.heading(col, text=col)
            self.results_tree.column(col, width=col_widths.get(col, 80), anchor=tk.W)
        
        # Configure tags for row highlighting - dark theme
        self.results_tree.tag_configure("expired", background="#3B1219", foreground="#FCA5A5")
        self.results_tree.tag_configure("normal", background="#1E293B")
        self.results_tree.tag_configure("ocr", background="#2D2B1E", foreground="#FDE68A")  # amber tint
        self.results_tree.tag_configure("structured", background="#1E2D3B", foreground="#93C5FD")  # blue tint
        self.results_tree.tag_configure("warning", background="#3B2F1E", foreground="#FBBF24")  # warning amber
        
        # Scrollbars
        tree_scroll_y = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.results_tree.yview)
        tree_scroll_x = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.results_tree.xview)
        self.results_tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        
        self.results_tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")
        
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        
        # Bind double-click for editing
        self.results_tree.bind("<Double-1>", self.on_results_double_click)
        
        # Row count label for successful
        self.results_count_var = tk.StringVar(value="0 records ready for upload")
        ttk.Label(success_frame, textvariable=self.results_count_var, 
                  font=("Segoe UI", 9, "italic")).pack(anchor=tk.W, pady=(5, 0))
        
        # === HIDDEN errors treeview (used by rest of code, shown via popup) ===
        self.errors_columns = ["File Name", "Error", "Auth Type"]
        self.errors_tree = ttk.Treeview(main_frame, columns=self.errors_columns, show="headings", height=0)
        self.errors_tree.heading("File Name", text="File Name")
        self.errors_tree.heading("Error", text="Error Details")
        self.errors_tree.heading("Auth Type", text="Type")
        self.errors_tree.tag_configure("error", background="#3B1219", foreground="#FCA5A5")
        self.errors_count_var = tk.StringVar(value="0 errors")
        
        # === BOTTOM BUTTON BAR ===
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(4, 4))
        
        self.refresh_review_btn = ttk.Button(button_frame, text="🔄 Refresh", 
                                              command=self.refresh_review_from_excel)
        self.refresh_review_btn.pack(side=tk.LEFT, padx=2)
        
        self.open_btn = ttk.Button(button_frame, text="📂 Open File", 
                                   command=self.open_output)
        self.open_btn.pack(side=tk.LEFT, padx=2)
        
        self.caspio_btn = ttk.Button(button_frame, text="☁️ Upload All to Caspio", 
                                     command=lambda: self.show_caspio_upload_dialog(valid_only=False), 
                                     style='Action.TButton')
        self.caspio_btn.pack(side=tk.LEFT, padx=2)
        
        self.caspio_valid_btn = ttk.Button(button_frame, text="✅ Upload Valid Only", 
                                            command=lambda: self.show_caspio_upload_dialog(valid_only=True))
        self.caspio_valid_btn.pack(side=tk.LEFT, padx=2)

        self.view_errors_btn = ttk.Button(button_frame, text="⚠️ View Errors",
                                           command=self.show_errors_popup)
        self.view_errors_btn.pack(side=tk.LEFT, padx=2)
        
        # Valid count indicator and legend on same row
        self.valid_count_var = tk.StringVar(value="")
        ttk.Label(button_frame, textvariable=self.valid_count_var, 
                  font=("Segoe UI", 9), foreground="#4CAF50").pack(side=tk.LEFT, padx=(10, 5))
        
        tk.Label(button_frame, text="  ", bg="#5c2828", width=2).pack(side=tk.LEFT, padx=(5, 2))
        ttk.Label(button_frame, text="= Expired", 
                  font=("Segoe UI", 8)).pack(side=tk.LEFT)

        tk.Label(button_frame, text="  ", bg="#2D2B1E", width=2).pack(side=tk.LEFT, padx=(8, 2))
        ttk.Label(button_frame, text="= OCR", 
                  font=("Segoe UI", 8)).pack(side=tk.LEFT)

        tk.Label(button_frame, text="  ", bg="#1E2D3B", width=2).pack(side=tk.LEFT, padx=(8, 2))
        ttk.Label(button_frame, text="= Structured", 
                  font=("Segoe UI", 8)).pack(side=tk.LEFT)

    def show_errors_popup(self):
        """Show extraction errors in a popup window."""
        popup = tk.Toplevel(self.root)
        popup.title("Extraction Errors")
        popup.geometry("800x400")
        popup.transient(self.root)
        self.style_popup(popup)
        
        popup.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 400
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 200
        popup.geometry(f"+{x}+{y}")
        
        frame = ttk.Frame(popup, padding="12")
        frame.pack(fill=tk.BOTH, expand=True)
        
        error_count = len(self.errors_tree.get_children())
        ttk.Label(frame, text=f"Extraction Errors ({error_count})",
                  font=("Segoe UI", 12, "bold")).pack(anchor=tk.W, pady=(0, 8))
        
        container = ttk.Frame(frame)
        container.pack(fill=tk.BOTH, expand=True)
        
        cols = ["File Name", "Error", "Auth Type"]
        tree = ttk.Treeview(container, columns=cols, show="headings", height=12)
        tree.heading("File Name", text="File Name")
        tree.heading("Error", text="Error Details")
        tree.heading("Auth Type", text="Type")
        tree.column("File Name", width=240, anchor=tk.W)
        tree.column("Error", width=430, anchor=tk.W)
        tree.column("Auth Type", width=80, anchor=tk.CENTER)
        tree.tag_configure("error", background="#3B1219", foreground="#FCA5A5")
        
        sy = ttk.Scrollbar(container, orient=tk.VERTICAL, command=tree.yview)
        sx = ttk.Scrollbar(container, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        
        # Copy rows from hidden errors_tree into this popup tree
        for item in self.errors_tree.get_children():
            values = self.errors_tree.item(item, "values")
            tags = self.errors_tree.item(item, "tags")
            tree.insert("", tk.END, values=values, tags=tags)
        
        ttk.Button(frame, text="Close", command=popup.destroy).pack(pady=(8, 0))
    
    def show_add_entry_dialog(self):
        """Show dialog to manually add an auth entry."""
        popup = tk.Toplevel(self.root)
        popup.title("Add Auth Entry")
        popup.geometry("500x600")
        popup.transient(self.root)
        popup.grab_set()
        self.style_popup(popup)
        
        # Center the popup
        popup.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 250
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 300
        popup.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(popup, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="Add Auth Entry Manually", 
                  font=("Segoe UI", 14, "bold")).pack(pady=(0, 15))
        
        # Create entry fields for each column
        entries = {}
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create scrollable frame for fields
        canvas = tk.Canvas(fields_frame, height=450)
        scrollbar = ttk.Scrollbar(fields_frame, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        for i, col in enumerate(self.results_columns):
            row_frame = ttk.Frame(scrollable)
            row_frame.pack(fill=tk.X, pady=3)
            
            ttk.Label(row_frame, text=f"{col}:", width=22, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 10))
            entry = ttk.Entry(row_frame, width=35)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            entries[col] = entry
            
            # Add date hint for date fields
            if "Date" in col or "DOS" in col:
                ttk.Label(row_frame, text="(MM/DD/YYYY)", font=("Segoe UI", 8), 
                          foreground="gray").pack(side=tk.LEFT, padx=(5, 0))
        
        def add_entry():
            values = [entries[col].get().strip() for col in self.results_columns]
            
            # Validate required fields
            if not values[0] and not values[2]:  # Last Name or Patient Name
                messagebox.showwarning("Required Field", "Please enter Last Name or Patient Name.")
                return
            
            # Check if this is an expired auth
            is_expired = False
            try:
                date_expired_idx = self.results_columns.index("Date Auth Expired")
                last_dos_idx = self.results_columns.index("Last DOS")
                
                if values[date_expired_idx] and values[last_dos_idx]:
                    from datetime import datetime
                    expire_date = datetime.strptime(values[date_expired_idx], "%m/%d/%Y")
                    dos_date = datetime.strptime(values[last_dos_idx], "%m/%d/%Y")
                    if dos_date > expire_date:
                        is_expired = True
            except (ValueError, IndexError):
                pass
            
            tag = "expired" if is_expired else "normal"
            self.results_tree.insert("", tk.END, values=values, tags=(tag,))
            
            # Update counts
            self.update_review_counts()
            
            popup.destroy()
            messagebox.showinfo("Added", "Auth entry added successfully.")
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        
        ttk.Button(btn_frame, text="✓ Add Entry", command=add_entry, 
                   style='Action.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=popup.destroy).pack(side=tk.LEFT, padx=5)
    
    def delete_selected_row(self):
        """Delete the selected row from the results table."""
        selected = self.results_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a row to delete.")
            return
        
        if len(selected) == 1:
            confirm_msg = "Delete the selected row?"
        else:
            confirm_msg = f"Delete {len(selected)} selected rows?"
        
        if messagebox.askyesno("Confirm Delete", confirm_msg):
            for item in selected:
                self.results_tree.delete(item)
            self.update_review_counts()
    
    def mark_selected_as_error(self):
        """Mark selected rows as extraction errors and move them to the Errors section."""
        selected = self.results_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a row to mark as error.")
            return
        
        if len(selected) == 1:
            confirm_msg = "Mark the selected row as an error?\n\nThis will move it to the Extraction Errors section."
        else:
            confirm_msg = f"Mark {len(selected)} selected rows as errors?\n\nThis will move them to the Extraction Errors section."
        
        if messagebox.askyesno("Confirm Mark as Error", confirm_msg):
            for item in selected:
                values = self.results_tree.item(item, 'values')
                if values:
                    # Get patient name for display
                    patient_name = ""
                    for i, col in enumerate(self.results_columns):
                        if col == "Patient Name" and i < len(values):
                            patient_name = values[i]
                            break
                        elif col == "Last Name" and i < len(values):
                            first_name = ""
                            for j, c in enumerate(self.results_columns):
                                if c == "First Name" and j < len(values):
                                    first_name = values[j]
                                    break
                            patient_name = f"{values[i]}, {first_name}".strip(", ")
                            break
                    
                    # Determine auth type from CPT codes in the row
                    auth_type = ""
                    for i, col in enumerate(self.results_columns):
                        if col == "CPT Code" and i < len(values):
                            cpt = values[i]
                            if cpt == "S9123" or cpt == "S9124":
                                auth_type = "Skilled"
                            elif cpt == "S9122":
                                # Could be Unskilled or Escort - check CPT Code 2 or default to Unskilled
                                auth_type = "Unskilled"
                            break
                    
                    # Use patient name as filename for display
                    filename = patient_name if patient_name else "Unknown"
                    error_msg = "Manually marked - data not found in PDF"
                    
                    # Add to errors tree with auth type
                    self.errors_tree.insert("", tk.END, 
                        values=(filename, error_msg, auth_type), 
                        tags=("error",))
                    
                    # Remove from results tree
                    self.results_tree.delete(item)
            
            self.update_review_counts()
            messagebox.showinfo("Marked as Error", 
                f"{len(selected)} row(s) moved to Extraction Errors.\n\n"
                "These will appear in the Email tab queue when you refresh.")
    
    def edit_selected_entry(self):
        """Edit the selected entry in the results table."""
        selected = self.results_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a record to edit.")
            return
        
        if len(selected) > 1:
            messagebox.showwarning("Multiple Selection", "Please select only one record to edit.")
            return
        
        item = selected[0]
        current_values = self.results_tree.item(item, 'values')
        
        popup = tk.Toplevel(self.root)
        popup.title("Edit Auth Entry")
        popup.geometry("500x600")
        popup.transient(self.root)
        popup.grab_set()
        self.style_popup(popup)
        
        # Center the popup
        popup.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 250
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 300
        popup.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(popup, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="Edit Auth Entry", 
                  font=("Segoe UI", 14, "bold")).pack(pady=(0, 15))
        
        # Create entry fields for each column
        entries = {}
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill=tk.BOTH, expand=True)
        
        # Create scrollable frame for fields
        canvas = tk.Canvas(fields_frame, height=450)
        scrollbar = ttk.Scrollbar(fields_frame, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        
        scrollable.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        for i, col in enumerate(self.results_columns):
            row_frame = ttk.Frame(scrollable)
            row_frame.pack(fill=tk.X, pady=3)
            
            ttk.Label(row_frame, text=f"{col}:", width=22, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 10))
            entry = ttk.Entry(row_frame, width=35)
            # Pre-fill with current value
            if i < len(current_values):
                entry.insert(0, current_values[i])
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            entries[col] = entry
        
        def save_entry():
            # Get values from entries
            new_values = [entries[col].get() for col in self.results_columns]
            
            # Determine if expired based on dates
            tags = ["normal"]
            try:
                from datetime import datetime
                date_expired_str = entries.get("Date Auth Expired", None)
                last_dos_str = entries.get("Last DOS", None)
                if date_expired_str and last_dos_str:
                    date_expired_val = date_expired_str.get()
                    last_dos_val = last_dos_str.get()
                    if date_expired_val and last_dos_val:
                        date_expired = pd.to_datetime(date_expired_val, errors="coerce")
                        last_dos = pd.to_datetime(last_dos_val, errors="coerce")
                        if pd.notna(date_expired) and pd.notna(last_dos):
                            if last_dos > date_expired:
                                tags = ["expired"]
            except:
                pass
            
            # Update the tree item
            self.results_tree.item(item, values=new_values, tags=tags)
            self.update_review_counts()
            popup.destroy()
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        
        ttk.Button(btn_frame, text="✓ Save Changes", command=save_entry, 
                   style='Action.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=popup.destroy).pack(side=tk.LEFT, padx=5)

    def update_review_counts(self):
        """Update the count labels on the Review & Edit tab."""
        total = 0
        expired = 0
        valid = 0
        
        for item in self.results_tree.get_children():
            total += 1
            tags = self.results_tree.item(item, 'tags')
            if 'expired' in tags:
                expired += 1
            else:
                valid += 1
        
        self.results_count_var.set(f"{total} records ({valid} valid, {expired} expired)")
        self.valid_count_var.set(f"✓ {valid} valid auths ready for upload")
        
        # Update summary
        error_count = len(self.errors_tree.get_children()) if hasattr(self, 'errors_tree') else 0
        self.review_summary_var.set(
            f"Total Records: {total}  |  Valid: {valid}  |  Expired: {expired}  |  Errors: {error_count}"
        )
    
    def get_results_from_table(self, valid_only=False):
        """Get current data from results table as DataFrame.
        
        Args:
            valid_only: If True, only return non-expired rows.
        """
        data = []
        for item in self.results_tree.get_children():
            tags = self.results_tree.item(item, 'tags')
            
            # Skip expired rows if valid_only is True
            if valid_only and 'expired' in tags:
                continue
            
            values = self.results_tree.item(item, "values")
            row = {col: values[i] if i < len(values) else "" for i, col in enumerate(self.results_columns)}
            data.append(row)
        return pd.DataFrame(data) if data else None

    def refresh_review_from_excel(self):
        """Reload review tab data from the output Excel file."""
        output_file = self.output_file.get()
        if not output_file or not pathlib.Path(output_file).exists():
            messagebox.showwarning("No File", "No output file found. Run extraction first.")
            return
        
        try:
            xl = pd.ExcelFile(output_file)
            if "Formatted" in xl.sheet_names:
                df = pd.read_excel(output_file, sheet_name="Formatted")
                self.populate_results_table(df)
                messagebox.showinfo("Refreshed", f"Loaded {len(df)} records from {output_file}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file: {str(e)}")

    def setup_finder_tab(self):
        """Set up the File Finder tab UI."""
        main_frame = ttk.Frame(self.finder_tab, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Description
        desc_label = ttk.Label(main_frame, 
            text="Search for PDFs by patient name, then copy matches to a destination folder",
            style='Desc.TLabel')
        desc_label.pack(anchor=tk.W, pady=(0, 15))
        
        # Top row: Folders on left, Date filter on right
        top_row = ttk.Frame(main_frame)
        top_row.pack(fill=tk.X, pady=5)
        
        # Left side: Source and Destination folders (takes most of the width)
        folders_frame = ttk.Frame(top_row)
        folders_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Source type selection — Local Folder or Dropbox
        src_type_row = ttk.Frame(folders_frame)
        src_type_row.pack(fill=tk.X, pady=(0, 6), padx=(0, 15))
        ttk.Label(src_type_row, text="Source:", font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(src_type_row, text="📁 Local Folder", variable=self.finder_source_type,
                        value="local", command=self._toggle_finder_source).pack(side=tk.LEFT, padx=(0, 20))
        ttk.Radiobutton(src_type_row, text="☁️ Dropbox Folder", variable=self.finder_source_type,
                        value="dropbox", command=self._toggle_finder_source).pack(side=tk.LEFT)

        # Container that switches between local source frame and Dropbox source frame
        self._finder_source_container = ttk.Frame(folders_frame)
        self._finder_source_container.pack(fill=tk.X, pady=(0, 5), padx=(0, 15))

        # Local source folder (shown by default)
        self.finder_local_source_frame = ttk.LabelFrame(self._finder_source_container, text="Source Folder", padding="8")
        self.finder_local_source_frame.pack(fill=tk.X)

        source_entry = ttk.Entry(self.finder_local_source_frame, textvariable=self.finder_source, width=55)
        source_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))

        source_btn = ttk.Button(self.finder_local_source_frame, text="📁 Browse", command=self.browse_finder_source)
        source_btn.pack(side=tk.RIGHT)

        # Dropbox source (hidden until Dropbox radio selected)
        self.finder_dropbox_source_frame = ttk.LabelFrame(self._finder_source_container, text="☁️ Dropbox Source", padding="8")
        # Not packed yet — shown by _toggle_finder_source

        # Row 1: Connect + status + folder dropdown
        _dbx_r1 = ttk.Frame(self.finder_dropbox_source_frame)
        _dbx_r1.pack(fill=tk.X, pady=(0, 4))
        self.dropbox_connect_btn = ttk.Button(_dbx_r1, text="🔗 Connect", command=self._connect_dropbox)
        self.dropbox_connect_btn.pack(side=tk.LEFT, padx=(0, 8))
        ttk.Label(_dbx_r1, textvariable=self.dropbox_status_var,
                  font=("Segoe UI", 9, "italic"), foreground="gray").pack(side=tk.LEFT, padx=(0, 12))
        ttk.Label(_dbx_r1, text="Folder:").pack(side=tk.LEFT, padx=(4, 2))
        self.dropbox_folder_combo = ttk.Combobox(_dbx_r1, textvariable=self.dropbox_folder_var,
                                                  width=30, state="normal")
        self.dropbox_folder_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        ttk.Button(_dbx_r1, text="🔄", width=3,
                   command=self._refresh_dropbox_folders).pack(side=tk.LEFT)

        # Row 2: Type/keyword filter
        _dbx_r3 = ttk.Frame(self.finder_dropbox_source_frame)
        _dbx_r3.pack(fill=tk.X)
        ttk.Label(_dbx_r3, text="Type filter:").pack(side=tk.LEFT, padx=(0, 6))
        for _kw in ("Skilled", "Unskilled", "Escort"):
            ttk.Radiobutton(_dbx_r3, text=_kw, variable=self.dropbox_keyword_var,
                            value=_kw).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(_dbx_r3, text="All", variable=self.dropbox_keyword_var,
                        value="").pack(side=tk.LEFT)
        
        # Destination folder selection
        dest_frame = ttk.LabelFrame(folders_frame, text="Destination Folder", padding="8")
        dest_frame.pack(fill=tk.X, padx=(0, 15))
        
        dest_entry = ttk.Entry(dest_frame, textvariable=self.finder_dest, width=55)
        dest_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        
        dest_btn = ttk.Button(dest_frame, text="📁 Browse", command=self.browse_finder_dest)
        dest_btn.pack(side=tk.RIGHT)
        
        # Right side: Date range filter (fixed width, larger)
        date_frame = ttk.LabelFrame(top_row, text="Date Range Filter", padding="12")
        date_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        # From date row
        from_frame = ttk.Frame(date_frame)
        from_frame.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Label(from_frame, text="From:", width=6).pack(side=tk.LEFT)
        self.finder_date_from = tk.StringVar()
        from_entry = ttk.Entry(from_frame, textvariable=self.finder_date_from, width=14)
        from_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # To date row
        to_frame = ttk.Frame(date_frame)
        to_frame.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Label(to_frame, text="To:", width=6).pack(side=tk.LEFT)
        self.finder_date_to = tk.StringVar()
        to_entry = ttk.Entry(to_frame, textvariable=self.finder_date_to, width=14)
        to_entry.pack(side=tk.LEFT, padx=(0, 5))
        
        # Date format hint
        ttk.Label(date_frame, text="Format: MM/DD/YYYY", 
                  font=("Segoe UI", 8), foreground="gray").pack(anchor=tk.W, pady=(5, 0))
        
        # Toggle buttons frame
        toggle_frame = ttk.Frame(main_frame)
        toggle_frame.pack(fill=tk.X, pady=(10, 5))
        
        self.finder_view_var = tk.StringVar(value="criteria")
        
        self.criteria_toggle_btn = ttk.Button(toggle_frame, text="📝 Search Criteria", 
                                               command=lambda: self.toggle_finder_view("criteria"))
        self.criteria_toggle_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.results_toggle_btn = ttk.Button(toggle_frame, text="📊 View Results", 
                                              command=lambda: self.toggle_finder_view("results"))
        self.results_toggle_btn.pack(side=tk.LEFT, padx=5)
        
        # Results summary (shown next to toggle)
        self.finder_results_summary_var = tk.StringVar(value="")
        ttk.Label(toggle_frame, textvariable=self.finder_results_summary_var, 
                  font=("Segoe UI", 9, "bold")).pack(side=tk.LEFT, padx=(20, 0))
        
        # Container for toggleable content
        self.finder_content_container = ttk.Frame(main_frame)

        # Status line - pack at bottom FIRST so it's always visible
        status_frame = ttk.Frame(main_frame)
        status_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(0, 5))
        self.finder_status_var = tk.StringVar(value="Ready to search")
        self.finder_status_label = ttk.Label(status_frame, textvariable=self.finder_status_var,
                                              font=("Segoe UI", 9), foreground="gray")
        self.finder_status_label.pack(side=tk.LEFT)

        # Buttons - pack at bottom above status so they're always visible
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(5, 2))

        self.find_btn = ttk.Button(button_frame, text="\U0001f50d Find & Copy Files",
                                   command=self.run_file_finder, style='Action.TButton')
        self.find_btn.pack(side=tk.LEFT, padx=5)

        self.clear_btn = ttk.Button(button_frame, text="\U0001f5d1\ufe0f Clear", command=self.clear_finder)
        self.clear_btn.pack(side=tk.LEFT, padx=5)

        self.open_dest_btn = ttk.Button(button_frame, text="\U0001f4c2 Open Destination", command=self.open_finder_dest)
        self.open_dest_btn.pack(side=tk.LEFT, padx=5)

        self.export_finder_btn = ttk.Button(button_frame, text="\U0001f4ca Export Results to Excel", command=self.export_finder_results)
        self.export_finder_btn.pack(side=tk.LEFT, padx=5)

        # Pack content container last - it fills remaining space between top widgets and bottom buttons
        self.finder_content_container.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # === SEARCH CRITERIA VIEW ===
        self.search_criteria_frame = ttk.LabelFrame(self.finder_content_container, text="Search Criteria", padding="10")
        
        # Entry row for adding new criteria
        entry_frame = ttk.Frame(self.search_criteria_frame)
        entry_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(entry_frame, text="Patient Name:").pack(side=tk.LEFT, padx=(0, 5))
        self.finder_name_entry = ttk.Entry(entry_frame, width=30)
        self.finder_name_entry.pack(side=tk.LEFT, padx=(0, 15))
        
        ttk.Label(entry_frame, text="Auth Type:").pack(side=tk.LEFT, padx=(0, 5))
        self.finder_type_combo = ttk.Combobox(entry_frame, width=10, 
            values=["Unskilled", "Skilled"], state="readonly")
        self.finder_type_combo.pack(side=tk.LEFT, padx=(0, 15))
        self.finder_type_combo.set("Unskilled")
        
        add_btn = ttk.Button(entry_frame, text="➕ Add", command=self.add_finder_row)
        add_btn.pack(side=tk.LEFT, padx=5)
        
        remove_btn = ttk.Button(entry_frame, text="➖ Remove", command=self.remove_finder_row)
        remove_btn.pack(side=tk.LEFT, padx=5)
        
        bulk_btn = ttk.Button(entry_frame, text="📋 Bulk Import", command=self.bulk_import_finder)
        bulk_btn.pack(side=tk.LEFT, padx=5)
        
        # Auth type legend
        legend = ttk.Label(entry_frame, 
            text="Unskilled=Unskilled/Escort Assistance | Skilled=Skilled",
            font=("Segoe UI", 8), foreground="gray")
        legend.pack(side=tk.RIGHT)
        
        # Table (Treeview) for criteria
        table_frame = ttk.Frame(self.search_criteria_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(10, 0))
        
        columns = ("name", "auth_type", "searches_for", "last_dos")
        self.finder_table = ttk.Treeview(table_frame, columns=columns, show="headings", height=10)
        
        self.finder_table.heading("name", text="Patient Name")
        self.finder_table.heading("auth_type", text="Auth Type")
        self.finder_table.heading("searches_for", text="Searches For")
        self.finder_table.heading("last_dos", text="Last DOS")
        
        self.finder_table.column("name", width=250, anchor=tk.W)
        self.finder_table.column("auth_type", width=100, anchor=tk.CENTER)
        self.finder_table.column("searches_for", width=180, anchor=tk.W)
        self.finder_table.column("last_dos", width=120, anchor=tk.CENTER)
        
        table_scroll = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.finder_table.yview)
        self.finder_table.configure(yscrollcommand=table_scroll.set)
        
        self.finder_table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        table_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind Enter key to add row
        self.finder_name_entry.bind('<Return>', lambda e: self.add_finder_row())
        
        # === RESULTS VIEW (Split: Found left, Not Found right) ===
        self.results_view_frame = ttk.Frame(self.finder_content_container)
        
        # Left side: Files Found
        found_frame = ttk.LabelFrame(self.results_view_frame, text="Files Found", padding="8")
        found_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        found_container = ttk.Frame(found_frame)
        found_container.pack(fill=tk.BOTH, expand=True)
        
        self.finder_found_columns = ["Patient Name", "Auth Type", "File Name", "Last DOS"]
        self.finder_found_tree = ttk.Treeview(found_container, columns=self.finder_found_columns, 
                                               show="headings", height=12)
        
        self.finder_found_tree.heading("Patient Name", text="Patient Name")
        self.finder_found_tree.heading("Auth Type", text="Auth Type")
        self.finder_found_tree.heading("File Name", text="File Name")
        self.finder_found_tree.heading("Last DOS", text="Last DOS")
        
        self.finder_found_tree.column("Patient Name", width=150, anchor=tk.W)
        self.finder_found_tree.column("Auth Type", width=70, anchor=tk.CENTER)
        self.finder_found_tree.column("File Name", width=250, anchor=tk.W)
        self.finder_found_tree.column("Last DOS", width=80, anchor=tk.CENTER)
        
        self.finder_found_tree.tag_configure("duplicate", background="#422006", foreground="#FCD34D")
        self.finder_found_tree.tag_configure("normal", background="#1E293B")
        
        found_scroll_y = ttk.Scrollbar(found_container, orient=tk.VERTICAL, command=self.finder_found_tree.yview)
        self.finder_found_tree.configure(yscrollcommand=found_scroll_y.set)
        
        self.finder_found_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        found_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Found count label
        self.finder_found_count_var = tk.StringVar(value="0 files found")
        ttk.Label(found_frame, textvariable=self.finder_found_count_var, 
                  font=("Segoe UI", 9, "italic"), foreground="#4CAF50").pack(anchor=tk.W, pady=(5, 0))
        
        # Right side: Names Not Found
        not_found_frame = ttk.LabelFrame(self.results_view_frame, text="Names Not Found", padding="8")
        not_found_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        not_found_container = ttk.Frame(not_found_frame)
        not_found_container.pack(fill=tk.BOTH, expand=True)
        
        self.finder_not_found_columns = ["Patient Name", "Auth Type", "Reason"]
        self.finder_not_found_tree = ttk.Treeview(not_found_container, columns=self.finder_not_found_columns, 
                                                   show="headings", height=12)
        
        self.finder_not_found_tree.heading("Patient Name", text="Patient Name")
        self.finder_not_found_tree.heading("Auth Type", text="Auth Type")
        self.finder_not_found_tree.heading("Reason", text="Reason")
        
        self.finder_not_found_tree.column("Patient Name", width=150, anchor=tk.W)
        self.finder_not_found_tree.column("Auth Type", width=70, anchor=tk.CENTER)
        self.finder_not_found_tree.column("Reason", width=180, anchor=tk.W)
        
        self.finder_not_found_tree.tag_configure("not_found", background="#422006", foreground="#FCD34D")
        
        not_found_scroll_y = ttk.Scrollbar(not_found_container, orient=tk.VERTICAL, command=self.finder_not_found_tree.yview)
        self.finder_not_found_tree.configure(yscrollcommand=not_found_scroll_y.set)
        
        self.finder_not_found_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        not_found_scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Not found count label
        self.finder_not_found_count_var = tk.StringVar(value="0 not found")
        ttk.Label(not_found_frame, textvariable=self.finder_not_found_count_var, 
                  font=("Segoe UI", 9, "italic"), foreground="#ff6b6b").pack(anchor=tk.W, pady=(5, 0))
        
        # Show criteria view by default
        self.search_criteria_frame.pack(fill=tk.BOTH, expand=True)

    def _toggle_finder_source(self):
        """Toggle between local folder and Dropbox source in File Finder."""
        if self.finder_source_type.get() == "dropbox":
            self.finder_local_source_frame.pack_forget()
            self.finder_dropbox_source_frame.pack(fill=tk.X)
        else:
            self.finder_dropbox_source_frame.pack_forget()
            self.finder_local_source_frame.pack(fill=tk.X)

    def _auto_unlock_pdf(self, pdf_path):
        """Automatically decrypt a PDF in place if it's encrypted. Non-fatal on failure."""
        if not pdf_path.lower().endswith(".pdf"):
            return
        try:
            from services.pdf_unlock_service import PdfUnlockService
            unlock = PdfUnlockService()
            unlocked_path, was_encrypted = unlock.unlock(pdf_path)
            if was_encrypted:
                import shutil as _shutil
                _shutil.move(unlocked_path, pdf_path)
                self.finder_log_msg(f"      🔓 Decrypted: {os.path.basename(pdf_path)}")
        except Exception:
            pass  # Non-fatal — keep the file as-is

    def _run_finder_from_dropbox(self):
        """Run the File Finder using Dropbox as the source, downloading matches to destination."""
        if not self.dropbox_service or not self.dropbox_service.is_connected:
            messagebox.showerror("Error", "Connect to Dropbox first using the 🔗 Connect button.")
            return

        dest = self.finder_dest.get()
        if not dest:
            messagebox.showerror("Error", "Please select a destination folder.")
            return

        criteria = []
        for item in self.finder_table.get_children():
            values = self.finder_table.item(item, 'values')
            criteria.append((values[0], values[1], values[3] if len(values) > 3 else ""))

        # Parse the shared Date Range Filter (MM/DD/YYYY)
        date_from_str = self.finder_date_from.get().strip()
        date_to_str = self.finder_date_to.get().strip()
        from_date = None
        to_date = None

        if date_from_str:
            try:
                from_date = datetime.strptime(date_from_str, "%m/%d/%Y").date()
            except ValueError:
                messagebox.showerror("Error", f"Invalid 'From' date format: {date_from_str}\nUse MM/DD/YYYY")
                return

        if date_to_str:
            try:
                to_date = datetime.strptime(date_to_str, "%m/%d/%Y").date()
            except ValueError:
                messagebox.showerror("Error", f"Invalid 'To' date format: {date_to_str}\nUse MM/DD/YYYY")
                return

        if not criteria and not from_date and not to_date:
            messagebox.showerror("Error",
                "Please add patient names to the criteria table, set a date range, or both.")
            return

        self.finder_status_var.set("Listing Dropbox files...")
        self.root.update_idletasks()

        # List files from Dropbox, apply date range + keyword filters
        try:
            folder = self.dropbox_folder_var.get().strip()
            all_files = self.dropbox_service.list_supported_files(folder)

            if from_date or to_date:
                filtered = []
                for m in all_files:
                    mod = getattr(m, 'server_modified', None) or getattr(m, 'client_modified', None)
                    if mod is None:
                        continue
                    # Dropbox returns timezone-aware UTC datetimes; strip tz before
                    # extracting date so it compares cleanly against our naive dates.
                    if hasattr(mod, 'tzinfo') and mod.tzinfo is not None:
                        mod = mod.replace(tzinfo=None)
                    mod_date = mod.date() if hasattr(mod, 'date') else mod
                    if from_date and mod_date < from_date:
                        continue
                    if to_date and mod_date > to_date:
                        continue
                    filtered.append(m)
                self.finder_log_msg(f"   📅 Date range → {len(filtered)} of {len(all_files)} files")
                all_files = filtered

            keyword = self.dropbox_keyword_var.get().strip()
            if keyword:
                all_files = [m for m in all_files if keyword.lower() in m.name.lower()]
                self.finder_log_msg(f"   🔤 Type filter '{keyword}' → {len(all_files)} files")

            self.finder_log_msg(f"☁️  {len(all_files)} Dropbox files after filters.")
        except Exception as e:
            messagebox.showerror("Dropbox Error", f"Failed to list files:\n{e}")
            return

        dest_path = pathlib.Path(dest)
        dest_path.mkdir(parents=True, exist_ok=True)

        found_matches = []
        not_found_list = []
        matched = 0

        if not criteria:
            # ── Date-range-only mode: download every file that passed the filters ──
            self.finder_log_msg(f"   📥 Date-range mode: downloading {len(all_files)} files...")
            for meta in all_files:
                dest_file = dest_path / meta.name
                if not dest_file.exists():
                    try:
                        self.finder_log_msg(f"   ⬇️  {meta.name}")
                        self.dropbox_service.download_file(meta, str(dest_path))
                        # Auto-unlock encrypted PDFs in place
                        self._auto_unlock_pdf(str(dest_file))
                        matched += 1
                    except Exception as e:
                        self.finder_log_msg(f"   ❌ Download failed {meta.name}: {e}")
                        continue
                else:
                    matched += 1
                found_matches.append((meta.name, "", meta.name, ""))
            self.finder_duplicate_imports = []
            self.finder_original_count = len(found_matches)
        else:
            # ── Name-matching mode (date range already applied to all_files above) ──
            all_imports_by_key = {}
            original_criteria_count = len(criteria)
            for name, auth_type, last_dos in criteria:
                key = (name.upper().strip(), auth_type.upper().strip())
                all_imports_by_key.setdefault(key, []).append((name, auth_type, last_dos))

            duplicate_import_entries = [e for entries in all_imports_by_key.values() for e in entries[1:]]
            criteria_for_search = [(v[0][0], v[0][1], v[0][2]) for v in all_imports_by_key.values()]

            self.finder_log_msg(f"   🔍 Matching {len(criteria_for_search)} patients against {len(all_files)} files...")

            name_found_type_mismatch = []
            not_found_at_all = []

            for name, auth_type, last_dos in criteria_for_search:
                name_lower = name.lower()
                name_cleaned = re.sub(r'\s*\([^)]*\)\s*', ' ', name_lower)
                name_parts = [p for p in name_cleaned.replace(",", " ").split() if len(p) > 2]

                matching_files, name_only_matches = [], []

                for meta in all_files:
                    fn_lower = meta.name.lower()
                    if name_parts:
                        hits = sum(1 for p in name_parts if p in fn_lower)
                        n_match = hits == len(name_parts) if len(name_parts) <= 2 else hits >= len(name_parts) - 1
                    else:
                        n_match = False

                    if auth_type.lower() == "skilled":
                        k_match = "skilled" in fn_lower and "unskilled" not in fn_lower
                    elif auth_type.lower() == "escort":
                        k_match = "escort" in fn_lower
                    elif auth_type.lower() == "unskilled":
                        k_match = "unskilled" in fn_lower
                    else:
                        k_match = False

                    if n_match:
                        mod = getattr(meta, 'server_modified', None) or getattr(meta, 'client_modified', None)
                        if k_match:
                            matching_files.append((meta, mod))
                        else:
                            found_type = ("Skilled" if "skilled" in fn_lower and "unskilled" not in fn_lower
                                          else "Unskilled" if "unskilled" in fn_lower
                                          else "Escort" if "escort" in fn_lower else "")
                            if found_type:
                                name_only_matches.append((meta, found_type))

                key = (name.upper().strip(), auth_type.upper().strip())
                all_for_name = all_imports_by_key.get(key, [(name, auth_type, last_dos)])

                if matching_files:
                    matching_files.sort(key=lambda x: x[1] if x[1] else datetime.min, reverse=True)
                    best = matching_files[0][0]
                    dest_file = dest_path / best.name
                    if not dest_file.exists():
                        try:
                            self.finder_log_msg(f"   ⬇️  {best.name}")
                            self.dropbox_service.download_file(best, str(dest_path))
                            # Auto-unlock encrypted PDFs in place
                            self._auto_unlock_pdf(str(dest_file))
                            matched += 1
                        except Exception as e:
                            self.finder_log_msg(f"   ❌ Download failed {best.name}: {e}")
                            for n, t, d in all_for_name:
                                not_found_at_all.append((n, t, d))
                            continue
                    for n, t, d in all_for_name:
                        found_matches.append((n, t, best.name, d))
                elif name_only_matches:
                    found_type = name_only_matches[0][1]
                    for n, t, d in all_for_name:
                        name_found_type_mismatch.append((n, t, found_type, d))
                else:
                    for n, t, d in all_for_name:
                        not_found_at_all.append((n, t, d))

            self.finder_duplicate_imports = duplicate_import_entries
            self.finder_original_count = original_criteria_count
            for n, wt, ft, d in name_found_type_mismatch:
                not_found_list.append((n, wt, f"Type mismatch — wanted {wt}, found {ft}", d))
            for n, t, d in not_found_at_all:
                not_found_list.append((n, t, "Not found in Dropbox", d))

        # Store results
        self.finder_found_matches = found_matches
        self.finder_not_found = not_found_list

        if hasattr(self, 'finder_files_count_var'):
            self.update_finder_files_count()

        self.populate_finder_results()
        self.finder_status_var.set(f"Done — {matched} files downloaded, {len(not_found_list)} not found")

        messagebox.showinfo("Dropbox File Finder Complete",
            f"Downloaded {matched} files to:\n{dest}\n\n"
            f"Files found: {len(found_matches)}\n"
            f"Not found: {len(not_found_list)}")

    def toggle_finder_view(self, view):
        """Toggle between search criteria view and results view."""
        self.finder_view_var.set(view)
        
        if view == "criteria":
            self.results_view_frame.pack_forget()
            self.search_criteria_frame.pack(fill=tk.BOTH, expand=True)
            # Update button styles to show active state
            self.criteria_toggle_btn.configure(style='Action.TButton')
            self.results_toggle_btn.configure(style='TButton')
        else:  # results
            self.search_criteria_frame.pack_forget()
            self.results_view_frame.pack(fill=tk.BOTH, expand=True)
            # Update button styles
            self.criteria_toggle_btn.configure(style='TButton')
            self.results_toggle_btn.configure(style='Action.TButton')
    
    def proceed_to_extract_from_finder(self):
        """Navigate to Extract tab and set up for extraction from finder destination."""
        dest = self.finder_dest.get()
        if dest and pathlib.Path(dest).exists():
            self.input_folder.set(dest)
            # Suggest output file name
            output_file = pathlib.Path(dest) / "Auth_Results_Combined.xlsx"
            self.output_file.set(str(output_file))
        self.notebook.select(self.extractor_tab)
    
    def populate_finder_results(self):
        """Populate the finder results view with current finder data."""
        # Clear existing items
        for item in self.finder_found_tree.get_children():
            self.finder_found_tree.delete(item)
        for item in self.finder_not_found_tree.get_children():
            self.finder_not_found_tree.delete(item)
        
        # Track which files are used by multiple names
        file_to_names = {}  # filename -> list of names
        for name, auth_type, filename, last_dos in self.finder_found_matches:
            if filename not in file_to_names:
                file_to_names[filename] = []
            file_to_names[filename].append((name, auth_type, last_dos))
        
        # Identify duplicate files (same file matched to multiple names)
        duplicate_files = {f for f, names in file_to_names.items() if len(names) > 1}
        
        # Populate found matches
        found_count = 0
        duplicate_count = 0
        unique_files = set()
        
        for name, auth_type, filename, last_dos in self.finder_found_matches:
            is_duplicate = filename in duplicate_files
            tag = "duplicate" if is_duplicate else "normal"
            if is_duplicate:
                duplicate_count += 1
            
            self.finder_found_tree.insert("", tk.END, 
                                          values=(name, auth_type, filename, last_dos),
                                          tags=(tag,))
            unique_files.add(filename)
            found_count += 1
        
        # Populate not found
        not_found_count = 0
        for name, auth_type, reason, last_dos in self.finder_not_found:
            self.finder_not_found_tree.insert("", tk.END,
                                               values=(name, auth_type, reason),
                                               tags=("not_found",))
            not_found_count += 1
        
        # Update counts
        self.finder_found_count_var.set(f"{found_count} files found")
        self.finder_not_found_count_var.set(f"{not_found_count} not found")
        
        # Update summary in toggle bar
        self.finder_results_summary_var.set(
            f"Found: {found_count}  |  Not Found: {not_found_count}  |  Unique Files: {len(unique_files)}"
        )
        
        # Auto-switch to results view
        self.toggle_finder_view("results")

    def setup_email_tab(self):
        """Set up the Email tab with exception queue and invalid auth table."""
        main_frame = ttk.Frame(self.email_tab, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Description
        desc_label = ttk.Label(main_frame, 
            text="Work all errors from this page - expired auths, extraction errors, and not found names",
            style='Desc.TLabel')
        desc_label.pack(anchor=tk.W, pady=(0, 10))
        
        # === TOP SECTION: Exception Queue ===
        queue_frame = ttk.LabelFrame(main_frame, text="Exception Queue - All Errors", padding="10")
        queue_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Queue description
        ttk.Label(queue_frame, 
            text="Expired auths, not found names, and extraction errors - select to move to email list or mark as fixed",
            font=("Segoe UI", 9), foreground="gray").pack(anchor=tk.W, pady=(0, 5))
        
        # Queue toolbar
        queue_toolbar = ttk.Frame(queue_frame)
        queue_toolbar.pack(fill=tk.X, pady=(0, 5))
        
        self.refresh_queue_btn = ttk.Button(queue_toolbar, text="🔄 Refresh Queue", 
                                              command=self.refresh_email_queue)
        self.refresh_queue_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.move_to_invalid_btn = ttk.Button(queue_toolbar, text="⬇️ Move to Email List", 
                                                command=self.move_to_invalid_auth,
                                                style='Action.TButton')
        self.move_to_invalid_btn.pack(side=tk.LEFT, padx=5)
        
        self.mark_fixed_btn = ttk.Button(queue_toolbar, text="✅ Mark as Fixed/Added", 
                                          command=self.mark_as_fixed)
        self.mark_fixed_btn.pack(side=tk.LEFT, padx=5)
        
        # Queue count
        self.queue_count_var = tk.StringVar(value="0 items in queue")
        ttk.Label(queue_toolbar, textvariable=self.queue_count_var, 
                  font=("Segoe UI", 9, "italic")).pack(side=tk.RIGHT)
        
        # Queue treeview
        queue_container = ttk.Frame(queue_frame)
        queue_container.pack(fill=tk.BOTH, expand=True)
        
        self.queue_columns = ["Patient Name", "Type", "Source", "Reason"]
        self.email_queue_tree = ttk.Treeview(queue_container, columns=self.queue_columns, 
                                              show="headings", height=8)
        
        self.email_queue_tree.heading("Patient Name", text="Patient Name")
        self.email_queue_tree.heading("Type", text="Type")
        self.email_queue_tree.heading("Source", text="Source")
        self.email_queue_tree.heading("Reason", text="Reason")
        
        self.email_queue_tree.column("Patient Name", width=200, anchor=tk.W)
        self.email_queue_tree.column("Type", width=100, anchor=tk.CENTER)
        self.email_queue_tree.column("Source", width=120, anchor=tk.CENTER)
        self.email_queue_tree.column("Reason", width=300, anchor=tk.W)
        
        self.email_queue_tree.tag_configure("not_found", background="#422006", foreground="#FCD34D")
        self.email_queue_tree.tag_configure("error", background="#3B1219", foreground="#FCA5A5")
        self.email_queue_tree.tag_configure("expired", background="#3B1219", foreground="#FCA5A5")
        
        queue_scroll = ttk.Scrollbar(queue_container, orient=tk.VERTICAL, command=self.email_queue_tree.yview)
        self.email_queue_tree.configure(yscrollcommand=queue_scroll.set)
        
        self.email_queue_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        queue_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # === BOTTOM SECTION: Split into Invalid Auth (left) and Email Settings (right) ===
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.BOTH, expand=True)
        
        # LEFT: Invalid Auth Table (75%)
        invalid_frame = ttk.LabelFrame(bottom_frame, text="Email List", padding="8")
        invalid_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        # Invalid auth toolbar
        invalid_toolbar = ttk.Frame(invalid_frame)
        invalid_toolbar.pack(fill=tk.X, pady=(0, 5))
        
        self.remove_invalid_btn = ttk.Button(invalid_toolbar, text="⬆️ Remove", 
                                               command=self.remove_from_invalid)
        self.remove_invalid_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.clear_invalid_btn = ttk.Button(invalid_toolbar, text="🗑️ Clear", 
                                              command=self.clear_invalid_list)
        self.clear_invalid_btn.pack(side=tk.LEFT, padx=5)
        
        # Invalid count
        self.invalid_count_var = tk.StringVar(value="0 patients")
        ttk.Label(invalid_toolbar, textvariable=self.invalid_count_var, 
                  font=("Segoe UI", 9, "italic"), foreground="#ff6b6b").pack(side=tk.RIGHT)
        
        # Invalid auth treeview - compact columns
        invalid_container = ttk.Frame(invalid_frame)
        invalid_container.pack(fill=tk.BOTH, expand=True)
        
        self.invalid_columns = ["Patient Name", "Auth Type", "Reason"]
        self.invalid_auth_tree = ttk.Treeview(invalid_container, columns=self.invalid_columns, 
                                               show="headings", height=6)
        
        self.invalid_auth_tree.heading("Patient Name", text="Patient Name")
        self.invalid_auth_tree.heading("Auth Type", text="Type")
        self.invalid_auth_tree.heading("Reason", text="Reason")
        
        self.invalid_auth_tree.column("Patient Name", width=100, anchor=tk.W)
        self.invalid_auth_tree.column("Auth Type", width=100, anchor=tk.CENTER)
        self.invalid_auth_tree.column("Reason", width=100, anchor=tk.W)
        
        self.invalid_auth_tree.tag_configure("expired", background="#3B1219", foreground="#FCA5A5")
        self.invalid_auth_tree.tag_configure("confirmed", background="#1E293B", foreground="#94A3B8")
        
        invalid_scroll = ttk.Scrollbar(invalid_container, orient=tk.VERTICAL, command=self.invalid_auth_tree.yview)
        self.invalid_auth_tree.configure(yscrollcommand=invalid_scroll.set)
        
        self.invalid_auth_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        invalid_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # RIGHT: Gmail Account & Send (25%)
        email_section = ttk.LabelFrame(bottom_frame, text="Gmail Account & Send", padding="8")
        email_section.pack(side=tk.RIGHT, fill=tk.BOTH, padx=(5, 0))
        email_section.configure(width=350)
        email_section.pack_propagate(False)
        
        # Sender Email Dropdown
        ttk.Label(email_section, text="Send From:").pack(anchor=tk.W, pady=(0, 2))
        
        email_select_frame = ttk.Frame(email_section)
        email_select_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.saved_emails = self.load_saved_emails()
        self.sender_email_var = tk.StringVar()
        self.sender_email_combo = ttk.Combobox(email_select_frame, textvariable=self.sender_email_var, 
                                                values=self.saved_emails, state="readonly", width=24)
        self.sender_email_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        if self.saved_emails:
            self.sender_email_combo.current(0)
        
        # Add/Remove email buttons
        email_btn_frame = ttk.Frame(email_section)
        email_btn_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.add_email_btn = ttk.Button(email_btn_frame, text="➕ Add Email", 
                                         command=self.add_gmail_account, width=12)
        self.add_email_btn.pack(side=tk.LEFT, padx=(0, 5))
        
        self.remove_email_btn = ttk.Button(email_btn_frame, text="➖ Remove", 
                                            command=self.remove_gmail_account, width=10)
        self.remove_email_btn.pack(side=tk.LEFT)
        
        # Gmail API status
        api_status = "✓ Gmail API Ready" if GMAIL_API_AVAILABLE else "✗ Install google-api-python-client"
        api_color = "green" if GMAIL_API_AVAILABLE else "red"
        ttk.Label(email_section, text=api_status, font=("Segoe UI", 8), 
                  foreground=api_color).pack(anchor=tk.W, pady=(0, 8))
        
        # Preview button
        self.preview_attachment_btn = ttk.Button(email_section, text="👁️ Preview Attachment", 
                                                  command=self.preview_email_attachment)
        self.preview_attachment_btn.pack(fill=tk.X, pady=(0, 5))
        
        # Send Email button
        self.send_email_btn = ttk.Button(email_section, text="📧 Send Email", 
                                          command=self.show_send_email_dialog,
                                          style='Action.TButton')
        self.send_email_btn.pack(fill=tk.X, pady=(0, 5))
        
        # Status label
        self.email_status_var = tk.StringVar(value="")
        ttk.Label(email_section, textvariable=self.email_status_var, 
                  font=("Segoe UI", 8), foreground="gray", wraplength=180).pack(pady=(5, 0))
    
    def load_saved_emails(self):
        """Load list of saved Gmail accounts."""
        config_file = APP_DIR / "data" / "gmail_accounts.json"
        try:
            if config_file.exists():
                with open(config_file, 'r') as f:
                    config = json.load(f)
                    return config.get("emails", [])
        except:
            pass
        return []
    
    def save_email_list(self, emails):
        """Save list of Gmail accounts."""
        config_file = APP_DIR / "data" / "gmail_accounts.json"
        try:
            with open(config_file, 'w') as f:
                json.dump({"emails": emails}, f)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save email list: {str(e)}")
    
    def add_gmail_account(self):
        """Add a new Gmail account with OAuth authorization."""
        if not GMAIL_API_AVAILABLE:
            messagebox.showerror("Missing Libraries", 
                "Gmail API libraries not installed.\n\n"
                "Run this command in terminal:\n"
                "pip install google-api-python-client google-auth-oauthlib")
            return
        
        # Check for credentials.json
        creds_file = APP_DIR / "credentials.json"
        if not creds_file.exists():
            messagebox.showerror("Missing Credentials", 
                "credentials.json not found in the application folder.\n\n"
                "To set up Gmail API:\n"
                "1. Go to console.cloud.google.com\n"
                "2. Create a project and enable Gmail API\n"
                "3. Create OAuth 2.0 credentials (Desktop app)\n"
                "4. Download and save as 'credentials.json' in:\n"
                f"   {APP_DIR}")
            return
        
        try:
            # OAuth flow - will open browser for authorization
            SCOPES = ['https://www.googleapis.com/auth/gmail.send',
                      'https://www.googleapis.com/auth/gmail.readonly']
            flow = InstalledAppFlow.from_client_secrets_file(str(creds_file), SCOPES)
            creds = flow.run_local_server(port=0)
            
            # Get the email address from the token
            service = build('gmail', 'v1', credentials=creds)
            profile = service.users().getProfile(userId='me').execute()
            email_address = profile['emailAddress']
            
            # Save the token for this email
            token_file = APP_DIR / "data" / f"token_{email_address.replace('@', '_at_')}.json"
            with open(token_file, 'w') as f:
                f.write(creds.to_json())
            
            # Add to saved emails list
            if email_address not in self.saved_emails:
                self.saved_emails.append(email_address)
                self.save_email_list(self.saved_emails)
                self.sender_email_combo['values'] = self.saved_emails
                self.sender_email_combo.set(email_address)
                messagebox.showinfo("Success", f"Gmail account added: {email_address}")
            else:
                messagebox.showinfo("Already Added", f"Account {email_address} is already in the list.")
                
        except Exception as e:
            messagebox.showerror("Authorization Failed", f"Failed to authorize Gmail:\n{str(e)}")
    
    def remove_gmail_account(self):
        """Remove the selected Gmail account."""
        email = self.sender_email_var.get()
        if not email:
            messagebox.showwarning("No Selection", "No email account selected to remove.")
            return
        
        if messagebox.askyesno("Confirm Remove", f"Remove {email} from the list?"):
            # Remove token file
            token_file = APP_DIR / "data" / f"token_{email.replace('@', '_at_')}.json"
            if token_file.exists():
                token_file.unlink()
            
            # Remove from list
            if email in self.saved_emails:
                self.saved_emails.remove(email)
                self.save_email_list(self.saved_emails)
                self.sender_email_combo['values'] = self.saved_emails
                if self.saved_emails:
                    self.sender_email_combo.current(0)
                else:
                    self.sender_email_var.set("")
    
    def get_gmail_service(self, email):
        """Get authenticated Gmail service for the specified email."""
        if not GMAIL_API_AVAILABLE:
            raise Exception("Gmail API libraries not installed")
        
        token_file = APP_DIR / "data" / f"token_{email.replace('@', '_at_')}.json"
        creds_file = APP_DIR / "credentials.json"
        
        if not token_file.exists():
            raise Exception(f"No authorization token found for {email}. Please re-add the account.")
        
        creds = Credentials.from_authorized_user_file(str(token_file), 
                                                       ['https://www.googleapis.com/auth/gmail.send',
                                                        'https://www.googleapis.com/auth/gmail.readonly'])
        
        # Refresh token if expired
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
            with open(token_file, 'w') as f:
                f.write(creds.to_json())
        
        return build('gmail', 'v1', credentials=creds)
    
    def get_email_attachment_data(self):
        """Generate CSV data for email attachment from invalid auth table."""
        data = []
        for item in self.invalid_auth_tree.get_children():
            values = self.invalid_auth_tree.item(item, 'values')
            if values:
                patient_name = values[0] if len(values) > 0 else ""
                auth_type = values[1] if len(values) > 1 else ""
                data.append({"Patient Name": patient_name, "Auth Type": auth_type})
        return data
    
    def preview_email_attachment(self):
        """Show preview of the CSV attachment with edit capability."""
        popup = tk.Toplevel(self.root)
        popup.title("Attachment Preview - Invalid Auths.csv")
        popup.geometry("550x450")
        popup.transient(self.root)
        self.style_popup(popup)
        
        main_frame = ttk.Frame(popup, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="CSV Attachment Preview (Editable)", 
                  font=("Segoe UI", 12, "bold")).pack(pady=(0, 10))
        
        record_count = len(self.invalid_auth_tree.get_children())
        self.preview_record_label = ttk.Label(main_frame, text=f"File: Invalid_Auths.csv  |  {record_count} records", 
                  font=("Segoe UI", 9))
        self.preview_record_label.pack(anchor=tk.W, pady=(0, 5))
        
        # Toolbar for editing
        toolbar = ttk.Frame(main_frame)
        toolbar.pack(fill=tk.X, pady=(0, 5))
        
        def edit_selected():
            selected = preview_tree.selection()
            if not selected:
                return
            item = selected[0]
            values = preview_tree.item(item, 'values')
            
            edit_popup = tk.Toplevel(popup)
            edit_popup.title("Edit Entry")
            edit_popup.geometry("350x150")
            edit_popup.transient(popup)
            edit_popup.grab_set()
            edit_popup.configure(bg=self.colors['bg'])
            
            ef = ttk.Frame(edit_popup, padding="15")
            ef.pack(fill=tk.BOTH, expand=True)
            
            ttk.Label(ef, text="Patient Name:").grid(row=0, column=0, sticky=tk.W, pady=5)
            name_var = tk.StringVar(value=values[0] if values else "")
            name_entry = ttk.Entry(ef, textvariable=name_var, width=30)
            name_entry.grid(row=0, column=1, pady=5, padx=5)
            
            ttk.Label(ef, text="Auth Type:").grid(row=1, column=0, sticky=tk.W, pady=5)
            type_var = tk.StringVar(value=values[1] if len(values) > 1 else "")
            type_entry = ttk.Entry(ef, textvariable=type_var, width=30)
            type_entry.grid(row=1, column=1, pady=5, padx=5)
            
            def save_edit():
                preview_tree.item(item, values=(name_var.get(), type_var.get()))
                # Also update the main invalid_auth_tree
                for main_item in self.invalid_auth_tree.get_children():
                    main_vals = self.invalid_auth_tree.item(main_item, 'values')
                    if main_vals[0] == values[0]:  # Match by original name
                        reason = main_vals[2] if len(main_vals) > 2 else ""
                        self.invalid_auth_tree.item(main_item, values=(name_var.get(), type_var.get(), reason))
                        break
                edit_popup.destroy()
            
            btn_frame = ttk.Frame(ef)
            btn_frame.grid(row=2, column=0, columnspan=2, pady=(15, 0))
            ttk.Button(btn_frame, text="Save", command=save_edit).pack(side=tk.LEFT, padx=5)
            ttk.Button(btn_frame, text="Cancel", command=edit_popup.destroy).pack(side=tk.LEFT, padx=5)
        
        def delete_selected():
            selected = preview_tree.selection()
            if not selected:
                return
            for item in selected:
                values = preview_tree.item(item, 'values')
                preview_tree.delete(item)
                # Also remove from main invalid_auth_tree
                for main_item in self.invalid_auth_tree.get_children():
                    main_vals = self.invalid_auth_tree.item(main_item, 'values')
                    if main_vals[0] == values[0]:
                        self.invalid_auth_tree.delete(main_item)
                        break
            self.preview_record_label.config(text=f"File: Invalid_Auths.csv  |  {len(preview_tree.get_children())} records")
            self.update_email_counts()
        
        ttk.Button(toolbar, text="✏️ Edit", command=edit_selected).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(toolbar, text="🗑️ Delete", command=delete_selected).pack(side=tk.LEFT, padx=5)
        
        # Preview table
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        preview_tree = ttk.Treeview(tree_frame, columns=["Patient Name", "Auth Type"], 
                                     show="headings", height=15)
        preview_tree.heading("Patient Name", text="Patient Name")
        preview_tree.heading("Auth Type", text="Auth Type")
        preview_tree.column("Patient Name", width=280)
        preview_tree.column("Auth Type", width=150)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=preview_tree.yview)
        preview_tree.configure(yscrollcommand=scrollbar.set)
        
        preview_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load data from invalid_auth_tree
        for item in self.invalid_auth_tree.get_children():
            values = self.invalid_auth_tree.item(item, 'values')
            patient_name = values[0] if len(values) > 0 else ""
            auth_type = values[1] if len(values) > 1 else ""
            preview_tree.insert("", tk.END, values=(patient_name, auth_type))
        
        # Enable double-click to edit
        preview_tree.bind("<Double-1>", lambda e: edit_selected())
        
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        ttk.Button(btn_frame, text="Close", command=popup.destroy).pack(side=tk.RIGHT)
    
    def show_send_email_dialog(self):
        """Show dialog to compose and send email."""
        data = self.get_email_attachment_data()
        
        if not data:
            messagebox.showinfo("No Data", "No patients in the email list to send.")
            return
        
        if not self.sender_email_var.get():
            messagebox.showwarning("No Account", 
                "Please add a Gmail account first using the '+ Add Email' button.")
            return
        
        if not GMAIL_API_AVAILABLE:
            messagebox.showerror("Missing Libraries", 
                "Gmail API libraries not installed.\n\n"
                "Run: pip install google-api-python-client google-auth-oauthlib")
            return
        
        popup = tk.Toplevel(self.root)
        popup.title("Send Email")
        popup.geometry("600x550")
        popup.transient(self.root)
        popup.grab_set()
        self.style_popup(popup)
        
        main_frame = ttk.Frame(popup, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="Compose Email", 
                  font=("Segoe UI", 14, "bold")).pack(pady=(0, 15))
        
        # Email fields
        fields_frame = ttk.Frame(main_frame)
        fields_frame.pack(fill=tk.BOTH, expand=True)
        
        # From
        from_frame = ttk.Frame(fields_frame)
        from_frame.pack(fill=tk.X, pady=5)
        ttk.Label(from_frame, text="From:", width=12, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 10))
        from_entry = ttk.Entry(from_frame, width=50)
        from_entry.insert(0, self.sender_email_var.get())
        from_entry.config(state="readonly")
        from_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # To
        to_frame = ttk.Frame(fields_frame)
        to_frame.pack(fill=tk.X, pady=5)
        ttk.Label(to_frame, text="To:", width=12, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 10))
        to_entry = ttk.Entry(to_frame, width=50)
        to_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Subject
        subject_frame = ttk.Frame(fields_frame)
        subject_frame.pack(fill=tk.X, pady=5)
        ttk.Label(subject_frame, text="Subject:", width=12, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 10))
        subject_entry = ttk.Entry(subject_frame, width=50)
        subject_entry.insert(0, "Patients Without Valid Authorization")
        subject_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Body
        body_frame = ttk.Frame(fields_frame)
        body_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        ttk.Label(body_frame, text="Body:", width=12, anchor=tk.NE).pack(side=tk.LEFT, padx=(0, 10))
        
        body_text = tk.Text(body_frame, height=10, width=50, font=("Segoe UI", 10))
        body_text.insert("1.0", f"""Hello,

Please find attached a list of {len(data)} patient(s) without valid authorization.

These patients require updated authorizations for continued services.

Thank you,
PACE Authorization Team""")
        body_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Attachment info
        attach_frame = ttk.Frame(fields_frame)
        attach_frame.pack(fill=tk.X, pady=10)
        ttk.Label(attach_frame, text="Attachment:", width=12, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Label(attach_frame, text=f"📎 Invalid_Auths.csv ({len(data)} records)", 
                  font=("Segoe UI", 10)).pack(side=tk.LEFT)
        ttk.Button(attach_frame, text="Preview", 
                   command=self.preview_email_attachment).pack(side=tk.LEFT, padx=(10, 0))
        
        # Status
        status_var = tk.StringVar(value="")
        status_label = ttk.Label(main_frame, textvariable=status_var, font=("Segoe UI", 9))
        status_label.pack(pady=(5, 0))
        
        def send_email():
            recipient = to_entry.get().strip()
            subject = subject_entry.get().strip()
            body = body_text.get("1.0", tk.END).strip()
            
            if not recipient:
                messagebox.showwarning("Missing Recipient", "Please enter a recipient email address.")
                return
            
            if not subject:
                messagebox.showwarning("Missing Subject", "Please enter a subject.")
                return
            
            status_var.set("Sending email...")
            popup.update()
            
            try:
                sender_email = self.sender_email_var.get()
                
                # Build the email message
                message_text = f"From: {sender_email}\r\n"
                message_text += f"To: {recipient}\r\n"
                message_text += f"Subject: {subject}\r\n"
                message_text += "MIME-Version: 1.0\r\n"
                message_text += 'Content-Type: multipart/mixed; boundary="boundary"\r\n\r\n'
                message_text += "--boundary\r\n"
                message_text += "Content-Type: text/plain; charset=utf-8\r\n\r\n"
                message_text += body + "\r\n\r\n"
                
                # Create CSV content
                csv_buffer = io.StringIO()
                writer = csv.DictWriter(csv_buffer, fieldnames=["Patient Name", "Auth Type"])
                writer.writeheader()
                writer.writerows(data)
                csv_content = csv_buffer.getvalue()
                
                # Add CSV attachment
                message_text += "--boundary\r\n"
                message_text += "Content-Type: text/csv; name=\"Invalid_Auths.csv\"\r\n"
                message_text += "Content-Disposition: attachment; filename=\"Invalid_Auths.csv\"\r\n"
                message_text += "Content-Transfer-Encoding: base64\r\n\r\n"
                message_text += base64.b64encode(csv_content.encode('utf-8')).decode('utf-8') + "\r\n"
                message_text += "--boundary--"
                
                # Encode the message
                raw_message = base64.urlsafe_b64encode(message_text.encode('utf-8')).decode('utf-8')
                
                # Send via Gmail API
                service = self.get_gmail_service(sender_email)
                service.users().messages().send(
                    userId='me',
                    body={'raw': raw_message}
                ).execute()
                
                status_var.set("✓ Email sent successfully!")
                status_label.config(foreground="green")
                self.email_status_var.set(f"Last sent: {datetime.now().strftime('%H:%M')} to {recipient}")
                
                messagebox.showinfo("Success", f"Email sent successfully to {recipient}")
                popup.destroy()
                
            except Exception as e:
                status_var.set("✗ Failed to send")
                status_label.config(foreground="red")
                messagebox.showerror("Error", f"Failed to send email:\n{str(e)}")
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(15, 0))
        
        ttk.Button(btn_frame, text="📧 Send Email", command=send_email, 
                   style='Action.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=popup.destroy).pack(side=tk.LEFT, padx=5)
    
    def refresh_email_queue(self):
        """Refresh the exception queue with current not-found names, errors, and expired auths."""
        # Clear existing items
        for item in self.email_queue_tree.get_children():
            self.email_queue_tree.delete(item)
        
        count = 0
        
        # Add expired auths from results tree
        if hasattr(self, 'results_tree'):
            for item in self.results_tree.get_children():
                tags = self.results_tree.item(item, 'tags')
                if 'expired' in tags:
                    values = self.results_tree.item(item, 'values')
                    if values:
                        # Get patient name and other info
                        patient_name = ""
                        last_dos = ""
                        auth_type = ""  # Will determine from CPT codes
                        cpt_code = ""
                        
                        # Find patient name and CPT code from values
                        for i, col in enumerate(self.results_columns):
                            if col == "Patient Name" and i < len(values):
                                patient_name = values[i]
                            elif col == "Last Name" and i < len(values) and not patient_name:
                                first_name = ""
                                for j, c in enumerate(self.results_columns):
                                    if c == "First Name" and j < len(values):
                                        first_name = values[j]
                                        break
                                patient_name = f"{values[i]}, {first_name}".strip(", ")
                            elif col == "Last DOS" and i < len(values):
                                last_dos = values[i]
                            elif col == "CPT Code" and i < len(values):
                                cpt_code = values[i]
                            elif col == "Date Auth Expired" and i < len(values):
                                auth_expired = values[i]
                        
                        # Determine auth type from CPT code
                        if cpt_code == "S9123" or cpt_code == "S9124":
                            auth_type = "Skilled"
                        elif cpt_code == "S9122":
                            auth_type = "Unskilled"
                        
                        reason = f"Auth expired"
                        self.email_queue_tree.insert("", tk.END, 
                            values=(patient_name, auth_type, "Extraction", reason),
                            tags=("expired",))
                        count += 1
        
        # Add names not found from file finder
        if hasattr(self, 'finder_not_found'):
            for name, auth_type, reason, last_dos in self.finder_not_found:
                self.email_queue_tree.insert("", tk.END, 
                    values=(name, auth_type, "File Finder", reason),
                    tags=("not_found",))
                count += 1
        
        # Add extraction errors
        if hasattr(self, 'errors_tree'):
            for item in self.errors_tree.get_children():
                values = self.errors_tree.item(item, 'values')
                if values:
                    filename = values[0] if len(values) > 0 else "Unknown"
                    error = values[1] if len(values) > 1 else "Error"
                    auth_type = values[2] if len(values) > 2 else ""
                    # Extract patient name from filename - remove extension, numbers, dates
                    patient_name = filename.replace(".pdf", "").replace(".PDF", "")
                    patient_name = patient_name.replace("_", " ").replace("-", " ")
                    # Remove common patterns like dates, numbers at start/end
                    import re as re_mod
                    patient_name = re_mod.sub(r'\d{1,2}[\-\/]\d{1,2}[\-\/]\d{2,4}', '', patient_name)  # dates
                    patient_name = re_mod.sub(r'^\d+\s*', '', patient_name)  # leading numbers
                    patient_name = re_mod.sub(r'\s*\d+$', '', patient_name)  # trailing numbers
                    patient_name = ' '.join(patient_name.split()).strip()  # normalize spaces
                    
                    # If no auth type stored, try to determine from filename
                    if not auth_type and filename:
                        auth_type = self.get_auth_type_from_filename(filename)
                    
                    self.email_queue_tree.insert("", tk.END, 
                        values=(patient_name, auth_type, "Extraction", error),
                        tags=("error",))
                    count += 1
        
        self.queue_count_var.set(f"{count} items in queue")
    
    def move_to_invalid_auth(self):
        """Move selected queue items to the invalid auth table."""
        selected = self.email_queue_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select items from the queue to move.")
            return
        
        for item in selected:
            values = self.email_queue_tree.item(item, 'values')
            if values:
                patient_name = values[0] if len(values) > 0 else ""
                auth_type = values[1] if len(values) > 1 else ""
                reason = values[3] if len(values) > 3 else ""
                
                # Add to invalid auth table (3 columns: Patient Name, Auth Type, Reason)
                self.invalid_auth_tree.insert("", tk.END,
                    values=(patient_name, auth_type, reason),
                    tags=("confirmed",))
                
                # Remove from queue
                self.email_queue_tree.delete(item)
        
        self.update_email_counts()
    
    def mark_as_fixed(self):
        """Mark selected queue items as fixed/added - removes from queue without adding to email list."""
        selected = self.email_queue_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select items to mark as fixed.")
            return
        
        count = len(selected)
        for item in selected:
            self.email_queue_tree.delete(item)
        
        self.update_email_counts()
        messagebox.showinfo("Marked as Fixed", f"{count} item(s) marked as fixed/added and removed from queue.")

    def add_expired_to_invalid(self):
        """Add all expired auth rows (red highlighted) to the invalid auth table."""
        if not hasattr(self, 'results_tree'):
            messagebox.showinfo("No Data", "Run extraction first to get auth data.")
            return
        
        added = 0
        for item in self.results_tree.get_children():
            tags = self.results_tree.item(item, 'tags')
            if 'expired' in tags:
                values = self.results_tree.item(item, 'values')
                if values:
                    # Get patient name (could be combined or separate first/last)
                    patient_name = ""
                    last_dos = ""
                    
                    if len(self.results_columns) > 0:
                        # Find patient name column
                        for i, col in enumerate(self.results_columns):
                            if col == "Patient Name" and i < len(values):
                                patient_name = values[i]
                            elif col == "Last Name" and i < len(values):
                                patient_name = values[i]
                                if "First Name" in self.results_columns:
                                    fn_idx = self.results_columns.index("First Name")
                                    if fn_idx < len(values):
                                        patient_name = f"{values[fn_idx]} {patient_name}"
                            elif col == "Last DOS" and i < len(values):
                                last_dos = values[i]
                    
                    # Check if already in invalid list
                    already_exists = False
                    for existing in self.invalid_auth_tree.get_children():
                        existing_vals = self.invalid_auth_tree.item(existing, 'values')
                        if existing_vals and existing_vals[0] == patient_name:
                            already_exists = True
                            break
                    
                    if not already_exists and patient_name:
                        self.invalid_auth_tree.insert("", tk.END,
                            values=(patient_name, "", "Auth expired"),
                            tags=("expired",))
                        added += 1
        
        self.update_email_counts()
        if added > 0:
            messagebox.showinfo("Added", f"Added {added} patients with expired authorizations.")
        else:
            messagebox.showinfo("No New Entries", "No new expired auths to add (may already be in list).")
    
    def remove_from_invalid(self):
        """Remove selected items from the invalid auth table."""
        selected = self.invalid_auth_tree.selection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select items to remove.")
            return
        
        for item in selected:
            self.invalid_auth_tree.delete(item)
        
        self.update_email_counts()
    
    def clear_invalid_list(self):
        """Clear all items from the invalid auth table."""
        if not self.invalid_auth_tree.get_children():
            return
        
        if messagebox.askyesno("Confirm Clear", "Clear all items from the invalid auth list?"):
            for item in self.invalid_auth_tree.get_children():
                self.invalid_auth_tree.delete(item)
            self.update_email_counts()
    
    def update_email_counts(self):
        """Update the count labels on the Email tab."""
        queue_count = len(self.email_queue_tree.get_children())
        invalid_count = len(self.invalid_auth_tree.get_children())
        
        self.queue_count_var.set(f"{queue_count} items in queue")
        self.invalid_count_var.set(f"{invalid_count} patients")

    def setup_search_tab(self):
        """Set up the Search tab UI for searching Authorizations and Patients side by side."""
        main_frame = ttk.Frame(self.search_tab, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Description
        desc_label = ttk.Label(
            main_frame,
            text="Search Caspio tables to find patient profiles and existing authorizations",
            style='Desc.TLabel'
        )
        desc_label.pack(anchor=tk.W, pady=(0, 10))
        
        # Create paned window for side-by-side panels
        paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # === LEFT SIDE: Authorizations Panel ===
        auth_frame = ttk.LabelFrame(paned, text="Authorizations (a_Authorizations)", padding="10")
        paned.add(auth_frame, weight=1)
        
        # Auth filter controls
        auth_filter_frame = ttk.Frame(auth_frame)
        auth_filter_frame.pack(fill=tk.X, pady=(0, 8))
        
        # Row 1: Field and Operator
        auth_filter_row1 = ttk.Frame(auth_filter_frame)
        auth_filter_row1.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(auth_filter_row1, text="Field:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        # Field names must match actual Caspio a_Authorizations table columns
        self.auth_search_field_var = tk.StringVar(value="Last_Name")
        auth_field_options = ["Last_Name", "a_First_Name_", "Patient_ID", "Authorization_", "Auth_Start_Date", "Auth_Expire_Date"]
        self.auth_field_combo = ttk.Combobox(auth_filter_row1, textvariable=self.auth_search_field_var, 
                                              values=auth_field_options, state="readonly", width=18)
        self.auth_field_combo.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(auth_filter_row1, text="Operator:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.auth_operator_var = tk.StringVar(value="Contains")
        operator_options = ["Equals", "Contains", "Starts With", "Ends With", "Not Equal", "Is Empty", "Is Not Empty"]
        self.auth_operator_combo = ttk.Combobox(auth_filter_row1, textvariable=self.auth_operator_var, 
                                                 values=operator_options, state="readonly", width=12)
        self.auth_operator_combo.pack(side=tk.LEFT)
        self.auth_operator_combo.bind("<<ComboboxSelected>>", self.on_auth_operator_change)
        
        # Row 2: Value and Search button
        auth_filter_row2 = ttk.Frame(auth_filter_frame)
        auth_filter_row2.pack(fill=tk.X)
        
        ttk.Label(auth_filter_row2, text="Value:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_auth_entry = ttk.Entry(auth_filter_row2, textvariable=self.search_auth_term_var, width=20)
        self.search_auth_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.search_auth_entry.bind("<Return>", lambda e: self.run_auth_search())
        
        self.search_auth_btn = ttk.Button(
            auth_filter_row2, 
            text="🔍 Search", 
            command=self.run_auth_search,
            style='Action.TButton'
        )
        self.search_auth_btn.pack(side=tk.LEFT, padx=(0, 3))
        
        self.clear_auth_btn = ttk.Button(
            auth_filter_row2, 
            text="Clear", 
            command=self.clear_auth_search,
            width=5
        )
        self.clear_auth_btn.pack(side=tk.LEFT)
        
        # Authorizations treeview
        auth_container = ttk.Frame(auth_frame)
        auth_container.pack(fill=tk.BOTH, expand=True)
        
        # Define columns for authorizations - must match actual Caspio field names
        self.search_auth_columns = [
            "Last_Name", "a_First_Name_", "Patient_ID", "Authorization_", 
            "Auth_Start_Date", "Auth_Expire_Date"
        ]
        self.search_auth_pk = "Authorization_ID"  # Primary key field for auth table
        
        # Display name mapping for friendlier column headers
        self.auth_display_names = {
            "Last_Name": "Last Name",
            "a_First_Name_": "First Name",
            "Patient_ID": "Patient ID",
            "Authorization_": "Auth #",
            "Auth_Start_Date": "Date Approved",
            "Auth_Expire_Date": "Date Expired"
        }
        
        self.search_auth_tree = ttk.Treeview(
            auth_container, 
            columns=self.search_auth_columns, 
            show="headings", 
            height=12
        )
        
        # Configure authorization columns
        auth_col_widths = {
            "Last_Name": 100, "a_First_Name_": 90, "Patient_ID": 80, 
            "Authorization_": 95, "Auth_Start_Date": 95, "Auth_Expire_Date": 105
        }
        for col in self.search_auth_columns:
            display_name = self.auth_display_names.get(col, col.replace("_", " "))
            self.search_auth_tree.heading(col, text=display_name)
            self.search_auth_tree.column(col, width=auth_col_widths.get(col, 80), anchor=tk.W)
        
        # Scrollbars for auth tree
        auth_scroll_y = ttk.Scrollbar(auth_container, orient=tk.VERTICAL, command=self.search_auth_tree.yview)
        auth_scroll_x = ttk.Scrollbar(auth_container, orient=tk.HORIZONTAL, command=self.search_auth_tree.xview)
        self.search_auth_tree.configure(yscrollcommand=auth_scroll_y.set, xscrollcommand=auth_scroll_x.set)
        
        self.search_auth_tree.grid(row=0, column=0, sticky="nsew")
        auth_scroll_y.grid(row=0, column=1, sticky="ns")
        auth_scroll_x.grid(row=1, column=0, sticky="ew")
        
        auth_container.grid_rowconfigure(0, weight=1)
        auth_container.grid_columnconfigure(0, weight=1)
        
        # Double-click to edit
        self.search_auth_tree.bind("<Double-1>", lambda e: self.edit_auth_record())
        
        # Auth bottom controls: Count and Action buttons
        auth_bottom = ttk.Frame(auth_frame)
        auth_bottom.pack(fill=tk.X, pady=(5, 0))
        
        self.search_auth_count_var = tk.StringVar(value="0 authorizations found")
        ttk.Label(auth_bottom, textvariable=self.search_auth_count_var, 
                  font=("Segoe UI", 9, "italic")).pack(side=tk.LEFT)
        
        # Action buttons
        auth_btn_frame = ttk.Frame(auth_bottom)
        auth_btn_frame.pack(side=tk.RIGHT)
        
        ttk.Button(auth_btn_frame, text="➕ Add", command=self.add_auth_record, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Button(auth_btn_frame, text="✏️ Edit", command=self.edit_auth_record, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Button(auth_btn_frame, text="🗑️ Delete", command=self.delete_auth_record, width=8).pack(side=tk.LEFT, padx=2)
        
        # === RIGHT SIDE: Patients Panel ===
        patient_frame = ttk.LabelFrame(paned, text="Patients (a_Patient)", padding="10")
        paned.add(patient_frame, weight=1)
        
        # Patient filter controls
        patient_filter_frame = ttk.Frame(patient_frame)
        patient_filter_frame.pack(fill=tk.X, pady=(0, 8))
        
        # Row 1: Field and Operator
        patient_filter_row1 = ttk.Frame(patient_filter_frame)
        patient_filter_row1.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(patient_filter_row1, text="Field:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.patient_search_field_var = tk.StringVar(value="Box_2__Patient_Last_Name")
        patient_field_options = ["Box_2__Patient_Last_Name", "Box_2__Patient_First_Name", "Concantenated_Patient_Name", "Patient_ID"]
        self.patient_field_combo = ttk.Combobox(patient_filter_row1, textvariable=self.patient_search_field_var, 
                                                 values=patient_field_options, state="readonly", width=22)
        self.patient_field_combo.pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Label(patient_filter_row1, text="Operator:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.patient_operator_var = tk.StringVar(value="Contains")
        self.patient_operator_combo = ttk.Combobox(patient_filter_row1, textvariable=self.patient_operator_var, 
                                                    values=operator_options, state="readonly", width=12)
        self.patient_operator_combo.pack(side=tk.LEFT)
        self.patient_operator_combo.bind("<<ComboboxSelected>>", self.on_patient_operator_change)
        
        # Row 2: Value and Search button
        patient_filter_row2 = ttk.Frame(patient_filter_frame)
        patient_filter_row2.pack(fill=tk.X)
        
        ttk.Label(patient_filter_row2, text="Value:", font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=(0, 5))
        
        self.search_patient_entry = ttk.Entry(patient_filter_row2, textvariable=self.search_patient_term_var, width=20)
        self.search_patient_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        self.search_patient_entry.bind("<Return>", lambda e: self.run_patient_search())
        
        self.search_patient_btn = ttk.Button(
            patient_filter_row2, 
            text="🔍 Search", 
            command=self.run_patient_search,
            style='Action.TButton'
        )
        self.search_patient_btn.pack(side=tk.LEFT, padx=(0, 3))
        
        self.clear_patient_btn = ttk.Button(
            patient_filter_row2, 
            text="Clear", 
            command=self.clear_patient_search,
            width=5
        )
        self.clear_patient_btn.pack(side=tk.LEFT)
        
        # Patients treeview
        patient_container = ttk.Frame(patient_frame)
        patient_container.pack(fill=tk.BOTH, expand=True)
        
        # Define columns for patients - display columns
        self.search_patient_columns = [
            "Box_2__Patient_Last_Name", "Box_2__Patient_First_Name", 
            "Concantenated_Patient_Name", "Patient_ID"
        ]
        self.search_patient_pk = "Patient_ID"  # Primary key field for patient table
        
        self.search_patient_tree = ttk.Treeview(
            patient_container, 
            columns=self.search_patient_columns, 
            show="headings", 
            height=12
        )
        
        # Configure patient columns with friendlier display names
        patient_display_names = {
            "Box_2__Patient_Last_Name": "Last Name",
            "Box_2__Patient_First_Name": "First Name",
            "Concantenated_Patient_Name": "Full Name",
            "Patient_ID": "Patient ID"
        }
        patient_col_widths = {
            "Box_2__Patient_Last_Name": 110, "Box_2__Patient_First_Name": 100,
            "Concantenated_Patient_Name": 180, "Patient_ID": 90
        }
        for col in self.search_patient_columns:
            display_name = patient_display_names.get(col, col.replace("_", " "))
            self.search_patient_tree.heading(col, text=display_name)
            self.search_patient_tree.column(col, width=patient_col_widths.get(col, 100), anchor=tk.W)
        
        # Scrollbars for patient tree
        patient_scroll_y = ttk.Scrollbar(patient_container, orient=tk.VERTICAL, command=self.search_patient_tree.yview)
        patient_scroll_x = ttk.Scrollbar(patient_container, orient=tk.HORIZONTAL, command=self.search_patient_tree.xview)
        self.search_patient_tree.configure(yscrollcommand=patient_scroll_y.set, xscrollcommand=patient_scroll_x.set)
        
        self.search_patient_tree.grid(row=0, column=0, sticky="nsew")
        patient_scroll_y.grid(row=0, column=1, sticky="ns")
        patient_scroll_x.grid(row=1, column=0, sticky="ew")
        
        patient_container.grid_rowconfigure(0, weight=1)
        patient_container.grid_columnconfigure(0, weight=1)
        
        # Double-click to edit
        self.search_patient_tree.bind("<Double-1>", lambda e: self.edit_patient_record())
        
        # Patient bottom controls: Count and Action buttons
        patient_bottom = ttk.Frame(patient_frame)
        patient_bottom.pack(fill=tk.X, pady=(5, 0))
        
        self.search_patient_count_var = tk.StringVar(value="0 patients found")
        ttk.Label(patient_bottom, textvariable=self.search_patient_count_var, 
                  font=("Segoe UI", 9, "italic")).pack(side=tk.LEFT)
        
        # Action buttons
        patient_btn_frame = ttk.Frame(patient_bottom)
        patient_btn_frame.pack(side=tk.RIGHT)
        
        ttk.Button(patient_btn_frame, text="➕ Add", command=self.add_patient_record, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Button(patient_btn_frame, text="✏️ Edit", command=self.edit_patient_record, width=7).pack(side=tk.LEFT, padx=2)
        ttk.Button(patient_btn_frame, text="🗑️ Delete", command=self.delete_patient_record, width=8).pack(side=tk.LEFT, padx=2)
        
        # Status bar at bottom
        self.search_status_var = tk.StringVar(value="Select a field, operator, and enter a value to search")
        status_bar = ttk.Label(main_frame, textvariable=self.search_status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(fill=tk.X, pady=(5, 0))
    
    def on_auth_operator_change(self, event=None):
        """Handle operator change for auth search - disable value entry for Is Empty/Is Not Empty."""
        op = self.auth_operator_var.get()
        if op in ("Is Empty", "Is Not Empty"):
            self.search_auth_entry.configure(state='disabled')
            self.search_auth_term_var.set("")
        else:
            self.search_auth_entry.configure(state='normal')
    
    def on_patient_operator_change(self, event=None):
        """Handle operator change for patient search - disable value entry for Is Empty/Is Not Empty."""
        op = self.patient_operator_var.get()
        if op in ("Is Empty", "Is Not Empty"):
            self.search_patient_entry.configure(state='disabled')
            self.search_patient_term_var.set("")
        else:
            self.search_patient_entry.configure(state='normal')
    
    def _operator_to_api(self, operator):
        """Convert UI operator name to API operator string."""
        mapping = {
            "Equals": "equals",
            "Contains": "contains",
            "Starts With": "starts_with",
            "Ends With": "ends_with",
            "Not Equal": "not_equal",
            "Is Empty": "is_empty",
            "Is Not Empty": "is_not_empty"
        }
        return mapping.get(operator, "contains")
    
    def run_auth_search(self):
        """Execute search against the Authorizations table using selected field and operator."""
        field = self.auth_search_field_var.get()
        operator = self.auth_operator_var.get()
        value = self.search_auth_term_var.get().strip()
        
        # For Is Empty / Is Not Empty, value is not needed
        if operator not in ("Is Empty", "Is Not Empty"):
            if not value:
                messagebox.showwarning("Warning", "Please enter a search value")
                return
            if len(value) < 2:
                messagebox.showwarning("Warning", "Please enter at least 2 characters")
                return
        
        self.search_status_var.set(f"Searching authorizations: {field} {operator} '{value}'...")
        self.search_auth_btn.configure(state='disabled')
        self.root.update_idletasks()
        
        api_operator = self._operator_to_api(operator)
        
        def search_thread():
            try:
                caspio = CaspioAPI()
                # Include PK_ID in select so we can edit/delete
                select_fields = [self.search_auth_pk] + self.search_auth_columns
                auth_results = caspio.search_with_operator(
                    "a_Authorizations",
                    field,
                    api_operator,
                    value,
                    select_fields=select_fields
                )
                self.root.after(0, lambda: self.update_auth_results(auth_results))
            except Exception as e:
                self.root.after(0, lambda: self.auth_search_error(str(e)))
        
        Thread(target=search_thread, daemon=True).start()
    
    def run_patient_search(self):
        """Execute search against the Patient table using selected field and operator."""
        field = self.patient_search_field_var.get()
        operator = self.patient_operator_var.get()
        value = self.search_patient_term_var.get().strip()
        
        # For Is Empty / Is Not Empty, value is not needed
        if operator not in ("Is Empty", "Is Not Empty"):
            if not value:
                messagebox.showwarning("Warning", "Please enter a search value")
                return
            if len(value) < 2:
                messagebox.showwarning("Warning", "Please enter at least 2 characters")
                return
        
        self.search_status_var.set(f"Searching patients: {field} {operator} '{value}'...")
        self.search_patient_btn.configure(state='disabled')
        self.root.update_idletasks()
        
        api_operator = self._operator_to_api(operator)
        
        def search_thread():
            try:
                caspio = CaspioAPI()
                # Include PK_ID in select so we can edit/delete
                select_fields = [self.search_patient_pk] + self.search_patient_columns
                patient_results = caspio.search_with_operator(
                    "a_Patient",
                    field,
                    api_operator,
                    value,
                    select_fields=select_fields
                )
                self.root.after(0, lambda: self.update_patient_results(patient_results))
            except Exception as e:
                self.root.after(0, lambda: self.patient_search_error(str(e)))
        
        Thread(target=search_thread, daemon=True).start()
    
    def update_auth_results(self, auth_results):
        """Update the authorization search results treeview."""
        # Clear existing results
        for item in self.search_auth_tree.get_children():
            self.search_auth_tree.delete(item)
        
        # Populate authorizations
        self.search_auths_results = auth_results
        for record in auth_results:
            values = [record.get(col, "") or "" for col in self.search_auth_columns]
            self.search_auth_tree.insert("", tk.END, values=values)
        
        # Update count
        self.search_auth_count_var.set(f"{len(auth_results)} authorization(s) found")
        self.search_status_var.set(f"Authorization search complete: {len(auth_results)} found")
        self.search_auth_btn.configure(state='normal')
    
    def update_patient_results(self, patient_results):
        """Update the patient search results treeview."""
        # Clear existing results
        for item in self.search_patient_tree.get_children():
            self.search_patient_tree.delete(item)
        
        # Populate patients
        self.search_patients_results = patient_results
        for record in patient_results:
            values = [record.get(col, "") or "" for col in self.search_patient_columns]
            self.search_patient_tree.insert("", tk.END, values=values)
        
        # Update count
        self.search_patient_count_var.set(f"{len(patient_results)} patient(s) found")
        self.search_status_var.set(f"Patient search complete: {len(patient_results)} found")
        self.search_patient_btn.configure(state='normal')
    
    def auth_search_error(self, error_msg):
        """Handle authorization search errors."""
        self.search_status_var.set(f"Auth search error: {error_msg}")
        self.search_auth_btn.configure(state='normal')
        messagebox.showerror("Search Error", f"Failed to search authorizations:\n\n{error_msg}")
    
    def patient_search_error(self, error_msg):
        """Handle patient search errors."""
        self.search_status_var.set(f"Patient search error: {error_msg}")
        self.search_patient_btn.configure(state='normal')
        messagebox.showerror("Search Error", f"Failed to search patients:\n\n{error_msg}")
    
    def clear_auth_search(self):
        """Clear authorization search results."""
        self.search_auth_term_var.set("")
        for item in self.search_auth_tree.get_children():
            self.search_auth_tree.delete(item)
        self.search_auths_results = []
        self.search_auth_count_var.set("0 authorizations found")
        self.search_status_var.set("Authorization results cleared")
    
    def clear_patient_search(self):
        """Clear patient search results."""
        self.search_patient_term_var.set("")
        for item in self.search_patient_tree.get_children():
            self.search_patient_tree.delete(item)
        self.search_patients_results = []
        self.search_patient_count_var.set("0 patients found")
        self.search_status_var.set("Patient results cleared")
    
    # ===== Edit/Add/Delete Record Functions =====
    
    def _get_selected_auth_record(self):
        """Get the selected authorization record and its index."""
        selection = self.search_auth_tree.selection()
        if not selection:
            return None, None
        
        item = selection[0]
        idx = self.search_auth_tree.index(item)
        if idx < len(self.search_auths_results):
            return self.search_auths_results[idx], idx
        return None, None
    
    def _get_selected_patient_record(self):
        """Get the selected patient record and its index."""
        selection = self.search_patient_tree.selection()
        if not selection:
            return None, None
        
        item = selection[0]
        idx = self.search_patient_tree.index(item)
        if idx < len(self.search_patients_results):
            return self.search_patients_results[idx], idx
        return None, None
    
    def edit_auth_record(self):
        """Edit the selected authorization record."""
        record, idx = self._get_selected_auth_record()
        if not record:
            messagebox.showwarning("Warning", "Please select an authorization to edit")
            return
        
        pk_value = record.get(self.search_auth_pk)
        if not pk_value:
            messagebox.showerror("Error", "Cannot edit: record has no primary key")
            return
        
        # Show edit dialog - pass None for fields to fetch all from schema
        self._show_edit_dialog(
            "Edit Authorization",
            "a_Authorizations",
            self.search_auth_pk,
            pk_value,
            record,
            None,  # Will fetch all fields from table schema
            self.run_auth_search  # Refresh callback
        )
    
    def edit_patient_record(self):
        """Edit the selected patient record."""
        record, idx = self._get_selected_patient_record()
        if not record:
            messagebox.showwarning("Warning", "Please select a patient to edit")
            return
        
        pk_value = record.get(self.search_patient_pk)
        if not pk_value:
            messagebox.showerror("Error", "Cannot edit: record has no primary key")
            return
        
        # Show edit dialog - pass None for fields to fetch all from schema
        self._show_edit_dialog(
            "Edit Patient",
            "a_Patient",
            self.search_patient_pk,
            pk_value,
            record,
            None,  # Will fetch all fields from table schema
            self.run_patient_search  # Refresh callback
        )
    
    def _show_edit_dialog(self, title, table_name, pk_field, pk_value, record, fields, refresh_callback):
        """Show a dialog to edit a record. If fields is None, fetches all fields from table schema."""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center dialog - larger to accommodate all fields
        dialog.geometry("600x600")
        dialog.resizable(True, True)
        
        # Apply dark theme
        dialog.configure(bg='#1a1d21')
        
        # Fetch all fields from table schema if not provided
        if fields is None:
            try:
                caspio = CaspioAPI()
                schema = caspio.get_table_schema(table_name)
                fields = [f["name"] for f in schema]
                readonly_fields = {f["name"] for f in schema if f.get("readonly", False)}
            except Exception as e:
                messagebox.showerror("Error", f"Failed to fetch table schema:\n{e}")
                dialog.destroy()
                return
        else:
            readonly_fields = set()
        
        # We also need the full record data - fetch it if we only have display columns
        if record and pk_value:
            try:
                caspio = CaspioAPI()
                # Fetch the full record
                import urllib.parse
                where_clause = f"{pk_field}='{pk_value}'"
                encoded_where = urllib.parse.quote(where_clause)
                url = f"{caspio.base_url}/tables/{table_name}/records?q.where={encoded_where}"
                headers = {
                    "Authorization": f"Bearer {caspio.get_access_token()}",
                    "Content-Type": "application/json"
                }
                response = requests.get(url, headers=headers)
                if response.status_code == 200:
                    data = response.json()
                    results = data.get("Result", [])
                    if results:
                        record = results[0]  # Use the full record
            except Exception as e:
                pass  # Continue with partial record
        
        # Create scrollable frame for fields
        canvas = tk.Canvas(dialog, bg='#1a1d21', highlightthickness=0)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Entry variables
        entries = {}
        
        for i, field in enumerate(fields):
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, pady=3, padx=5)
            
            # Display friendly field name
            display_name = field.replace("_", " ").replace("Box 2  Patient", "").strip()
            is_readonly = field in readonly_fields or field == pk_field
            
            label_text = f"{display_name}:" + (" (read-only)" if is_readonly else "")
            ttk.Label(frame, text=label_text, width=28, anchor='e').pack(side=tk.LEFT, padx=(0, 10))
            
            var = tk.StringVar(value=str(record.get(field, "") or ""))
            entry = ttk.Entry(frame, textvariable=var, width=45)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            
            if is_readonly:
                entry.configure(state='disabled')
            else:
                entries[field] = var
        
        # Unbind mousewheel when dialog closes
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_close)
        
        # Buttons frame
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, pady=10, padx=10)
        
        def save_changes():
            # Collect updated values
            updates = {}
            for field, var in entries.items():
                new_val = var.get().strip()
                old_val = str(record.get(field, "") or "")
                if new_val != old_val:
                    updates[field] = new_val
            
            if not updates:
                messagebox.showinfo("Info", "No changes to save")
                on_close()
                return
            
            # Update in Caspio
            self.search_status_var.set("Saving changes to Caspio...")
            dialog.update_idletasks()
            
            def update_thread():
                try:
                    caspio = CaspioAPI()
                    result = caspio.update_record(table_name, pk_field, pk_value, updates)
                    
                    if result["success"]:
                        self.root.after(0, lambda: self._on_edit_success_cleanup(on_close, refresh_callback))
                    else:
                        self.root.after(0, lambda: self._on_edit_error(result["message"]))
                except Exception as e:
                    self.root.after(0, lambda: self._on_edit_error(str(e)))
            
            Thread(target=update_thread, daemon=True).start()
        
        ttk.Button(btn_frame, text="💾 Save Changes", command=save_changes, style='Action.TButton').pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=on_close).pack(side=tk.RIGHT, padx=5)
    
    def _on_edit_success_cleanup(self, cleanup_fn, refresh_callback):
        """Handle successful edit with cleanup."""
        cleanup_fn()
        messagebox.showinfo("Success", "Record updated successfully in Caspio")
        self.search_status_var.set("Record updated successfully")
        refresh_callback()
    
    def _on_edit_error(self, error_msg):
        """Handle edit error."""
        messagebox.showerror("Update Error", f"Failed to update record:\n\n{error_msg}")
        self.search_status_var.set(f"Update failed: {error_msg}")
    
    def add_auth_record(self):
        """Add a new authorization record."""
        self._show_add_dialog(
            "Add Authorization",
            "a_Authorizations",
            None,  # Will fetch all fields from table schema
            self.run_auth_search
        )
    
    def add_patient_record(self):
        """Add a new patient record."""
        self._show_add_dialog(
            "Add Patient",
            "a_Patient",
            None,  # Will fetch all fields from table schema
            self.run_patient_search
        )
    
    def _show_add_dialog(self, title, table_name, fields, refresh_callback):
        """Show a dialog to add a new record. If fields is None, fetches all fields from table schema."""
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.transient(self.root)
        dialog.grab_set()
        
        dialog.geometry("600x600")
        dialog.resizable(True, True)
        dialog.configure(bg='#1a1d21')
        
        # Fetch all fields from table schema if not provided
        if fields is None:
            try:
                caspio = CaspioAPI()
                schema = caspio.get_table_schema(table_name)
                # Exclude auto-generated fields for adding
                fields = [f["name"] for f in schema if not f.get("readonly", False)]
                readonly_fields = {f["name"] for f in schema if f.get("readonly", False)}
            except Exception as e:
                messagebox.showerror("Error", f"Failed to fetch table schema:\n{e}")
                dialog.destroy()
                return
        else:
            readonly_fields = set()
        
        # Create scrollable frame
        canvas = tk.Canvas(dialog, bg='#1a1d21', highlightthickness=0)
        scrollbar = ttk.Scrollbar(dialog, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y")
        
        # Enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        entries = {}
        
        for i, field in enumerate(fields):
            if field in readonly_fields:
                continue  # Skip read-only fields for add
                
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, pady=3, padx=5)
            
            display_name = field.replace("_", " ").replace("Box 2  Patient", "").strip()
            ttk.Label(frame, text=f"{display_name}:", width=28, anchor='e').pack(side=tk.LEFT, padx=(0, 10))
            
            var = tk.StringVar()
            entry = ttk.Entry(frame, textvariable=var, width=45)
            entry.pack(side=tk.LEFT, fill=tk.X, expand=True)
            entries[field] = var
        
        # Unbind mousewheel when dialog closes
        def on_close():
            canvas.unbind_all("<MouseWheel>")
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_close)
        
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, pady=10, padx=10)
        
        def add_record():
            # Collect values
            new_record = {}
            for field, var in entries.items():
                val = var.get().strip()
                if val:
                    new_record[field] = val
            
            if not new_record:
                messagebox.showwarning("Warning", "Please enter at least one field value")
                return
            
            self.search_status_var.set("Adding record to Caspio...")
            dialog.update_idletasks()
            
            def insert_thread():
                try:
                    caspio = CaspioAPI()
                    result = caspio.insert_single_record(table_name, new_record)
                    
                    if result["success"]:
                        self.root.after(0, lambda: self._on_add_success(dialog, refresh_callback))
                    else:
                        self.root.after(0, lambda: self._on_add_error(result["message"]))
                except Exception as e:
                    self.root.after(0, lambda: self._on_add_error(str(e)))
            
            Thread(target=insert_thread, daemon=True).start()
        
        ttk.Button(btn_frame, text="➕ Add Record", command=add_record, style='Action.TButton').pack(side=tk.RIGHT, padx=5)
        ttk.Button(btn_frame, text="Cancel", command=on_close).pack(side=tk.RIGHT, padx=5)
    
    def _on_add_success(self, dialog, refresh_callback):
        """Handle successful add."""
        try:
            dialog.destroy()
        except:
            pass
        messagebox.showinfo("Success", "Record added successfully to Caspio")
        self.search_status_var.set("Record added successfully")
        refresh_callback()
    
    def _on_add_error(self, error_msg):
        """Handle add error."""
        messagebox.showerror("Insert Error", f"Failed to add record:\n\n{error_msg}")
        self.search_status_var.set(f"Add failed: {error_msg}")
    
    def delete_auth_record(self):
        """Delete the selected authorization record."""
        record, idx = self._get_selected_auth_record()
        if not record:
            messagebox.showwarning("Warning", "Please select an authorization to delete")
            return
        
        pk_value = record.get(self.search_auth_pk)
        if not pk_value:
            messagebox.showerror("Error", "Cannot delete: record has no primary key")
            return
        
        # Confirmation - use correct field names
        name = f"{record.get('Last_Name', '')}, {record.get('a_First_Name_', '')}"
        auth_num = record.get('Authorization_', '')
        if not messagebox.askyesno("Confirm Delete", 
            f"Are you sure you want to delete this authorization?\n\n"
            f"Patient: {name}\nAuth #: {auth_num}\n\n"
            f"This action cannot be undone!"):
            return
        
        self._delete_record("a_Authorizations", self.search_auth_pk, pk_value, self.run_auth_search)
    
    def delete_patient_record(self):
        """Delete the selected patient record."""
        record, idx = self._get_selected_patient_record()
        if not record:
            messagebox.showwarning("Warning", "Please select a patient to delete")
            return
        
        pk_value = record.get(self.search_patient_pk)
        if not pk_value:
            messagebox.showerror("Error", "Cannot delete: record has no primary key")
            return
        
        # Confirmation
        name = f"{record.get('Box_2__Patient_Last_Name', '')}, {record.get('Box_2__Patient_First_Name', '')}"
        patient_id = record.get('Patient_ID', '')
        if not messagebox.askyesno("Confirm Delete", 
            f"Are you sure you want to delete this patient?\n\n"
            f"Name: {name}\nPatient ID: {patient_id}\n\n"
            f"This action cannot be undone!"):
            return
        
        self._delete_record("a_Patient", self.search_patient_pk, pk_value, self.run_patient_search)
    
    def _delete_record(self, table_name, pk_field, pk_value, refresh_callback):
        """Delete a record from Caspio."""
        self.search_status_var.set("Deleting record from Caspio...")
        self.root.update_idletasks()
        
        def delete_thread():
            try:
                caspio = CaspioAPI()
                result = caspio.delete_record(table_name, pk_field, pk_value)
                
                if result["success"]:
                    self.root.after(0, lambda: self._on_delete_success(refresh_callback))
                else:
                    self.root.after(0, lambda: self._on_delete_error(result["message"]))
            except Exception as e:
                self.root.after(0, lambda: self._on_delete_error(str(e)))
        
        Thread(target=delete_thread, daemon=True).start()
    
    def _on_delete_success(self, refresh_callback):
        """Handle successful delete."""
        messagebox.showinfo("Success", "Record deleted successfully from Caspio")
        self.search_status_var.set("Record deleted successfully")
        refresh_callback()
    
    def _on_delete_error(self, error_msg):
        """Handle delete error."""
        messagebox.showerror("Delete Error", f"Failed to delete record:\n\n{error_msg}")
        self.search_status_var.set(f"Delete failed: {error_msg}")
    
    def browse_finder_source(self):
        """Browse for source folder."""
        folder = filedialog.askdirectory(title="Select the folder to search for PDFs")
        if folder:
            self.finder_source.set(folder)
            pdf_count = len(list(pathlib.Path(folder).glob("*.pdf")))
            self.finder_log_msg(f"📁 Source: {folder}")
            self.finder_log_msg(f"   Found {pdf_count} PDF files")

    def browse_finder_dest(self):
        """Browse for destination folder."""
        folder = filedialog.askdirectory(title="Select the destination folder for matched files")
        if folder:
            self.finder_dest.set(folder)
            self.finder_log_msg(f"📁 Destination: {folder}")
    
    def add_finder_row(self):
        """Add a row to the finder table."""
        name = self.finder_name_entry.get().strip()
        auth_type = self.finder_type_combo.get().strip()
        
        if not name:
            messagebox.showwarning("Warning", "Please enter a patient name")
            return
        
        if not auth_type:
            messagebox.showwarning("Warning", "Please select an auth type")
            return
        
        # Get what this auth type searches for (for display)
        keywords = self.get_keywords_for_type(auth_type)
        searches_for = ", ".join(keywords).title() if keywords else "Unknown"
        
        # Manual entries don't have a last DOS - leave blank
        last_dos = ""
        
        # Add to table
        self.finder_table.insert("", tk.END, values=(name, auth_type, searches_for, last_dos))
        
        # Clear the name entry for next input
        self.finder_name_entry.delete(0, tk.END)
        self.finder_name_entry.focus()
    
    def remove_finder_row(self):
        """Remove selected row(s) from the finder table."""
        selected = self.finder_table.selection()
        if not selected:
            messagebox.showwarning("Warning", "Please select a row to remove")
            return
        
        for item in selected:
            self.finder_table.delete(item)
    
    def bulk_import_finder(self):
        """Open dialog for bulk importing names with structured columns."""
        # Create popup window - larger size for visibility
        popup = tk.Toplevel(self.root)
        popup.title("Bulk Import - Patient Search Criteria")
        popup.geometry("800x600")
        popup.transient(self.root)
        popup.grab_set()
        self.style_popup(popup)
        
        # Center the popup
        popup.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (800 // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (600 // 2)
        popup.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(popup, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Instructions
        instructions = ttk.Label(main_frame, 
            text="Paste data into each column. You can paste a list of names, types, and dates separately.\n" +
                 "Tip: Copy a column from Excel and paste directly into the text area for that column.",
            font=("Segoe UI", 10), justify=tk.LEFT)
        instructions.pack(anchor=tk.W, pady=(0, 15))
        
        # Three-column layout
        columns_frame = ttk.Frame(main_frame)
        columns_frame.pack(fill=tk.BOTH, expand=True)
        
        # Configure grid weights
        columns_frame.columnconfigure(0, weight=3)  # Name column wider
        columns_frame.columnconfigure(1, weight=1)
        columns_frame.columnconfigure(2, weight=1)
        columns_frame.rowconfigure(1, weight=1)
        
        # Column 1: Patient Names
        name_label = ttk.Label(columns_frame, text="Patient Names", font=("Segoe UI", 10, "bold"))
        name_label.grid(row=0, column=0, sticky="w", padx=(0, 5), pady=(0, 5))
        
        name_frame = ttk.Frame(columns_frame)
        name_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 5))
        
        name_text = tk.Text(name_frame, font=("Consolas", 10), width=35)
        name_scroll = ttk.Scrollbar(name_frame, orient=tk.VERTICAL, command=name_text.yview)
        name_text.config(yscrollcommand=name_scroll.set)
        name_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        name_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Column 2: Auth Type
        type_frame_outer = ttk.Frame(columns_frame)
        type_frame_outer.grid(row=0, column=1, rowspan=2, sticky="nsew", padx=5)
        
        type_label = ttk.Label(type_frame_outer, text="Auth Type", font=("Segoe UI", 10, "bold"))
        type_label.pack(anchor=tk.W, pady=(0, 5))
        
        type_inner = ttk.Frame(type_frame_outer)
        type_inner.pack(fill=tk.BOTH, expand=True)
        
        type_text = tk.Text(type_inner, font=("Consolas", 10), width=12)
        type_scroll = ttk.Scrollbar(type_inner, orient=tk.VERTICAL, command=type_text.yview)
        type_text.config(yscrollcommand=type_scroll.set)
        type_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        type_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Type helper label
        type_hint = ttk.Label(type_frame_outer, 
            text="Enter: Unskilled, Skilled,\nor Escort (one per line)",
            font=("Segoe UI", 8), foreground="gray")
        type_hint.pack(anchor=tk.W, pady=(5, 0))
        
        # Column 3: Last DOS
        dos_frame_outer = ttk.Frame(columns_frame)
        dos_frame_outer.grid(row=0, column=2, rowspan=2, sticky="nsew", padx=(5, 0))
        
        dos_label = ttk.Label(dos_frame_outer, text="Last DOS", font=("Segoe UI", 10, "bold"))
        dos_label.pack(anchor=tk.W, pady=(0, 5))
        
        dos_inner = ttk.Frame(dos_frame_outer)
        dos_inner.pack(fill=tk.BOTH, expand=True)
        
        dos_text = tk.Text(dos_inner, font=("Consolas", 10), width=12)
        dos_scroll = ttk.Scrollbar(dos_inner, orient=tk.VERTICAL, command=dos_text.yview)
        dos_text.config(yscrollcommand=dos_scroll.set)
        dos_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        dos_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # DOS helper label
        dos_hint = ttk.Label(dos_frame_outer, 
            text="Format: MM/DD/YYYY\n(one per line)",
            font=("Segoe UI", 8), foreground="gray")
        dos_hint.pack(anchor=tk.W, pady=(5, 0))
        
        # Quick fill options for Auth Type
        quick_frame = ttk.LabelFrame(main_frame, text="Quick Fill Auth Type", padding="10")
        quick_frame.pack(fill=tk.X, pady=(15, 10))
        
        def fill_type_column(type_value):
            """Fill the type column with a value for all name rows."""
            names = name_text.get(1.0, tk.END).strip().split("\n")
            names = [n.strip() for n in names if n.strip()]
            if names:
                type_text.delete(1.0, tk.END)
                type_text.insert(1.0, "\n".join([type_value] * len(names)))
            else:
                messagebox.showwarning("Warning", "Paste patient names first", parent=popup)
        
        ttk.Button(quick_frame, text="Fill All Unskilled", 
                   command=lambda: fill_type_column("Unskilled")).pack(side=tk.LEFT, padx=5)
        ttk.Button(quick_frame, text="Fill All Skilled", 
                   command=lambda: fill_type_column("Skilled")).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(quick_frame, text="(Click after pasting names to auto-fill the type column)", 
                  font=("Segoe UI", 8), foreground="gray").pack(side=tk.LEFT, padx=(15, 0))
        
        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        def do_import():
            # Get data from each column
            names = name_text.get(1.0, tk.END).strip().split("\n")
            types = type_text.get(1.0, tk.END).strip().split("\n")
            dates = dos_text.get(1.0, tk.END).strip().split("\n")
            
            # Clean up empty lines
            names = [n.strip() for n in names if n.strip()]
            types = [t.strip() for t in types]
            dates = [d.strip() for d in dates]
            
            if not names:
                messagebox.showwarning("Warning", "Please paste patient names first", parent=popup)
                return
            
            # Extend types and dates arrays to match names length
            while len(types) < len(names):
                types.append("")
            while len(dates) < len(names):
                dates.append("")
            
            added = 0
            errors = []
            
            for i, name in enumerate(names):
                if not name:
                    continue
                
                # Get type for this row
                auth_type = types[i].strip() if i < len(types) else ""
                last_dos = dates[i].strip() if i < len(dates) else ""
                
                # Normalize auth type
                auth_type_lower = auth_type.lower()
                if 'unskilled' in auth_type_lower or auth_type_lower in ['u', 'un', 'uns']:
                    auth_type = "Unskilled"
                elif auth_type_lower == 'skilled' or auth_type_lower in ['s', 'sk', 'ski']:
                    # Must be exact 'skilled' to avoid matching 'unskilled'
                    auth_type = "Skilled"
                elif 'escort' in auth_type_lower or auth_type_lower in ['e', 'es', 'esc']:
                    auth_type = "Escort"
                elif auth_type == "":
                    errors.append(f"Row {i+1}: '{name[:30]}' - missing auth type")
                    continue
                else:
                    errors.append(f"Row {i+1}: '{name[:30]}' - unknown type '{auth_type}'")
                    continue
                
                # Add to table
                keywords = self.get_keywords_for_type(auth_type)
                searches_for = ", ".join(keywords).title() if keywords else "Unknown"
                self.finder_table.insert("", tk.END, values=(name, auth_type, searches_for, last_dos))
                added += 1
            
            # Show result
            if added > 0:
                msg = f"Successfully added {added} entries to the table."
                if errors:
                    msg += f"\n\n{len(errors)} row(s) had issues:\n" + "\n".join(errors[:5])
                    if len(errors) > 5:
                        msg += f"\n... and {len(errors) - 5} more"
                messagebox.showinfo("Bulk Import Complete", msg, parent=popup)
                popup.destroy()
            else:
                messagebox.showerror("Error", 
                    "Could not import any entries.\n\n" +
                    "Make sure you have:\n" +
                    "1. Pasted patient names in the first column\n" +
                    "2. Either used 'Fill All' buttons or pasted types in the second column\n" +
                    "3. Valid auth types: Unskilled, Skilled, or Escort", parent=popup)
        
        import_btn = ttk.Button(btn_frame, text="✅ Import All", command=do_import, style='Action.TButton')
        import_btn.pack(side=tk.LEFT, padx=5)
        
        clear_btn = ttk.Button(btn_frame, text="🗑️ Clear All", 
                               command=lambda: [name_text.delete(1.0, tk.END), 
                                                type_text.delete(1.0, tk.END), 
                                                dos_text.delete(1.0, tk.END)])
        clear_btn.pack(side=tk.LEFT, padx=5)
        
        cancel_btn = ttk.Button(btn_frame, text="Cancel", command=popup.destroy)
        cancel_btn.pack(side=tk.LEFT, padx=5)
        
        # Row count helper
        def update_row_count(event=None):
            names = name_text.get(1.0, tk.END).strip().split("\n")
            names = [n for n in names if n.strip()]
            count_label.config(text=f"Names: {len(names)} rows")
        
        count_label = ttk.Label(btn_frame, text="Names: 0 rows", font=("Segoe UI", 9), foreground="gray")
        count_label.pack(side=tk.RIGHT, padx=10)
        
        name_text.bind("<KeyRelease>", update_row_count)
        name_text.bind("<<Paste>>", lambda e: popup.after(50, update_row_count))
        
        # Focus the text area
        name_text.focus()
    
    def clear_finder(self):
        """Clear the finder table and log."""
        # Clear the table
        for item in self.finder_table.get_children():
            self.finder_table.delete(item)
        # Clear entry fields
        self.finder_name_entry.delete(0, tk.END)
        self.finder_type_combo.set("Unskilled")
        # Clear status
        if hasattr(self, 'finder_status_var'):
            self.finder_status_var.set("Ready to search")
    
    def open_finder_dest(self):
        """Open the destination folder."""
        dest = self.finder_dest.get()
        if dest and pathlib.Path(dest).exists():
            os.startfile(dest)
        else:
            messagebox.showwarning("Warning", "Please select a destination folder first")
    
    def finder_log_msg(self, message):
        """Update finder status message."""
        # Update status line instead of log
        # Extract just the key info for status display
        if hasattr(self, 'finder_status_var'):
            self.finder_status_var.set(message)
        self.root.update_idletasks()

    def get_keywords_for_type(self, auth_type):
        """Get search keywords based on auth type."""
        auth = auth_type.strip().lower()
        if auth == "unskilled":
            return ["unskilled"]
        elif auth == "skilled":
            return ["skilled"]
        elif auth == "escort assistance":
            return ["escort assistance"]
        elif auth == "escort":
            return ["escort"]
        else:
            return []
    
    def parse_search_criteria(self, text):
        """Parse the search criteria text into list of (name, cpt_code) tuples."""
        criteria = []
        for line in text.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            
            # Try to split by comma, tab, or multiple spaces
            parts = None
            if "," in line:
                parts = [p.strip() for p in line.split(",", 1)]
            elif "\t" in line:
                parts = [p.strip() for p in line.split("\t", 1)]
            else:
                # Try splitting by multiple spaces
                parts = line.rsplit(None, 1)
            
            if parts and len(parts) >= 2:
                name = parts[0].strip()
                cpt = parts[1].strip()
                if name and cpt:
                    criteria.append((name, cpt))
        
        return criteria
    
    def run_file_finder(self):
        """Run the file finder process."""
        # Route to Dropbox handler if Dropbox source selected
        if self.finder_source_type.get() == "dropbox":
            self._run_finder_from_dropbox()
            return

        source = self.finder_source.get()
        dest = self.finder_dest.get()
        
        if not source:
            messagebox.showerror("Error", "Please select a source folder")
            return
        
        if not dest:
            messagebox.showerror("Error", "Please select a destination folder")
            return
        
        # Get criteria from table (now includes last_dos)
        criteria = []
        for item in self.finder_table.get_children():
            values = self.finder_table.item(item, 'values')
            name = values[0]
            auth_type = values[1]
            last_dos = values[3] if len(values) > 3 else ""
            criteria.append((name, auth_type, last_dos))
        
        if not criteria:
            messagebox.showerror("Error", "Please add at least one search criteria to the table")
            return
        
        # Reset status
        if hasattr(self, 'finder_status_var'):
            self.finder_status_var.set("Searching...")
        
        self.finder_log_msg(f"🔍 Searching with {len(criteria)} criteria...")
        
        # Parse date range
        date_from = self.finder_date_from.get().strip()
        date_to = self.finder_date_to.get().strip()
        
        from_date = None
        to_date = None
        
        if date_from:
            try:
                from_date = datetime.strptime(date_from, "%m/%d/%Y")
                self.finder_log_msg(f"📅 From date: {from_date.strftime('%m/%d/%Y')}")
            except ValueError:
                messagebox.showerror("Error", f"Invalid 'From' date format: {date_from}\nUse MM/DD/YYYY")
                return
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, "%m/%d/%Y")
                # Set to end of day for inclusive range
                to_date = to_date.replace(hour=23, minute=59, second=59)
                self.finder_log_msg(f"📅 To date: {to_date.strftime('%m/%d/%Y')}")
            except ValueError:
                messagebox.showerror("Error", f"Invalid 'To' date format: {date_to}\nUse MM/DD/YYYY")
                return
        
        # Get all PDFs in source folder
        source_path = pathlib.Path(source)
        dest_path = pathlib.Path(dest)
        all_pdf_files = list(source_path.glob("*.pdf"))
        
        # Filter by date range if specified
        pdf_files = []
        for pdf in all_pdf_files:
            file_mtime = datetime.fromtimestamp(pdf.stat().st_mtime)
            
            if from_date and file_mtime < from_date:
                continue
            if to_date and file_mtime > to_date:
                continue
            
            pdf_files.append(pdf)
        
        if from_date or to_date:
            self.finder_log_msg(f"📄 {len(pdf_files)} PDFs within date range (of {len(all_pdf_files)} total)")
        else:
            self.finder_log_msg(f"📄 Scanning {len(pdf_files)} PDFs...")
        
        # Group ALL original criteria by normalized name+auth_type (to track duplicates)
        # This means "John Smith, Skilled" appearing 3 times with different Last DOS will all be tracked
        all_imports_by_key = {}  # (name_normalized, auth_type) -> list of (name, auth_type, last_dos)
        original_criteria_count = len(criteria)
        
        for name, auth_type, last_dos in criteria:
            key = (name.upper().strip(), auth_type.upper().strip())
            if key not in all_imports_by_key:
                all_imports_by_key[key] = []
            all_imports_by_key[key].append((name, auth_type, last_dos))
        
        # Count duplicate import entries
        duplicate_import_entries = []
        for key, entries in all_imports_by_key.items():
            if len(entries) > 1:
                # All but the first are duplicates
                duplicate_import_entries.extend(entries[1:])
        
        # Create deduplicated list for searching (one per unique name+type)
        criteria_for_search = [(entries[0][0], entries[0][1], entries[0][2]) for entries in all_imports_by_key.values()]
        
        if duplicate_import_entries:
            self.finder_log_msg(f"ℹ️ Found {len(duplicate_import_entries)} duplicate name entries in import (each will get its own Last DOS)")
        
        # Search and copy - only copy MOST RECENT matching file per criteria
        matched = 0
        copied_files = []
        
        # Track results for reporting (now includes last_dos)
        found_matches = []  # (name, auth_type, filename, last_dos) - includes ALL duplicates
        encrypted_matches = []  # (name, auth_type, encrypted_filename, fallback_filename, last_dos)
        name_found_type_mismatch = []  # (name, auth_type, found_type, last_dos)
        not_found_at_all = []  # (name, auth_type, last_dos)
        
        for name, auth_type, last_dos in criteria_for_search:
            keywords = self.get_keywords_for_type(auth_type)
            if not keywords:
                self.finder_log_msg(f"⚠️ Unknown auth type: {auth_type} for {name}")
                continue
            
            # Search for matching files - STRICT MATCHING
            name_lower = name.lower()
            
            # Remove common suffixes like (FHC), (FCH) before parsing
            name_cleaned = re.sub(r'\s*\([^)]*\)\s*', ' ', name_lower)
            
            # Extract name parts, filter out very short ones (1-2 chars like middle initials)
            name_parts = [p for p in name_cleaned.replace(",", " ").split() if len(p) > 2]
            
            # Find all matching files and track the most recent one
            matching_files = []  # Full matches (name + type)
            name_only_matches = []  # Name matches but type doesn't
            
            for pdf in pdf_files:
                filename_lower = pdf.name.lower()
                
                # STRICT MATCH: ALL significant name parts must be found in the filename
                # This prevents "Jack Joy" from matching "Annette Jackson"
                if name_parts:
                    matching_parts = sum(1 for part in name_parts if part in filename_lower)
                    # Require ALL parts to match (or all but one if there are 3+ parts)
                    if len(name_parts) <= 2:
                        # For 1-2 name parts, ALL must match
                        name_match = matching_parts == len(name_parts)
                    else:
                        # For 3+ name parts, allow one to be missing (handles middle names)
                        name_match = matching_parts >= len(name_parts) - 1
                else:
                    name_match = False
                
                # Check if keyword matches - must be careful that "skilled" doesn't match "unskilled"
                if auth_type.lower() == "skilled":
                    # For Skilled: must have "skilled" but NOT "unskilled"
                    keyword_match = "skilled" in filename_lower and "unskilled" not in filename_lower
                elif auth_type.lower() == "escort":
                    # For Escort: must have "escort" in filename
                    keyword_match = "escort" in filename_lower
                elif auth_type.lower() == "unskilled":
                    # For Unskilled: must have "unskilled" only
                    keyword_match = "unskilled" in filename_lower
                else:
                    # Unknown type - no match
                    keyword_match = False
                
                if name_match:
                    file_mtime = datetime.fromtimestamp(pdf.stat().st_mtime)
                    if keyword_match:
                        # Full match - name and type both found
                        matching_files.append((pdf, file_mtime))
                    else:
                        # Name found but check what type IS in the filename
                        if auth_type.lower() == "unskilled":
                            # Looking for unskilled, check if skilled (but not unskilled) is there
                            if "skilled" in filename_lower and "unskilled" not in filename_lower:
                                name_only_matches.append((pdf, "Skilled"))
                            elif "escort" in filename_lower:
                                name_only_matches.append((pdf, "Escort"))
                        elif auth_type.lower() == "skilled":
                            # Looking for skilled, check if unskilled/escort is there
                            if "unskilled" in filename_lower or "escort assistance" in filename_lower:
                                name_only_matches.append((pdf, "Unskilled"))
                            elif "escort" in filename_lower:
                                name_only_matches.append((pdf, "Escort"))
                        elif auth_type.lower() == "escort":
                            # Looking for escort, check if skilled/unskilled is there
                            if "skilled" in filename_lower and "unskilled" not in filename_lower:
                                name_only_matches.append((pdf, "Skilled"))
                            elif "unskilled" in filename_lower:
                                name_only_matches.append((pdf, "Unskilled"))
            
            # If we found full matches, copy only the most recent one
            if matching_files:
                # Sort by modification time (newest first) and take the first one
                matching_files.sort(key=lambda x: x[1], reverse=True)
                most_recent_pdf, most_recent_time = matching_files[0]
                
                # Track if file is encrypted (for informational purposes)
                is_encrypted = "encrypted" in most_recent_pdf.name.lower()
                if is_encrypted:
                    # Track encrypted files (informational only - will still process them)
                    encrypted_matches.append((name, auth_type, most_recent_pdf.name, None, last_dos))
                
                # Get ALL import entries with this name+auth_type (includes duplicates with different Last DOS)
                key = (name.upper().strip(), auth_type.upper().strip())
                all_imports_for_this_name = all_imports_by_key.get(key, [(name, auth_type, last_dos)])
                
                # Add entry to found_matches for EACH import (so duplicates get their own Last DOS)
                for import_name, import_type, import_last_dos in all_imports_for_this_name:
                    found_matches.append((import_name, import_type, most_recent_pdf.name, import_last_dos))
                
                # ALWAYS use the most recent file (password 92020 handles decryption during extraction)
                dest_file = dest_path / most_recent_pdf.name
                if dest_file not in copied_files:
                    try:
                        shutil.copy2(most_recent_pdf, dest_file)
                        # Auto-unlock encrypted PDFs in place
                        self._auto_unlock_pdf(str(dest_file))
                        copied_files.append(dest_file)
                        matched += 1
                    except Exception as e:
                        self.finder_log_msg(f"❌ Error copying {most_recent_pdf.name}: {e}")
            elif name_only_matches:
                # Name was found but type didn't match
                found_type = name_only_matches[0][1]  # Get the type that was found
                # Add entry for ALL imports with this name
                key = (name.upper().strip(), auth_type.upper().strip())
                all_imports_for_this_name = all_imports_by_key.get(key, [(name, auth_type, last_dos)])
                for import_name, import_type, import_last_dos in all_imports_for_this_name:
                    name_found_type_mismatch.append((import_name, import_type, found_type, import_last_dos))
            else:
                # Name not found at all
                # Add entry for ALL imports with this name
                key = (name.upper().strip(), auth_type.upper().strip())
                all_imports_for_this_name = all_imports_by_key.get(key, [(name, auth_type, last_dos)])
                for import_name, import_type, import_last_dos in all_imports_for_this_name:
                    not_found_at_all.append((import_name, import_type, import_last_dos))
        
        # Group found matches by filename to detect duplicates
        file_to_criteria = {}  # filename -> list of (name, auth_type, last_dos)
        for name, auth_type, filename, last_dos in found_matches:
            if filename not in file_to_criteria:
                file_to_criteria[filename] = []
            file_to_criteria[filename].append((name, auth_type, last_dos))
        
        # Separate into unique matches and duplicate matches (multiple criteria -> same file)
        unique_matches = []  # Files matched by exactly one criteria
        duplicate_matches = []  # Files matched by multiple criteria
        
        for filename, criteria_list in file_to_criteria.items():
            if len(criteria_list) == 1:
                unique_matches.append((criteria_list[0][0], criteria_list[0][1], filename, criteria_list[0][2]))
            else:
                duplicate_matches.append((filename, criteria_list))
        
        # Output results in sections
        self.finder_log_msg(f"\n{'='*60}")
        self.finder_log_msg(f"✅ FOUND & COPIED - UNIQUE MATCHES ({len(unique_matches)} files)")
        self.finder_log_msg(f"{'='*60}")
        for name, auth_type, filename, last_dos in unique_matches:
            self.finder_log_msg(f"  ✓ {name} ({auth_type}) → {filename}")
        
        if duplicate_matches:
            self.finder_log_msg(f"\n{'='*60}")
            self.finder_log_msg(f"🔄 DUPLICATE MATCHES - Multiple names matched same file ({len(duplicate_matches)} files)")
            self.finder_log_msg(f"{'='*60}")
            for filename, criteria_list in duplicate_matches:
                self.finder_log_msg(f"  🔄 FILE: {filename}")
                self.finder_log_msg(f"     Matched to {len(criteria_list)} patients:")
                for name, auth_type, last_dos in criteria_list:
                    self.finder_log_msg(f"       - {name} ({auth_type})")
                self.finder_log_msg("")  # Blank line between entries
        
        if encrypted_matches:
            self.finder_log_msg(f"\n{'='*60}")
            self.finder_log_msg(f"🔒 ENCRYPTED FILES ({len(encrypted_matches)})")
            self.finder_log_msg(f"{'='*60}")
            for name, auth_type, encrypted_file, fallback_file, last_dos in encrypted_matches:
                self.finder_log_msg(f"  🔒 {name} ({auth_type})")
                self.finder_log_msg(f"     File: {encrypted_file} (password will be used for extraction)")
        
        if name_found_type_mismatch:
            self.finder_log_msg(f"\n{'='*60}")
            self.finder_log_msg(f"⚠️ NAME FOUND BUT TYPE MISMATCH ({len(name_found_type_mismatch)})")
            self.finder_log_msg(f"{'='*60}")
            for name, wanted_type, found_type, last_dos in name_found_type_mismatch:
                self.finder_log_msg(f"  ⚠️ {name} - wanted {wanted_type}, found {found_type}")
        
        if not_found_at_all:
            self.finder_log_msg(f"\n{'='*60}")
            self.finder_log_msg(f"❌ NOT FOUND AT ALL ({len(not_found_at_all)})")
            self.finder_log_msg(f"{'='*60}")
            for name, auth_type, last_dos in not_found_at_all:
                self.finder_log_msg(f"  ❌ {name} ({auth_type})")
        
        # Calculate how many criteria entries matched vs didn't
        total_criteria_matched = len(found_matches)  # Total import entries that found a file (includes duplicates)
        total_criteria_not_matched = len(name_found_type_mismatch) + len(not_found_at_all)
        
        self.finder_log_msg(f"\n{'='*60}")
        self.finder_log_msg(f"📊 SUMMARY")
        self.finder_log_msg(f"{'='*60}")
        self.finder_log_msg(f"  Total names in import: {original_criteria_count}")
        self.finder_log_msg(f"  Unique names searched: {len(criteria_for_search)}")
        self.finder_log_msg(f"  Duplicate names in import: {len(duplicate_import_entries)}")
        self.finder_log_msg(f"  ---")
        self.finder_log_msg(f"  Import entries matched: {total_criteria_matched} (including duplicates)")
        self.finder_log_msg(f"  - Unique files found: {len(unique_matches) + len(duplicate_matches)}")
        self.finder_log_msg(f"  - Entries per file (avg): {total_criteria_matched / max(len(unique_matches) + len(duplicate_matches), 1):.1f}")
        self.finder_log_msg(f"  Encrypted files (processed with password): {len(encrypted_matches)}")
        self.finder_log_msg(f"  Import entries - type mismatch: {len(name_found_type_mismatch)}")
        self.finder_log_msg(f"  Import entries - not found: {len(not_found_at_all)}")
        self.finder_log_msg(f"  ---")
        self.finder_log_msg(f"  📁 UNIQUE FILES COPIED: {matched}")
        
        # Store results for Excel export (now includes last_dos)
        self.finder_found_matches = found_matches  # (name, auth_type, filename, last_dos) - includes ALL duplicates
        self.finder_duplicate_imports = duplicate_import_entries  # Track duplicates from import
        self.finder_original_count = original_criteria_count  # Track original count
        self.finder_not_found = []
        
        # Update Extract tab file count
        if hasattr(self, 'finder_files_count_var'):
            self.update_finder_files_count()
        
        # Add type mismatches to not found (with reason and last_dos)
        for name, wanted_type, found_type, last_dos in name_found_type_mismatch:
            self.finder_not_found.append((name, wanted_type, f"Type mismatch - wanted {wanted_type}, found {found_type}", last_dos))
        # Note: Encrypted files are now processed normally (password handles decryption)
        # Add not found at all (with reason and last_dos)
        for name, auth_type, last_dos in not_found_at_all:
            self.finder_not_found.append((name, auth_type, "Not found in source folder", last_dos))
        
        # Populate the results view (this auto-switches to results view)
        self.populate_finder_results()
        
        self.finder_log_msg(f"\n💡 TIP: Click 'View Results' button to see detailed results")
        
        if matched > 0:
            messagebox.showinfo("File Finder Complete", 
                f"Copied {matched} unique PDF files to:\n{dest}\n\n" +
                f"Names in import: {original_criteria_count}\n" +
                f"Unique names: {len(criteria_for_search)}\n" +
                f"Duplicate names: {len(duplicate_import_entries)}\n\n" +
                f"Import entries matched: {total_criteria_matched}\n" +
                f"Unique files returned: {matched}\n" +
                f"Type mismatch: {len(name_found_type_mismatch)}\n" +
                f"Not found: {len(not_found_at_all)}\n\n" +
                f"Results are displayed below.")
        else:
            messagebox.showinfo("File Finder Complete", 
                "No matching files found.\nResults view shows names that weren't found.")
    
    def start_combined_extraction(self, folder_path):
        """Start extraction process after file finder, combining results into one workbook."""
        # Ask for output file location
        output_file = filedialog.asksaveasfilename(
            title="Save Combined Results As",
            initialdir=folder_path,
            initialfile="Auth_Results_Combined.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not output_file:
            return
        
        # Switch to extractor tab to show progress
        self.notebook.select(self.extractor_tab)
        
        # Set the input/output paths
        self.input_folder.set(folder_path)
        self.output_file.set(output_file)
        
        # Update UI
        self.log(f"\n{'='*60}")
        self.log(f"🔄 COMBINED EXTRACTION - Starting...")
        self.log(f"{'='*60}")
        self.log(f"📁 Source: {folder_path}")
        self.log(f"📊 Output: {output_file}")
        
        # Disable buttons during processing
        self.is_processing = True
        self.extract_btn.config(state=tk.DISABLED)
        
        # Run extraction in background thread
        thread = Thread(target=self.run_combined_extraction, args=(folder_path, output_file))
        thread.start()
    
    def run_combined_extraction(self, input_folder, output_file):
        """Run extraction and combine with file finder results."""
        try:
            self.log(f"📁 Scanning folder: {input_folder}")
            
            # Count all supported files
            from config import SUPPORTED_EXTENSIONS
            folder = pathlib.Path(input_folder)
            total_files = sum(len(list(folder.rglob(f"*{ext}"))) for ext in SUPPORTED_EXTENSIONS)
            pdf_count = len(list(folder.rglob("*.pdf")))
            self.log(f"📄 Found {total_files} supported files ({pdf_count} PDFs)")
            
            if total_files == 0:
                self.root.after(0, lambda: messagebox.showwarning("Warning", "No supported files found in folder"))
                return
            
            # Process all supported file types
            self.log("\n🔍 Extracting data (routing each file to the best method)...")
            results = self.extractor.process_all_files(
                input_folder, 
                progress_callback=lambda c, t, f: self.root.after(0, lambda: self.update_progress(c, t, f))
            )
            
            # Export combined workbook
            self.log(f"\n📊 Creating combined Excel workbook...")
            self.export_combined_workbook(output_file)
            
            # Summary
            successful = len([r for r in results if "error" not in r])
            errors = len([r for r in results if "error" in r])
            ocr_count = len([r for r in results if r.get("extraction_method") in ("ocr", "tesseract_image")])
            
            self.log(f"\n✅ Combined Export Complete!")
            self.log(f"   PDFs Processed: {successful}")
            self.log(f"   Used OCR: {ocr_count}")
            self.log(f"   Extraction Errors: {errors}")
            self.log(f"   Found Auths: {len(self.finder_found_matches)}")
            self.log(f"   Not Found Auths: {len(self.finder_not_found)}")
            self.log(f"   Output: {output_file}")
            
            self.root.after(0, lambda: self.status_text.set(f"Done! Combined export complete"))
            self.root.after(0, lambda: self.open_btn.config(state=tk.NORMAL))
            self.root.after(0, lambda: messagebox.showinfo("Success", 
                f"Combined extraction complete!\n\n" +
                f"PDFs Processed: {pdf_count}\n" +
                f"Successful: {successful}\n" +
                f"Errors: {errors}\n\n" +
                f"Excel workbook contains:\n" +
                f"• Raw Data & Formatted sheets\n" +
                f"• Found Auths ({len(self.finder_found_matches)})\n" +
                f"• Not Found Auths ({len(self.finder_not_found)})\n" +
                f"• Import Summary with dates\n\n" +
                f"Output: {output_file}"))
            
        except Exception as e:
            err_msg = str(e)
            self.log(f"\n❌ Error: {err_msg}")
            self.root.after(0, lambda m=err_msg: messagebox.showerror("Error", m))
        finally:
            self.is_processing = False
            self.root.after(0, lambda: self.extract_btn.config(state=tk.NORMAL))
    
    def export_combined_workbook(self, output_path):
        """Export combined workbook with extraction results AND file finder results."""
        if not pd:
            raise ImportError("pandas is required for Excel export")
        
        from openpyxl.styles import PatternFill
        
        # Build lookup from filename to ALL import entries (including duplicates)
        # This maps filename -> list of (name, auth_type, last_dos) for ALL matching import entries
        filename_to_imports = {}  # filename -> list of (name, auth_type, last_dos)
        for name, auth_type, filename, last_dos in self.finder_found_matches:
            if filename not in filename_to_imports:
                filename_to_imports[filename] = []
            filename_to_imports[filename].append((name, auth_type, last_dos))
            # Also add lowercase version
            if filename.lower() not in filename_to_imports:
                filename_to_imports[filename.lower()] = []
            if (name, auth_type, last_dos) not in filename_to_imports[filename.lower()]:
                filename_to_imports[filename.lower()].append((name, auth_type, last_dos))
        
        # Build name -> list of (auth_type, last_dos) for ALL import entries from the table
        # This captures duplicates that were in the original import
        name_to_all_imports = {}  # normalized_name -> list of (auth_type, last_dos)
        if hasattr(self, 'finder_table'):
            for item in self.finder_table.get_children():
                values = self.finder_table.item(item, "values")
                if len(values) >= 4:
                    name, auth_type, searches_for, last_dos = values[0], values[1], values[2], values[3]
                    if name:
                        normalized = name.upper().strip()
                        normalized = re.sub(r'\s+', ' ', normalized)
                        if normalized not in name_to_all_imports:
                            name_to_all_imports[normalized] = []
                        name_to_all_imports[normalized].append((auth_type, last_dos))
        
        # === EXTRACTION DATA ===
        # Raw Data
        df_raw = pd.DataFrame(self.extractor.results) if self.extractor.results else pd.DataFrame()
        if not df_raw.empty:
            for col in ["Date Approved", "Date Auth Expire"]:
                if col in df_raw.columns:
                    df_raw[col] = pd.to_datetime(df_raw[col], errors="coerce")
            ordered = [
                "file", "Patient Name", "Auth #", "Date Approved", "Date Auth Expire", "Patient ID",
                "extraction_method", "raw_text_preview", "extracted_at", "error"
            ]
            cols = [c for c in ordered if c in df_raw.columns] + [c for c in df_raw.columns if c not in ordered]
            df_raw = df_raw[cols]
        
        # Formatted Data - Generate ONE ROW PER IMPORT ENTRY (including duplicates)
        formatted_data = self.extractor.format_results() if self.extractor.results else []
        
        # Expand formatted data to include one row per import entry
        expanded_formatted_data = []
        
        for idx in range(len(formatted_data)):
            row_data = formatted_data[idx].copy()
            
            # Find the corresponding file from extractor results
            if idx < len(self.extractor.results):
                result = self.extractor.results[idx]
                filename = pathlib.Path(result.get("file", "")).name
                extracted_name = result.get("Patient Name", "")
                
                # Get ALL import entries that match this file
                import_entries = filename_to_imports.get(filename, [])
                if not import_entries:
                    import_entries = filename_to_imports.get(filename.lower(), [])
                
                # If no filename match, try to match by patient name
                if not import_entries and extracted_name and name_to_all_imports:
                    normalized = extracted_name.upper().strip()
                    normalized = re.sub(r'\s+', ' ', normalized)
                    
                    # Try exact match
                    if normalized in name_to_all_imports:
                        import_entries = [(auth_type, last_dos) for auth_type, last_dos in name_to_all_imports[normalized]]
                        import_entries = [(normalized, auth_type, last_dos) for auth_type, last_dos in name_to_all_imports[normalized]]
                    
                    # Try LAST, FIRST format
                    if not import_entries and "," in normalized:
                        parts = normalized.split(",", 1)
                        last_first = f"{parts[0].strip()}, {parts[1].strip()}"
                        if last_first in name_to_all_imports:
                            import_entries = [(last_first, auth_type, last_dos) for auth_type, last_dos in name_to_all_imports[last_first]]
                    
                    # Try reverse: FIRST LAST -> LAST, FIRST
                    if not import_entries and "," not in normalized:
                        parts = normalized.split()
                        if len(parts) >= 2:
                            last_first = f"{parts[-1]}, {parts[0]}"
                            if last_first in name_to_all_imports:
                                import_entries = [(last_first, auth_type, last_dos) for auth_type, last_dos in name_to_all_imports[last_first]]
                    
                    # Try partial match by last name
                    if not import_entries:
                        if "," in normalized:
                            last_name = normalized.split(",")[0].strip()
                        else:
                            last_name = normalized.split()[-1] if normalized.split() else ""
                        
                        if last_name:
                            for db_name, imports_list in name_to_all_imports.items():
                                if last_name in db_name:
                                    import_entries = [(db_name, auth_type, last_dos) for auth_type, last_dos in imports_list]
                                    break
                
                # Create one row for EACH import entry (handles duplicates with different Last DOS)
                if import_entries:
                    for import_entry in import_entries:
                        if len(import_entry) == 3:
                            _, auth_type, last_dos = import_entry
                        else:
                            auth_type, last_dos = import_entry[0], import_entry[1] if len(import_entry) > 1 else ""
                        
                        new_row = row_data.copy()
                        new_row["Last DOS"] = last_dos
                        expanded_formatted_data.append(new_row)
                else:
                    # No import match - add row without Last DOS
                    row_data["Last DOS"] = ""
                    expanded_formatted_data.append(row_data)
            else:
                row_data["Last DOS"] = ""
                expanded_formatted_data.append(row_data)
        
        df_formatted = pd.DataFrame(expanded_formatted_data) if expanded_formatted_data else pd.DataFrame()
        
        # Track which rows should be highlighted (Last DOS > Date Auth Expired)
        rows_to_highlight = []
        
        if not df_formatted.empty:
            # Convert date columns for display
            for col in ["Date Approved", "Date Auth Expired"]:
                if col in df_formatted.columns:
                    df_formatted[col] = pd.to_datetime(df_formatted[col], errors="coerce")
            
            # Parse Last DOS with explicit format (MM/DD/YYYY from bulk import)
            if "Last DOS" in df_formatted.columns:
                df_formatted["Last DOS"] = pd.to_datetime(df_formatted["Last DOS"], format="%m/%d/%Y", errors="coerce")
            
            # Now determine which rows to highlight by comparing the ACTUAL DataFrame values
            for idx, row in df_formatted.iterrows():
                last_dos_val = row.get("Last DOS")
                expire_val = row.get("Date Auth Expired")
                
                # Only highlight if BOTH dates are valid AND Last DOS > Expire Date
                if pd.notna(last_dos_val) and pd.notna(expire_val):
                    if last_dos_val > expire_val:
                        rows_to_highlight.append(idx)
            
            # Column order with Last DOS added after Date Auth Expired
            formatted_columns = [
                "Last Name", "First Name", "Patient Name", "Extracted Name", "Patient ID",
                "CPT Code", "CPT Code 2", "CPT Code 3", "CPT Code 4", "CPT Code 5",
                "Auth Number", "Date Approved", "Date Auth Expired", "Last DOS",
                "Clearing House Payer ID", "Location ID", "Unique Payer Identifier"
            ]
            df_formatted = df_formatted[[c for c in formatted_columns if c in df_formatted.columns]]
        
        # Errors
        error_records = [r for r in self.extractor.results if r.get("error")] if self.extractor.results else []
        df_errors = pd.DataFrame(error_records) if error_records else pd.DataFrame(columns=["file", "error"])
        if not df_errors.empty:
            error_cols = [c for c in ["file", "error", "extraction_method"] if c in df_errors.columns]
            df_errors = df_errors[error_cols]
        
        # === FILE FINDER DATA ===
        # Found Auths
        found_data = []
        for name, auth_type, filename, last_dos in self.finder_found_matches:
            found_data.append({
                "Patient Name": name, "Auth Type": auth_type, "File Found": filename, "Last DOS": last_dos
            })
        df_found = pd.DataFrame(found_data) if found_data else pd.DataFrame(columns=["Patient Name", "Auth Type", "File Found", "Last DOS"])
        
        # Not Found Auths
        not_found_data = []
        for name, auth_type, reason, last_dos in self.finder_not_found:
            not_found_data.append({
                "Patient Name": name, "Auth Type": auth_type, "Reason": reason, "Last DOS": last_dos
            })
        df_not_found = pd.DataFrame(not_found_data) if not_found_data else pd.DataFrame(columns=["Patient Name", "Auth Type", "Reason", "Last DOS"])
        
        # === IMPORT SUMMARY (combine file finder + extraction data) ===
        # Build a lookup of extracted data by filename
        extraction_lookup = {}
        if self.extractor.results:
            for result in self.extractor.results:
                filename = pathlib.Path(result.get("file", "")).name
                extraction_lookup[filename] = {
                    "Date Approved": result.get("Date Approved", ""),
                    "Date Expired": result.get("Date Auth Expire", ""),
                    "Auth Number": result.get("Auth #", ""),
                    "Patient ID": result.get("Patient ID", "")
                }
        
        summary_data = []
        for name, auth_type, filename, last_dos in self.finder_found_matches:
            extracted = extraction_lookup.get(filename, {})
            summary_data.append({
                "Patient Name": name,
                "Auth Type": auth_type,
                "Date Approved": extracted.get("Date Approved", ""),
                "Date Expired": extracted.get("Date Expired", ""),
                "Last DOS": last_dos,
                "Auth Number": extracted.get("Auth Number", ""),
                "Patient ID": extracted.get("Patient ID", ""),
                "File": filename
            })
        df_summary = pd.DataFrame(summary_data) if summary_data else pd.DataFrame(
            columns=["Patient Name", "Auth Type", "Date Approved", "Date Expired", "Last DOS", "Auth Number", "Patient ID", "File"])
        
        # Parse dates in summary
        for col in ["Date Approved", "Date Expired"]:
            if col in df_summary.columns:
                df_summary[col] = pd.to_datetime(df_summary[col], errors="coerce")
        # Parse Last DOS with explicit MM/DD/YYYY format
        if "Last DOS" in df_summary.columns:
            df_summary["Last DOS"] = pd.to_datetime(df_summary["Last DOS"], format="%m/%d/%Y", errors="coerce")
        
        # Helper function for column width adjustment
        def adjust_column_widths(worksheet):
            for col_cells in worksheet.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for cell in col_cells:
                    try:
                        v = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, len(v))
                    except:
                        pass
                worksheet.column_dimensions[col_letter].width = min(max_len + 2, 50)
        
        # Red fill for expired auth rows
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        
        # Write to Excel
        with pd.ExcelWriter(output_path, engine="openpyxl", 
                           datetime_format="mm/dd/yyyy", date_format="mm/dd/yyyy") as writer:
            # Extraction sheets
            if not df_raw.empty:
                df_raw.to_excel(writer, sheet_name="Raw Data", index=False)
                writer.sheets["Raw Data"].freeze_panes = "A2"
                adjust_column_widths(writer.sheets["Raw Data"])
            
            if not df_formatted.empty:
                df_formatted.to_excel(writer, sheet_name="Formatted", index=False)
                ws_formatted = writer.sheets["Formatted"]
                ws_formatted.freeze_panes = "A2"
                adjust_column_widths(ws_formatted)
                
                # Apply red highlighting to rows where Last DOS > Date Auth Expired
                # Row indices in Excel are 1-based, and row 1 is the header
                for row_idx in rows_to_highlight:
                    excel_row = row_idx + 2  # +1 for 0-based to 1-based, +1 for header row
                    for col in range(1, ws_formatted.max_column + 1):
                        ws_formatted.cell(row=excel_row, column=col).fill = red_fill
            
            df_errors.to_excel(writer, sheet_name="Errors", index=False)
            writer.sheets["Errors"].freeze_panes = "A2"
            adjust_column_widths(writer.sheets["Errors"])
            
            # File Finder sheets
            df_found.to_excel(writer, sheet_name="Found Auths", index=False)
            writer.sheets["Found Auths"].freeze_panes = "A2"
            adjust_column_widths(writer.sheets["Found Auths"])
            
            df_not_found.to_excel(writer, sheet_name="Not Found Auths", index=False)
            writer.sheets["Not Found Auths"].freeze_panes = "A2"
            adjust_column_widths(writer.sheets["Not Found Auths"])
            
            # Combined Summary sheet
            df_summary.to_excel(writer, sheet_name="Import Summary", index=False)
            ws_summary = writer.sheets["Import Summary"]
            ws_summary.freeze_panes = "A2"
            adjust_column_widths(ws_summary)
            
            # Apply red highlighting to Import Summary rows where Last DOS > Date Expired
            for row_idx, row_data in df_summary.iterrows():
                last_dos_val = row_data.get("Last DOS")
                expire_val = row_data.get("Date Expired")
                if pd.notna(last_dos_val) and pd.notna(expire_val):
                    if last_dos_val > expire_val:
                        excel_row = row_idx + 2  # +1 for 0-based, +1 for header
                        for col in range(1, ws_summary.max_column + 1):
                            ws_summary.cell(row=excel_row, column=col).fill = red_fill
            
            # === SUMMARY STATISTICS SHEET ===
            # Use stored values from file finder if available, otherwise calculate
            total_names_imported = self.finder_original_count if self.finder_original_count > 0 else (len(self.finder_found_matches) + len(self.finder_not_found))
            duplicate_names_in_import = len(self.finder_duplicate_imports) if hasattr(self, 'finder_duplicate_imports') else 0
            
            # Unique names = total imports - duplicates
            unique_names_searched = total_names_imported - duplicate_names_in_import
            
            # Unique files returned (each file only counted once)
            unique_files = len(set(filename for name, auth_type, filename, last_dos in self.finder_found_matches))
            
            # Count expired auths (Last DOS > Date Expired) - from formatted sheet
            expired_count = len(rows_to_highlight) if rows_to_highlight else 0
            
            # Rows in formatted sheet (includes duplicates with their Last DOS)
            formatted_rows_count = len(df_formatted) if not df_formatted.empty else 0
            
            # Build Summary Statistics data
            stats_data = [
                {"Metric": "IMPORT STATISTICS", "Value": ""},
                {"Metric": "Total Entries in Import List", "Value": total_names_imported},
                {"Metric": "Duplicate Names in Import", "Value": duplicate_names_in_import},
                {"Metric": "Unique Names Searched", "Value": unique_names_searched},
                {"Metric": "", "Value": ""},
                {"Metric": "SEARCH RESULTS", "Value": ""},
                {"Metric": "Import Entries Matched", "Value": len(self.finder_found_matches)},
                {"Metric": "Unique Files Returned", "Value": unique_files},
                {"Metric": "Import Entries Not Found", "Value": len(self.finder_not_found)},
                {"Metric": "", "Value": ""},
                {"Metric": "OUTPUT", "Value": ""},
                {"Metric": "Rows in Formatted Sheet", "Value": formatted_rows_count},
                {"Metric": "Expired Auths (Last DOS > Expire)", "Value": expired_count},
                {"Metric": "Extraction Errors", "Value": len(error_records)},
            ]
            df_stats = pd.DataFrame(stats_data)
            
            df_stats.to_excel(writer, sheet_name="Summary Statistics", index=False)
            ws_stats = writer.sheets["Summary Statistics"]
            ws_stats.column_dimensions["A"].width = 40
            ws_stats.column_dimensions["B"].width = 15
            
            # Bold the section headers
            from openpyxl.styles import Font
            bold_font = Font(bold=True)
            for row in [2, 7, 12]:  # IMPORT STATISTICS, SEARCH RESULTS, OUTPUT rows
                ws_stats.cell(row=row, column=1).font = bold_font
        
        # Return the formatted DataFrame for in-app editing
        return df_formatted
    
    def export_finder_results(self):
        """Export File Finder results to Excel with Found, Not Found, and Import Summary sheets."""
        if not self.finder_found_matches and not self.finder_not_found:
            messagebox.showwarning("No Results", 
                "No file finder results to export.\n\nRun 'Find & Copy Files' first to generate results.")
            return
        
        # Ask user where to save
        dest_folder = self.finder_dest.get()
        default_filename = "File_Finder_Results.xlsx"
        if dest_folder:
            initial_dir = dest_folder
        else:
            initial_dir = None
        
        output_file = filedialog.asksaveasfilename(
            title="Save File Finder Results",
            initialdir=initial_dir,
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        
        if not output_file:
            return
        
        try:
            # Create DataFrames for each sheet
            # Found Auths sheet (includes last_dos now)
            found_data = []
            for name, auth_type, filename, last_dos in self.finder_found_matches:
                found_data.append({
                    "Patient Name": name,
                    "Auth Type": auth_type,
                    "File Found": filename,
                    "Last DOS": last_dos
                })
            df_found = pd.DataFrame(found_data) if found_data else pd.DataFrame(columns=["Patient Name", "Auth Type", "File Found", "Last DOS"])
            
            # Not Found Auths sheet (includes last_dos now)
            not_found_data = []
            for name, auth_type, reason, last_dos in self.finder_not_found:
                not_found_data.append({
                    "Patient Name": name,
                    "Auth Type": auth_type,
                    "Reason": reason,
                    "Last DOS": last_dos
                })
            df_not_found = pd.DataFrame(not_found_data) if not_found_data else pd.DataFrame(columns=["Patient Name", "Auth Type", "Reason", "Last DOS"])
            
            # Import Summary sheet - Patient Name, Date Approved, Date Expired, Last DOS
            # Date Approved and Date Expired are left blank to be filled in after extraction
            summary_data = []
            for name, auth_type, filename, last_dos in self.finder_found_matches:
                summary_data.append({
                    "Patient Name": name,
                    "Auth Type": auth_type,
                    "Date Approved": "",  # To be filled after extraction
                    "Date Expired": "",   # To be filled after extraction
                    "Last DOS": last_dos,
                    "File": filename
                })
            df_summary = pd.DataFrame(summary_data) if summary_data else pd.DataFrame(columns=["Patient Name", "Auth Type", "Date Approved", "Date Expired", "Last DOS", "File"])
            
            # Write to Excel
            with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
                # Found Auths sheet
                df_found.to_excel(writer, sheet_name="Found Auths", index=False)
                ws_found = writer.sheets["Found Auths"]
                ws_found.freeze_panes = "A2"
                
                # Auto-adjust column widths
                for col_cells in ws_found.columns:
                    max_len = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        try:
                            v = "" if cell.value is None else str(cell.value)
                            max_len = max(max_len, len(v))
                        except:
                            pass
                    ws_found.column_dimensions[col_letter].width = min(max_len + 2, 60)
                
                # Not Found Auths sheet
                df_not_found.to_excel(writer, sheet_name="Not Found Auths", index=False)
                ws_not_found = writer.sheets["Not Found Auths"]
                ws_not_found.freeze_panes = "A2"
                
                # Auto-adjust column widths
                for col_cells in ws_not_found.columns:
                    max_len = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        try:
                            v = "" if cell.value is None else str(cell.value)
                            max_len = max(max_len, len(v))
                        except:
                            pass
                    ws_not_found.column_dimensions[col_letter].width = min(max_len + 2, 60)
                
                # Import Summary sheet
                df_summary.to_excel(writer, sheet_name="Import Summary", index=False)
                ws_summary = writer.sheets["Import Summary"]
                ws_summary.freeze_panes = "A2"
                
                # Auto-adjust column widths
                for col_cells in ws_summary.columns:
                    max_len = 0
                    col_letter = col_cells[0].column_letter
                    for cell in col_cells:
                        try:
                            v = "" if cell.value is None else str(cell.value)
                            max_len = max(max_len, len(v))
                        except:
                            pass
                    ws_summary.column_dimensions[col_letter].width = min(max_len + 2, 60)
            
            self.finder_log_msg(f"\n📊 Results exported to: {output_file}")
            self.finder_log_msg(f"   - Found Auths: {len(found_data)} entries")
            self.finder_log_msg(f"   - Not Found Auths: {len(not_found_data)} entries")
            self.finder_log_msg(f"   - Import Summary: {len(summary_data)} entries")
            
            messagebox.showinfo("Export Complete", 
                f"File Finder results exported to:\n{output_file}\n\n" +
                f"Found Auths: {len(found_data)} entries\n" +
                f"Not Found Auths: {len(not_found_data)} entries\n" +
                f"Import Summary: {len(summary_data)} entries")
            
        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting results:\n{str(e)}")
            self.finder_log_msg(f"❌ Export error: {str(e)}")
    
    def update_source_selection(self):
        """Update UI based on source selection for extraction."""
        source_type = self.extract_source_type.get()
        self.manual_folder_frame.pack_forget()
        if source_type == "finder":
            self.update_finder_files_count()
        else:
            # manual
            self.manual_folder_frame.pack(fill=tk.X, pady=(5, 0))
    
    def update_finder_files_count(self):
        """Update the file count label for finder results."""
        dest = self.finder_dest.get()
        if dest and pathlib.Path(dest).exists():
            pdf_count = len(list(pathlib.Path(dest).glob("*.pdf")))
            if pdf_count > 0:
                self.finder_files_count_var.set(f"({pdf_count} PDF files ready in: {pathlib.Path(dest).name})")
                self.finder_files_label.configure(foreground="green")
            else:
                self.finder_files_count_var.set("(No PDFs in destination folder)")
                self.finder_files_label.configure(foreground="gray")
        else:
            self.finder_files_count_var.set("(No files found yet - run File Finder first)")
            self.finder_files_label.configure(foreground="gray")

    def browse_input(self):
        """Browse for input folder."""
        folder = filedialog.askdirectory(title="Select the FOLDER containing your PDFs (e.g., Archive Auth's PDF)")
        if folder:
            self.input_folder.set(folder)
            # Count PDFs found
            pdf_count = len(list(pathlib.Path(folder).glob("*.pdf")))
            if pdf_count > 0:
                self.log(f"📁 Selected folder: {folder}")
                self.log(f"📄 Found {pdf_count} PDF files ready to process")
            else:
                self.log(f"⚠️ No PDF files found in: {folder}")
            # Auto-set output file
            if not self.output_file.get():
                default_output = pathlib.Path(folder) / "Auth_Results_Combined.xlsx"
                self.output_file.set(str(default_output))
    
    def browse_output(self):
        """Browse for output file."""
        file = filedialog.asksaveasfilename(
            title="Save Excel file as",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file:
            self.output_file.set(file)
    
    def log(self, message):
        """Add message to log."""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.root.update_idletasks()
    
    def update_progress(self, current, total, filename):
        """Update progress bar and status."""
        percent = (current / total) * 100
        self.progress_bar["value"] = percent
        self.status_text.set(f"Processing {current}/{total}: {filename}")
        self.log(f"✓ Processed: {filename}")
        self.root.update_idletasks()
    
    def test_single_pdf(self):
        """Test extraction on a single PDF."""
        file = filedialog.askopenfilename(
            title="Select a PDF to test",
            filetypes=[("PDF files", "*.pdf")]
        )
        if file:
            self.log(f"\n🔍 Testing: {pathlib.Path(file).name}")
            try:
                result = self.extractor.process_pdf(file)
                self.log(f"   Method: {result.get('extraction_method', 'unknown')}")
                self.log(f"   Auth Form Page: {result.get('auth_page', 'unknown')}")
                self.log(f"   --- Extracted Fields ---")
                for field in FIELDS:
                    value = result.get(field, "NOT FOUND")
                    status = "✓" if value and value != "NOT FOUND" else "✗"
                    self.log(f"   {status} {field}: {value}")
                if result.get("raw_text_preview"):
                    self.log(f"\n   --- Raw Text (first 600 chars) ---")
                    # Show with line breaks preserved for readability
                    preview = result['raw_text_preview'][:600]
                    self.log(f"   {preview}")
            except Exception as e:
                self.log(f"   ❌ Error: {e}")
                import traceback
                self.log(f"   {traceback.format_exc()}")
    
    def sync_patient_names_from_caspio(self):
        """Sync patient names from Caspio a_Patient table to local JSON file."""
        if not REQUESTS_AVAILABLE:
            messagebox.showerror("Error", "requests library is required. Run: pip install requests")
            return
        
        self.log("\n🔄 Syncing patient names from Caspio...")
        self.sync_btn.config(state=tk.DISABLED)
        self.status_text.set("Syncing patient names from Caspio...")
        
        # Run in thread to keep UI responsive
        thread = Thread(target=self._do_caspio_sync)
        thread.start()
    
    def _do_caspio_sync(self):
        """Perform the Caspio sync in a background thread."""
        try:
            caspio = CaspioAPI()
            count = caspio.sync_patient_names_to_file()
            
            # Update UI on main thread
            self.root.after(0, lambda: self._sync_complete(count, None))
        except Exception as e:
            self.root.after(0, lambda: self._sync_complete(0, str(e)))
    
    def _sync_complete(self, count, error):
        """Handle sync completion on main thread."""
        self.sync_btn.config(state=tk.NORMAL)
        
        if error:
            self.log(f"   ❌ Error syncing from Caspio: {error}")
            self.status_text.set("Sync failed")
            messagebox.showerror("Sync Error", f"Failed to sync patient names:\n{error}")
        else:
            self.log(f"   ✅ Synced {count} patient names from Caspio")
            self.log(f"   📁 Saved to: {PATIENT_NAMES_FILE}")
            self.status_text.set(f"Synced {count} patient names")
            messagebox.showinfo("Sync Complete", 
                f"Successfully synced {count} patient names from Caspio.\n\n"
                f"Patient names will now be matched against this list when extracting data.")

    # ------------------------------------------------------------------
    # Dropbox integration methods
    # ------------------------------------------------------------------

    def _connect_dropbox(self):
        """Connect to Dropbox using credentials from .env."""
        try:
            from integrations.dropbox_service import DropboxService, DROPBOX_AVAILABLE
            if not DROPBOX_AVAILABLE:
                messagebox.showerror("Error",
                    "The 'dropbox' package is not installed.\nRun: pip install dropbox")
                return

            from config import DROPBOX_APP_KEY, DROPBOX_REFRESH_TOKEN
            if not DROPBOX_APP_KEY or not DROPBOX_REFRESH_TOKEN:
                messagebox.showerror("Dropbox Setup Required",
                    "Dropbox credentials not configured.\n\n"
                    "1. Create a Dropbox app at:\n"
                    "   https://www.dropbox.com/developers/apps\n\n"
                    "2. Run this in a terminal to get a refresh token:\n"
                    "   python -c \"from integrations.dropbox_service import "
                    "run_dropbox_oauth_flow; run_dropbox_oauth_flow()\"\n\n"
                    "3. Add DROPBOX_APP_KEY, DROPBOX_APP_SECRET,\n"
                    "   and DROPBOX_REFRESH_TOKEN to your .env file.")
                return

            self.log("🔗 Connecting to Dropbox...")
            self.dropbox_service = DropboxService()
            self.dropbox_service.connect()

            self.dropbox_status_var.set("✅ connected")
            self.dropbox_connect_btn.configure(text="✅ Connected")
            self.log("   ✅ Dropbox connected!")

            # Auto-fill the folder if DROPBOX_ROOT_FOLDER is set
            from config import DROPBOX_ROOT_FOLDER
            if DROPBOX_ROOT_FOLDER and not self.dropbox_folder_var.get():
                self.dropbox_folder_var.set(DROPBOX_ROOT_FOLDER)

            # Populate folder dropdown
            self._refresh_dropbox_folders()

        except Exception as e:
            self.dropbox_status_var.set("❌ error")
            self.log(f"   ❌ Dropbox connection failed: {e}")
            messagebox.showerror("Dropbox Error", str(e))

    def _refresh_dropbox_folders(self):
        """Populate the folder combobox with sub-folders from the Dropbox root."""
        if not self.dropbox_service or not self.dropbox_service.is_connected:
            return
        try:
            folders = self.dropbox_service.list_folders()
            if hasattr(self, 'dropbox_folder_combo'):
                self.dropbox_folder_combo['values'] = folders
        except Exception as e:
            self.log(f"   ⚠️ Could not list Dropbox folders: {e}")

    def _list_dropbox_files(self):
        """List supported files in the chosen Dropbox folder, with optional date filter."""
        if not hasattr(self, 'dropbox_tree'):
            return  # tree not present in current layout
        if not self.dropbox_service or not self.dropbox_service.is_connected:
            messagebox.showwarning("Warning", "Connect to Dropbox first.")
            return

        folder = self.dropbox_folder_var.get().strip()
        self.log(f"☁️  Listing files in Dropbox: {folder or '/'}")

        # Parse optional date filter
        filter_date = None
        date_str = self.dropbox_date_filter_var.get().strip()
        if date_str:
            try:
                from datetime import datetime
                filter_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                self.log(f"   📅 Filtering: uploaded on or after {date_str}")
            except ValueError:
                messagebox.showwarning("Invalid Date", f"Date format must be YYYY-MM-DD (e.g. 2026-03-20).\nGot: {date_str}")
                return

        try:
            all_files = self.dropbox_service.list_supported_files(folder)

            # Apply date filter if set
            if filter_date:
                filtered = []
                for meta in all_files:
                    mod = getattr(meta, "server_modified", None) or getattr(meta, "client_modified", None)
                    if mod and mod.date() >= filter_date:
                        filtered.append(meta)
                skipped = len(all_files) - len(filtered)
                if skipped:
                    self.log(f"   ⏭️  Skipped {skipped} files older than {date_str}")
                all_files = filtered

            # Apply keyword filter if set
            keyword = self.dropbox_keyword_var.get().strip()
            if keyword:
                before_kw = len(all_files)
                all_files = [m for m in all_files if keyword.lower() in m.name.lower()]
                skipped_kw = before_kw - len(all_files)
                if skipped_kw:
                    self.log(f"   ⏭️  Skipped {skipped_kw} files not matching '{keyword}'")

            # Apply patient name list filter if set
            if self.dropbox_name_filter_list:
                before_name = len(all_files)
                all_files = [
                    m for m in all_files
                    if any(n.lower() in m.name.lower() for n in self.dropbox_name_filter_list)
                ]
                skipped_name = before_name - len(all_files)
                if skipped_name:
                    self.log(f"   ⏭️  Skipped {skipped_name} files not matching any of {len(self.dropbox_name_filter_list)} names")

            self.dropbox_files = all_files

            # Populate treeview
            for row in self.dropbox_tree.get_children():
                self.dropbox_tree.delete(row)

            for meta in self.dropbox_files:
                size_str = self._fmt_bytes(meta.size)
                mod = getattr(meta, "server_modified", None) or getattr(meta, "client_modified", None)
                mod_str = mod.strftime("%Y-%m-%d") if mod else ""
                self.dropbox_tree.insert("", tk.END,
                                         values=(meta.name, meta.path_display, mod_str, size_str))

            count = len(self.dropbox_files)
            status_parts = [f"✅ {count} files"]
            if filter_date:
                status_parts.append(f"from {date_str}")
            if keyword:
                status_parts.append(f"'{keyword}'")
            if self.dropbox_name_filter_list:
                status_parts.append(f"{len(self.dropbox_name_filter_list)} names")
            self.dropbox_status_var.set(" | ".join(status_parts))
            self.dropbox_status_var.set(" | ".join(status_parts))
            self.log(f"   Found {count} supported files")

            # Auto-set output file if not yet set
            if not self.output_file.get():
                default_output = pathlib.Path.cwd() / "Dropbox_Auth_Results.xlsx"
                self.output_file.set(str(default_output))

        except Exception as e:
            self.log(f"   ❌ Error listing Dropbox files: {e}")
            messagebox.showerror("Dropbox Error", str(e))

    def _clear_name_filter(self):
        """Clear the patient name filter list."""
        self.dropbox_name_filter_list = []
        self.dropbox_name_count_var.set("(none)")

    def _open_name_filter_popup(self):
        """Open a popup to manage the patient name filter list."""
        popup = tk.Toplevel(self.root)
        popup.title("Filter by Patient Names")
        popup.geometry("480x500")
        popup.transient(self.root)
        popup.grab_set()
        self.style_popup(popup)

        popup.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - 240
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - 250
        popup.geometry(f"+{x}+{y}")

        frame = ttk.Frame(popup, padding="15")
        frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(frame, text="Patient Name Filter List",
                  font=("Segoe UI", 13, "bold")).pack(anchor=tk.W, pady=(0, 4))
        ttk.Label(frame,
                  text="Only files containing at least one name below will be listed.\n"
                       "One name per line. Partial / case-insensitive match.",
                  font=("Segoe UI", 9), foreground="gray").pack(anchor=tk.W, pady=(0, 10))

        # Text area
        txt_frame = ttk.Frame(frame)
        txt_frame.pack(fill=tk.BOTH, expand=True)
        txt = tk.Text(txt_frame, font=("Consolas", 10), height=18,
                      bg="#1E293B", fg="#E5E7EB", insertbackground="#E5E7EB",
                      relief="solid", borderwidth=1, padx=6, pady=6)
        sb = ttk.Scrollbar(txt_frame, orient=tk.VERTICAL, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        txt.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        # Pre-fill with existing names
        if self.dropbox_name_filter_list:
            txt.insert("1.0", "\n".join(self.dropbox_name_filter_list))

        # Row count label
        count_lbl = ttk.Label(frame, text="", font=("Segoe UI", 8), foreground="gray")
        count_lbl.pack(anchor=tk.W, pady=(4, 0))

        def _update_count(event=None):
            names = [n.strip() for n in txt.get("1.0", tk.END).split("\n") if n.strip()]
            count_lbl.config(text=f"{len(names)} name(s) entered")

        txt.bind("<KeyRelease>", _update_count)
        txt.bind("<<Paste>>", lambda e: popup.after(50, _update_count))
        _update_count()

        def apply():
            raw = txt.get("1.0", tk.END)
            names = [n.strip() for n in raw.split("\n") if n.strip()]
            self.dropbox_name_filter_list = names
            if names:
                self.dropbox_name_count_var.set(f"{len(names)} name(s) active")
            else:
                self.dropbox_name_count_var.set("(none)")
            popup.destroy()

        btn_row = ttk.Frame(frame)
        btn_row.pack(fill=tk.X, pady=(12, 0))
        ttk.Button(btn_row, text="✅ Apply", command=apply,
                   style="Action.TButton").pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="🗑️ Clear All",
                   command=lambda: txt.delete("1.0", tk.END)).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="Cancel", command=popup.destroy).pack(side=tk.LEFT)

        txt.focus_set()

    def _log_dropbox_download(self, current, total, filename):
        """Log callback for Dropbox download progress."""
        self.status_text.set(f"Downloading {current}/{total}: {filename}")
        self.log(f"   ⬇️  {current}/{total} {filename}")
        pct = (current / total) * 30  # first 30% of progress bar = download
        self.progress_bar["value"] = pct
        self.root.update_idletasks()

    @staticmethod
    def _fmt_bytes(n):
        """Format bytes as human-readable string."""
        for unit in ("B", "KB", "MB", "GB"):
            if abs(n) < 1024:
                return f"{n:.0f} {unit}"
            n /= 1024
        return f"{n:.1f} TB"

    def start_extraction(self):
        """Start the extraction process."""
        if self.is_processing:
            return
        
        # Determine input folder based on source type
        source_type = self.extract_source_type.get()
        
        if source_type == "finder":
            # Use File Finder destination folder
            input_folder = self.finder_dest.get()
            if not input_folder:
                messagebox.showerror("Error", "No File Finder destination set.\n\nRun File Finder first to find and copy PDFs,\nor select 'Select folder manually' option.")
                return
        elif source_type == "dropbox":
            # Download Dropbox files to a temp folder, then extract from there
            if not self.dropbox_service or not self.dropbox_service.is_connected:
                messagebox.showerror("Error", "Not connected to Dropbox.\nClick 'Connect' first.")
                return
            if not self.dropbox_files:
                messagebox.showerror("Error", "No Dropbox files listed.\nClick 'List Files' first.")
                return
            # Download will happen in the extraction thread
            input_folder = "__dropbox__"  # sentinel value
        else:
            # Use manual folder selection
            input_folder = self.input_folder.get()
            if not input_folder:
                messagebox.showerror("Error", "Please select an input folder")
                return
        
        output_file = self.output_file.get()
        
        if not output_file:
            messagebox.showerror("Error", "Please specify an output file")
            return
        
        if input_folder != "__dropbox__" and not pathlib.Path(input_folder).exists():
            messagebox.showerror("Error", "Input folder does not exist")
            return
        
        # Clear log
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        
        self.is_processing = True
        self.extract_btn.config(state=tk.DISABLED)
        self.progress_bar["value"] = 0
        
        # Run extraction in thread
        thread = Thread(target=self.run_extraction, args=(input_folder, output_file))
        thread.start()
    
    def run_extraction(self, input_folder, output_file):
        """Run extraction process (in separate thread)."""
        try:
            # --- Dropbox download step ---
            if input_folder == "__dropbox__":
                self.log("☁️  Downloading files from Dropbox...")
                try:
                    downloaded = self.dropbox_service.download_files(
                        self.dropbox_files,
                        progress_callback=lambda c, t, f: self.root.after(
                            0, lambda _c=c, _t=t, _f=f: self._log_dropbox_download(_c, _t, _f)),
                    )
                    if not downloaded:
                        self.root.after(0, lambda: messagebox.showwarning("Warning", "No files downloaded from Dropbox"))
                        return
                    # Use the temp directory that contains downloaded files
                    input_folder = self.dropbox_service._get_temp_dir()
                    self.log(f"✅ Downloaded {len(downloaded)} files to temp folder")
                except Exception as e:
                    err_msg = str(e)
                    self.log(f"❌ Dropbox download failed: {err_msg}")
                    self.root.after(0, lambda m=err_msg: messagebox.showerror("Dropbox Error", m))
                    return

            self.log(f"📁 Scanning folder: {input_folder}")
            
            # Count PDF files
            folder = pathlib.Path(input_folder)
            pdf_files = sorted(folder.glob("*.pdf"))
            total_files = len(pdf_files)

            if total_files == 0:
                self.root.after(0, lambda: messagebox.showwarning("Warning", "No PDF files found in folder"))
                return

            self.log(f"   PDF files: {total_files}")

            # --- Unlock PDFs ---
            # Decrypts password-protected PDFs so they can be opened/uploaded freely
            self.log("\n🔓 Unlocking PDFs...")
            from services.pdf_unlock_service import PdfUnlockService
            import shutil as _shutil

            unlock_service = PdfUnlockService()
            # Save unlocked PDFs to a "Unlocked" subfolder next to the output file
            output_folder = str(pathlib.Path(output_file).parent / "Unlocked PDFs")
            os.makedirs(output_folder, exist_ok=True)

            successful = 0
            errors = 0
            for i, pdf_file in enumerate(pdf_files):
                self.root.after(0, lambda c=i+1, t=total_files, f=pdf_file.name: self.update_progress(c, t, f))
                try:
                    unlocked_path, was_encrypted = unlock_service.unlock(str(pdf_file))
                    dest_path = os.path.join(output_folder, pdf_file.name)
                    if was_encrypted:
                        _shutil.move(unlocked_path, dest_path)
                        self.log(f"  🔓 Unlocked: {pdf_file.name}")
                    else:
                        _shutil.copy2(str(pdf_file), dest_path)
                        self.log(f"  ✓ Copied (not encrypted): {pdf_file.name}")
                    successful += 1
                except Exception as ex:
                    self.log(f"  ❌ Failed: {pdf_file.name} - {ex}")
                    errors += 1

            self.log(f"\n✅ Complete!")
            self.log(f"   PDFs saved: {successful}")
            self.log(f"   Errors: {errors}")
            self.log(f"   Output folder: {output_folder}")
            
            self.root.after(0, lambda: self.status_text.set(f"Done! {successful} PDFs unlocked"))
            self.root.after(0, lambda f=output_folder, s=successful, e=errors, t=total_files: messagebox.showinfo("Success", 
                f"Download complete!\n\nProcessed: {t} PDFs\nSaved: {s}\nErrors: {e}\n\nFolder:\n{f}"))
            
            # Open the output folder in Explorer
            self.root.after(500, lambda f=output_folder: os.startfile(f) if os.path.isdir(f) else None)
            
        except Exception as e:
            err_msg = str(e)
            self.log(f"\n❌ Error: {err_msg}")
            self.root.after(0, lambda m=err_msg: messagebox.showerror("Error", m))
        finally:
            self.is_processing = False
            self.root.after(0, lambda: self.extract_btn.config(state=tk.NORMAL))
    
    def open_output(self):
        """Open the output Excel file."""
        output_file = self.output_file.get()
        if output_file and pathlib.Path(output_file).exists():
            os.startfile(output_file)
    
    def show_caspio_upload_dialog(self, valid_only=False):
        """Show dialog to map fields and upload to Caspio.
        
        Args:
            valid_only: If True, only upload non-expired auths.
        """
        if not REQUESTS_AVAILABLE:
            messagebox.showerror("Error", "requests library is required.\nRun: pip install requests")
            return
        
        self.upload_data_source = None  # Will hold DataFrame to upload
        
        # Check if we have in-app data from the results table
        in_app_data = self.get_results_from_table(valid_only=valid_only)
        all_data = self.get_results_from_table(valid_only=False)
        
        has_in_app_data = in_app_data is not None and len(in_app_data) > 0
        
        # Count for display
        total_count = len(all_data) if all_data is not None else 0
        valid_count = len(in_app_data) if in_app_data is not None else 0
        expired_count = total_count - valid_count
        
        if has_in_app_data:
            # Different message based on valid_only mode
            if valid_only:
                msg = (f"Uploading VALID AUTHS ONLY from results table.\n\n"
                       f"Total records: {total_count}\n"
                       f"Valid (non-expired): {valid_count}\n"
                       f"Expired (excluded): {expired_count}\n\n"
                       "YES = Upload {valid_count} valid records\n"
                       "NO = Load from Excel file instead\n"
                       "CANCEL = Cancel upload")
            else:
                msg = (f"You have {total_count} records in the results table.\n\n"
                       "YES = Upload ALL data from results table\n"
                       "NO = Load from Excel file instead\n"
                       "CANCEL = Cancel upload")
            
            choice = messagebox.askyesnocancel("Data Source", msg)
            
            if choice is None:  # Cancel
                return
            elif choice:  # Yes - use in-app data
                self.upload_data_source = in_app_data
                if valid_only:
                    self.log(f"\n📋 Using {len(self.upload_data_source)} valid records (excluding {expired_count} expired)")
                else:
                    self.log(f"\n📋 Using {len(self.upload_data_source)} records from results table")
            # else: choice is False, fall through to load from file
        
        # If no in-app data selected, load from Excel file
        if self.upload_data_source is None:
            # Default to Auth_Results_Combined.xlsx in workspace
            default_file = pathlib.Path(self.input_folder.get()).parent / "Auth_Results_Combined.xlsx"
            if not default_file.exists():
                default_file = pathlib.Path(self.input_folder.get()) / "Auth_Results_Combined.xlsx"
            if not default_file.exists():
                default_file = pathlib.Path.cwd() / "Auth_Results_Combined.xlsx"
            
            # Also check for the output file path if set
            if self.output_file.get() and pathlib.Path(self.output_file.get()).exists():
                default_file = pathlib.Path(self.output_file.get())
            
            file_path = filedialog.askopenfilename(
                title="Select Excel file to upload",
                initialdir=default_file.parent if default_file.exists() else pathlib.Path.cwd(),
                initialfile=default_file.name if default_file.exists() else "Auth_Results_Combined.xlsx",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if not file_path:
                return
            
            try:
                # Load the Formatted sheet from the Excel file
                xl = pd.ExcelFile(file_path)
                if "Formatted" in xl.sheet_names:
                    self.upload_data_source = pd.read_excel(file_path, sheet_name="Formatted")
                    self.log(f"\n📂 Loaded {len(self.upload_data_source)} records from '{file_path}' (Formatted sheet)")
                else:
                    # Show available sheets and let user pick
                    available_sheets = xl.sheet_names
                    messagebox.showwarning("Sheet Not Found", 
                        f"'Formatted' sheet not found in file.\n\nAvailable sheets: {', '.join(available_sheets)}\n\nLoading first sheet: {available_sheets[0]}")
                    self.upload_data_source = pd.read_excel(file_path, sheet_name=0)
                    self.log(f"\n📂 Loaded {len(self.upload_data_source)} records from '{file_path}' (first sheet)")
            except Exception as e:
                messagebox.showerror("Error", f"Failed to load Excel file:\n{str(e)}")
                return
        
        # Create popup window
        popup = tk.Toplevel(self.root)
        popup.title("Upload to Caspio")
        popup.geometry("700x550")
        popup.transient(self.root)
        popup.grab_set()
        self.style_popup(popup)
        
        # Center the popup
        popup.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (350)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (275)
        popup.geometry(f"+{x}+{y}")
        
        main_frame = ttk.Frame(popup, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        ttk.Label(main_frame, text="Map Fields to Caspio Table", 
                  font=("Segoe UI", 14, "bold")).pack(pady=(0, 5))
        ttk.Label(main_frame, text=f"Table: {CASPIO_TABLE_NAME}", 
                  font=("Segoe UI", 10)).pack(pady=(0, 10))
        
        # Status label
        status_var = tk.StringVar(value="Loading Caspio table schema...")
        status_label = ttk.Label(main_frame, textvariable=status_var, foreground="blue")
        status_label.pack(pady=(0, 10))
        
        # Mapping frame (will be populated after loading schema)
        mapping_frame = ttk.LabelFrame(main_frame, text="Field Mapping", padding="10")
        mapping_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Create scrollable frame for mappings
        canvas = tk.Canvas(mapping_frame)
        scrollbar = ttk.Scrollbar(mapping_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Local fields - use columns from the loaded Excel file
        if self.upload_data_source is not None:
            local_fields = [col for col in self.upload_data_source.columns if col not in ["Extracted_Name"]]
        else:
            # Fallback default fields
            local_fields = [
                "Last Name", "First Name", "Patient Name", "Patient ID",
                "Service Type Identifier", "CPT Code", "CPT Code 2", "CPT Code 3", 
                "CPT Code 4", "CPT Code 5", "Auth Number", "Date Approved", 
                "Date Auth Expired", "Clearing House Payer ID", "Location ID", 
                "Unique Payer Identifier"
            ]
        
        # Store mapping dropdowns
        mapping_vars = {}
        
        def load_caspio_schema():
            """Load Caspio table schema in background."""
            try:
                caspio = CaspioAPI()
                caspio_fields = caspio.get_table_schema(CASPIO_TABLE_NAME)
                caspio_field_names = ["-- Do not upload --"] + [f["name"] for f in caspio_fields]
                
                # Update UI in main thread
                popup.after(0, lambda: populate_mappings(caspio_field_names, caspio_fields))
                popup.after(0, lambda: status_var.set(f"✅ Connected! Found {len(caspio_fields)} fields in Caspio table."))
                popup.after(0, lambda: status_label.config(foreground="green"))
            except Exception as e:
                popup.after(0, lambda: status_var.set(f"❌ Error: {str(e)}"))
                popup.after(0, lambda: status_label.config(foreground="red"))
        
        def populate_mappings(caspio_field_names, caspio_fields):
            """Populate the mapping dropdowns."""
            # Header row
            ttk.Label(scrollable_frame, text="Local Field", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, padx=10, pady=5, sticky="w")
            ttk.Label(scrollable_frame, text="→", font=("Segoe UI", 10)).grid(row=0, column=1, padx=5, pady=5)
            ttk.Label(scrollable_frame, text="Caspio Field", font=("Segoe UI", 10, "bold")).grid(row=0, column=2, padx=10, pady=5, sticky="w")
            
            # Auto-match function
            def auto_match(local_name):
                """Try to find a matching Caspio field name."""
                local_normalized = local_name.lower().replace(" ", "_").replace("-", "_")
                for cf in caspio_field_names[1:]:  # Skip "-- Do not upload --"
                    cf_normalized = cf.lower()
                    # Exact match
                    if local_normalized == cf_normalized:
                        return cf
                    # Common variations
                    if local_normalized.replace("_", "") == cf_normalized.replace("_", ""):
                        return cf
                    # Partial matches for common fields
                    if "last" in local_normalized and "last" in cf_normalized and "name" in cf_normalized:
                        return cf
                    if "first" in local_normalized and "first" in cf_normalized and "name" in cf_normalized:
                        return cf
                    if "patient" in local_normalized and "name" in local_normalized and "patient" in cf_normalized and "name" in cf_normalized:
                        return cf
                    if "patient" in local_normalized and "id" in local_normalized and "patient" in cf_normalized and "id" in cf_normalized:
                        return cf
                    if "auth" in local_normalized and "number" in local_normalized and "auth" in cf_normalized:
                        return cf
                    if "cpt" in local_normalized and "cpt" in cf_normalized:
                        # Try to match CPT Code 2 with CPT_Code_2, etc.
                        if local_normalized == cf_normalized:
                            return cf
                    if "date" in local_normalized and "approved" in local_normalized and "approved" in cf_normalized:
                        return cf
                    if "expire" in local_normalized and "expire" in cf_normalized:
                        return cf
                    if "clearing" in local_normalized and "clearing" in cf_normalized:
                        return cf
                    if "location" in local_normalized and "id" in local_normalized and "location" in cf_normalized:
                        return cf
                    if "payer" in local_normalized and "identifier" in local_normalized and "payer" in cf_normalized:
                        return cf
                return "-- Do not upload --"
            
            # Create mapping rows
            for i, local_field in enumerate(local_fields):
                ttk.Label(scrollable_frame, text=local_field).grid(row=i+1, column=0, padx=10, pady=3, sticky="w")
                ttk.Label(scrollable_frame, text="→").grid(row=i+1, column=1, padx=5, pady=3)
                
                var = tk.StringVar()
                combo = ttk.Combobox(scrollable_frame, textvariable=var, values=caspio_field_names, 
                                     width=30, state="readonly")
                combo.grid(row=i+1, column=2, padx=10, pady=3, sticky="w")
                
                # Try to auto-match
                matched = auto_match(local_field)
                var.set(matched)
                
                mapping_vars[local_field] = var
            
            # Enable upload button
            upload_btn.config(state=tk.NORMAL)
        
        def do_upload():
            """Perform the upload to Caspio."""
            # Build field mapping (local -> caspio)
            field_map = {}
            for local_field, var in mapping_vars.items():
                caspio_field = var.get()
                if caspio_field and caspio_field != "-- Do not upload --":
                    field_map[local_field] = caspio_field
            
            if not field_map:
                messagebox.showwarning("Warning", "No fields mapped. Please map at least one field.", parent=popup)
                return
            
            # Get data from the loaded Excel file (always re-read to capture edits)
            if self.upload_data_source is None:
                messagebox.showwarning("Warning", "No data loaded. Please select an Excel file.", parent=popup)
                return
            
            # Convert DataFrame to list of dicts
            formatted_data = self.upload_data_source.to_dict('records')
            
            if not formatted_data:
                messagebox.showwarning("Warning", "No data to upload.", parent=popup)
                return
            
            # Transform data to Caspio format
            caspio_records = []
            for row in formatted_data:
                record = {}
                for local_field, caspio_field in field_map.items():
                    value = row.get(local_field, "")
                    
                    # Check for NaN/None/NaT values FIRST (pandas NaN/NaT breaks JSON)
                    try:
                        is_null = value is None
                        if not is_null and pd:
                            is_null = pd.isna(value)
                        if not is_null and isinstance(value, float):
                            import math
                            is_null = math.isnan(value) or math.isinf(value)
                    except (TypeError, ValueError):
                        is_null = False
                    
                    if is_null:
                        value = ""
                    # Handle date fields - convert to string format Caspio expects
                    elif value and ("Date" in local_field or "date" in local_field or "DOS" in local_field):
                        if hasattr(value, 'strftime'):
                            try:
                                value = value.strftime("%m/%d/%Y")
                            except:
                                value = ""
                        elif pd:
                            try:
                                if pd.notna(value):
                                    value = pd.to_datetime(value).strftime("%m/%d/%Y")
                                else:
                                    value = ""
                            except:
                                value = str(value) if value else ""
                    # Convert any remaining non-string values to strings
                    if value and not isinstance(value, str):
                        try:
                            value = str(value)
                        except:
                            value = ""
                    record[caspio_field] = value if value else ""
                caspio_records.append(record)
            
            # Confirm upload
            if not messagebox.askyesno("Confirm Upload", 
                f"Upload {len(caspio_records)} records to Caspio table '{CASPIO_TABLE_NAME}'?\n\nMapped fields:\n" + 
                "\n".join([f"  {k} → {v}" for k, v in field_map.items()]),
                parent=popup):
                return
            
            # Disable button and update status
            upload_btn.config(state=tk.DISABLED)
            status_var.set("Uploading to Caspio...")
            status_label.config(foreground="blue")
            popup.update()
            
            # Perform upload
            try:
                caspio = CaspioAPI()
                results = caspio.insert_records(CASPIO_TABLE_NAME, caspio_records)
                
                # Audit log the upload
                try:
                    from audit.logger import AuditLogger
                    audit = AuditLogger()
                    audit.log_upload(
                        source_file=f"batch_{len(caspio_records)}_records",
                        table_name=CASPIO_TABLE_NAME,
                        record_count=results['success'],
                        status="success" if results['failed'] == 0 else "partial",
                        error="; ".join(results.get('errors', [])[:3]) if results['failed'] > 0 else "",
                    )
                except Exception:
                    pass  # audit failure should not block upload feedback
                
                status_var.set(f"✅ Upload complete! Success: {results['success']}, Failed: {results['failed']}")
                status_label.config(foreground="green")
                
                if results['failed'] > 0:
                    error_msg = f"Upload completed with errors.\n\nSuccess: {results['success']}\nFailed: {results['failed']}\n\n"
                    if results['errors']:
                        error_msg += "First few errors:\n" + "\n".join(results['errors'][:5])
                    messagebox.showwarning("Upload Completed with Errors", error_msg, parent=popup)
                else:
                    messagebox.showinfo("Success", 
                        f"Successfully uploaded {results['success']} records to Caspio!", parent=popup)
                    popup.destroy()
                    self.log(f"\n☁️ Uploaded {results['success']} records to Caspio table '{CASPIO_TABLE_NAME}'")
                    
            except Exception as e:
                status_var.set(f"❌ Error: {str(e)}")
                status_label.config(foreground="red")
                messagebox.showerror("Upload Error", str(e), parent=popup)
            finally:
                upload_btn.config(state=tk.NORMAL)
        
        # Buttons frame
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        upload_btn = ttk.Button(btn_frame, text="⬆️ Upload to Caspio", command=do_upload, state=tk.DISABLED)
        upload_btn.pack(side=tk.LEFT, padx=5)
        
        ttk.Button(btn_frame, text="Cancel", command=popup.destroy).pack(side=tk.LEFT, padx=5)
        
        # Record count - use Excel file data if loaded, otherwise use in-memory results
        if self.upload_data_source is not None:
            formatted_count = len(self.upload_data_source)
            source_label = " (from Excel file)"
        else:
            formatted_count = len(self.extractor.format_results())
            source_label = ""
        ttk.Label(btn_frame, text=f"Records to upload: {formatted_count}{source_label}", 
                  font=("Segoe UI", 9), foreground="gray").pack(side=tk.RIGHT, padx=10)
        
        # Load schema in background thread
        Thread(target=load_caspio_schema, daemon=True).start()


def check_dependencies():
    """Check if required dependencies are installed."""
    missing = []
    
    if not pdfplumber and not PyPDF2:
        missing.append("pdfplumber")
    if not pd:
        missing.append("pandas openpyxl")
    
    return missing


def main():
    """Main entry point."""
    # Check dependencies
    missing = check_dependencies()
    if missing:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Missing Dependencies", 
            f"Please install required packages:\n\npip install {' '.join(missing)}")
        sys.exit(1)
    
    # Create root window hidden during splash
    root = tk.Tk()
    root.withdraw()

    # Set logo as app icon for ALL windows (taskbar, title bar, alt-tab)
    _icon_photo = None  # keep ref to prevent GC
    _icon_path = APP_DIR / "Auth Radar Logo.png"
    if _icon_path.exists():
        try:
            from PIL import Image, ImageTk
            _raw = Image.open(str(_icon_path)).resize((64, 64), Image.LANCZOS)
            _icon_photo = ImageTk.PhotoImage(_raw)
            root.iconphoto(True, _icon_photo)  # True = apply to all Toplevel windows
        except Exception:
            pass

    _app_ref = []  # keep LandingPage/AuthExtractorApp alive (prevents logo GC)

    def _launch_landing():
        _app_ref.append(LandingPage(root))
        root.deiconify()

    # Show splash screen, then launch landing page
    splash_win = tk.Toplevel(root)
    SplashScreen(splash_win, _launch_landing)
    root.mainloop()


if __name__ == "__main__":
    main()
