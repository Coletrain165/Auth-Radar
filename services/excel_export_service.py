"""
Auth Radar - Excel Export Service

Exports extraction results to a formatted Excel file with:
- Auto-filter headers
- Frozen header row
- Readable column widths
- One row per PDF
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column definitions for the output Excel table
COLUMNS = [
    "Source File",
    "Auth #",
    "Date Approved",
    "Date Auth. Expired",
    "Participant's Name",
    "Participant ID",
    "Source Page",
    "Auth # Confidence",
    "Date Approved Confidence",
    "Date Auth. Expired Confidence",
    "Participant Name Confidence",
    "Participant ID Confidence",
    "Notes",
]


class ExcelExportService:
    """Exports extraction results to a formatted Excel table."""

    def export(self, results: list[dict], output_path: str) -> str:
        """
        Write extraction results to an Excel file.

        Args:
            results: List of extraction result dicts (one per PDF).
                Each dict should have the keys from the OpenAI extraction schema
                plus a 'source_file' key.
            output_path: Path to the output .xlsx file.

        Returns:
            The absolute path to the written file.
        """
        rows = []
        for r in results:
            confidence = r.get("confidence", {})
            rows.append({
                "Source File": r.get("source_file", ""),
                "Auth #": r.get("auth_number") or "",
                "Date Approved": r.get("date_approved") or "",
                "Date Auth. Expired": r.get("date_auth_expired") or "",
                "Participant's Name": r.get("participant_name") or "",
                "Participant ID": r.get("participant_id") or "",
                "Source Page": r.get("source_page", ""),
                "Auth # Confidence": confidence.get("auth_number", ""),
                "Date Approved Confidence": confidence.get("date_approved", ""),
                "Date Auth. Expired Confidence": confidence.get("date_auth_expired", ""),
                "Participant Name Confidence": confidence.get("participant_name", ""),
                "Participant ID Confidence": confidence.get("participant_id", ""),
                "Notes": r.get("notes", ""),
            })

        df = pd.DataFrame(rows, columns=COLUMNS)

        # Write to Excel with formatting
        output_path = str(Path(output_path).resolve())
        self._write_formatted_excel(df, output_path)

        logger.info("Excel output written: %s (%d rows)", output_path, len(rows))
        return output_path

    def _write_formatted_excel(self, df: pd.DataFrame, output_path: str):
        """Write a DataFrame to Excel with table formatting."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Auth Extractions"

        # Write header and data
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Freeze the header row
        ws.freeze_panes = "A2"

        # Create an Excel Table with autofilter
        if len(df) > 0:
            last_col = get_column_letter(len(COLUMNS))
            last_row = len(df) + 1  # +1 for header
            table_ref = f"A1:{last_col}{last_row}"

            table = Table(displayName="AuthExtractions", ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)

        # Set reasonable column widths
        col_widths = {
            "Source File": 35,
            "Auth #": 15,
            "Date Approved": 14,
            "Date Auth. Expired": 16,
            "Participant's Name": 25,
            "Participant ID": 15,
            "Source Page": 12,
            "Auth # Confidence": 16,
            "Date Approved Confidence": 22,
            "Date Auth. Expired Confidence": 24,
            "Participant Name Confidence": 24,
            "Participant ID Confidence": 20,
            "Notes": 40,
        }

        for col_idx, col_name in enumerate(COLUMNS, 1):
            width = col_widths.get(col_name, 15)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Format confidence columns as percentages (0-1 range display)
        from openpyxl.styles import numbers
        confidence_cols = [8, 9, 10, 11, 12]  # 1-indexed columns for confidence
        for col_idx in confidence_cols:
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

        wb.save(output_path)
